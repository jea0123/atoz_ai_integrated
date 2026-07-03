from __future__ import annotations

from copy import copy as copy_object
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
import posixpath
import re
import shutil
import time
import zipfile
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape
from uuid import uuid4

from document_update.document_number import write_updated_document, write_updated_project_title
from document_update.hwpx_text import is_hwpx_zip
from document_update.metadata_update import update_metadata_in_document
from document_update.runtime_conversion import prepare_target_file
from output_file_check.content_identity import read_file_identity, read_standard_project_title
from output_file_check.file_noise import is_noise_filename
from output_file_check.models import StandardOutput
from output_file_check.normalization import filesystem_safe_stem, normalize_for_match
from output_file_check.requirement_generation import (
    RequirementSource,
    create_requirement_source_folders,
    extract_requirement_ids,
)
from output_file_check.standard_reader import extract_standard_text, read_standard_outputs
from web_uploads import (
    ARTIFACT_UPLOAD_SUPPORTED_SUFFIXES,
    IGNORED_UPLOAD_FOLDER_KEYS,
    save_proposal_requirement_uploads,
    save_requirement_uploads,
    safe_relative_upload_path,
    safe_upload_filename,
)


TEMPLATE_ROOT = Path("templates")
TEMPLATE_CATEGORY_DIRS = {
    "management": TEMPLATE_ROOT / "관리산출물",
    "development": TEMPLATE_ROOT / "개발산출물",
}
COVER_TEMPLATE_DIR = TEMPLATE_ROOT / "표지"
DEFAULT_DOCUMENT_VERSION = "v0.1"
STANDARD_SUFFIXES = {".pdf", ".hwp", ".hwpx"}
TOKEN_PATTERN = re.compile(r"\{\{\s*([A-Z0-9_]+)\s*\}\}")
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
SUPPORTED_TEMPLATE_SUFFIXES = ARTIFACT_UPLOAD_SUPPORTED_SUFFIXES
COVER_SUFFIXES_BY_TYPE = {
    "ppt": {".pptx", ".pptm"},
    "excel": {".xlsx", ".xlsm", ".xltx", ".xltm"},
    "hwp": {".hwp", ".hwpx"},
}


@dataclass(frozen=True)
class BodyTemplate:
    path: Path
    relative_path: str
    suffix: str
    name_key: str
    relative_key: str
    stem_key: str
    cover_type: str


@dataclass(frozen=True)
class TemplateBuildInput:
    path: Path
    relative_path: str
    standard_output: StandardOutput | None
    standard_match_type: str
    cover_type: str
    cover_template: Path | None
    body_template: BodyTemplate | None
    body_match_type: str
    output_relative_path: str
    document_title: str
    document_number: str
    project_title: str
    input_has_cover: bool
    token_values: dict[str, str]
    status: str
    message: str


@dataclass(frozen=True)
class TemplateOutputVariant:
    requirement_id: str
    output_relative_path: str
    token_values: dict[str, str]


def run_template_build(
    fields: dict[str, str],
    file_items: dict[str, list[tuple[str, bytes]]],
    *,
    temp_dir: Path,
    result_dir: Path,
) -> dict[str, object]:
    category = parse_template_category(fields.get("artifact_category"))
    apply_mode = parse_bool(fields.get("apply_mode")) or clean_token_value(fields.get("template_mode")).lower() == "apply"
    standard_path = save_standard_file(temp_dir, file_items)
    standard_text = extract_standard_text(standard_path)
    standard_outputs = read_standard_outputs(standard_path, standard_text, category=category)
    standard_project_title = read_standard_project_title(standard_path, standard_text)
    source_dir = save_artifact_source(temp_dir, file_items, category)
    body_templates = load_body_templates(TEMPLATE_CATEGORY_DIRS[category])
    requirement_paths = save_requirement_sources_for_category(temp_dir, file_items, category)
    requirement_sources = build_requirement_sources(requirement_paths)
    requirement_ids = collect_requirement_ids(requirement_sources)
    token_values = build_cover_token_values(fields, requirement_ids, standard_project_title)

    output_root = result_dir / "template-builds" / f"{category}-{uuid4().hex[:10]}"
    output_root.mkdir(parents=True, exist_ok=True)

    inputs = scan_artifact_inputs(source_dir, standard_outputs, body_templates, token_values)
    copied_items = (
        copy_template_outputs(inputs, output_root, work_root=temp_dir / "template-output-work")
        if apply_mode
        else []
    )
    requirement_folder_items = (
        create_requirement_source_folders(output_root / "outputs", requirement_sources)
        if apply_mode and category == "development"
        else []
    )
    report_path = None
    if apply_mode:
        report_path = write_template_build_report(
            output_root,
            category=category,
            standard_path=standard_path,
            standard_output_count=len(standard_outputs),
            source_dir=source_dir,
            body_template_dir=TEMPLATE_CATEGORY_DIRS[category],
            body_template_count=len(body_templates),
            requirement_ids=requirement_ids,
            requirement_sources=requirement_paths,
            token_values=token_values,
            items=inputs,
        )

    files = [
        {
            "name": item["name"],
            "path": str(item["path"]),
            "kind": "template_output",
            "requirement_id": item.get("requirement_id", ""),
            "source_mode": item.get("source_mode", ""),
        }
        for item in copied_items
    ]
    if report_path is not None:
        files.append({"name": report_path.name, "path": str(report_path), "kind": "report"})

    return {
        "ok": True,
        "applied": apply_mode,
        "artifact_category": category,
        "standard_file": str(standard_path),
        "standard_project_title": standard_project_title,
        "standard_output_count": len(standard_outputs),
        "source_dir": str(source_dir),
        "body_template_dir": str(TEMPLATE_CATEGORY_DIRS[category]),
        "dump_root": str(output_root),
        "artifact_count": len(inputs),
        "template_count": len(body_templates),
        "output_count": len(copied_items),
        "matched_count": sum(1 for item in inputs if item.status == "matched"),
        "requirement_count": len(requirement_ids),
        "requirement_ids": requirement_ids,
        "requirement_generated_folder_count": sum(
            1 for item in requirement_folder_items if item.get("status") != "error"
        ),
        "requirement_generation_created_folder_count": sum(
            1 for item in requirement_folder_items if item.get("status") == "created"
        ),
        "requirement_generation_folder_items": requirement_folder_items,
        "token_values": token_values,
        "items": [serialize_template_item(item, output_root, applied=apply_mode) for item in inputs],
        "files": files,
    }


def parse_template_category(value: str | None) -> str:
    category = (value or "development").strip().lower()
    if category not in TEMPLATE_CATEGORY_DIRS:
        raise ValueError("템플릿 종류는 management 또는 development만 지원합니다.")
    return category


def parse_bool(value: str | None) -> bool:
    return clean_token_value(value).casefold() in {"1", "true", "yes", "y", "on", "apply"}


def save_artifact_source(
    temp_dir: Path,
    file_items: dict[str, list[tuple[str, bytes]]],
    category: str,
) -> Path:
    upload_items = file_items.get("artifact_files") or file_items.get("template_files") or []
    if not upload_items:
        label = "관리산출물" if category == "management" else "개발산출물"
        raise ValueError(f"표지를 읽을 {label} 파일 또는 폴더를 업로드해주세요.")

    artifact_dir = temp_dir / "artifacts"
    artifact_dir.mkdir(parents=True, exist_ok=True)
    saved_count = 0
    for index, (filename, payload) in enumerate(upload_items, start=1):
        if not payload or is_noise_filename(filename):
            continue
        suffix = Path(filename).suffix.lower()
        if suffix not in SUPPORTED_TEMPLATE_SUFFIXES:
            continue
        relative_path = safe_relative_upload_path(filename, f"artifact-{index}{suffix}")
        target_path = artifact_dir / relative_path
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(payload)
        saved_count += 1

    if saved_count == 0:
        raise ValueError("업로드된 산출물에서 지원하는 문서 파일을 찾지 못했습니다.")
    return artifact_dir


def save_standard_file(temp_dir: Path, file_items: dict[str, list[tuple[str, bytes]]]) -> Path:
    items = file_items.get("standard_file") or []
    if not items or not items[0][1]:
        raise ValueError("문서관리표준 파일을 선택하세요.")
    filename, payload = items[0]
    suffix = Path(filename).suffix.lower()
    if suffix not in STANDARD_SUFFIXES:
        raise ValueError("문서관리표준은 .pdf, .hwp, .hwpx만 사용할 수 있습니다.")
    target = temp_dir / safe_upload_filename(filename, "standard_file", suffix)
    target.write_bytes(payload)
    return target


def save_requirement_sources_for_category(
    temp_dir: Path,
    file_items: dict[str, list[tuple[str, bytes]]],
    category: str,
) -> list[Path]:
    if category == "management":
        return save_proposal_requirement_uploads(temp_dir, file_items)
    return save_requirement_uploads(temp_dir, file_items)


def load_body_templates(template_dir: Path) -> list[BodyTemplate]:
    if not template_dir.exists() or not template_dir.is_dir():
        raise ValueError(f"본문 템플릿 폴더를 찾지 못했습니다: {template_dir}")

    templates: list[BodyTemplate] = []
    for path in sorted(template_dir.rglob("*"), key=lambda item: str(item).casefold()):
        if not path.is_file() or is_noise_filename(path.name):
            continue
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_TEMPLATE_SUFFIXES:
            continue
        try:
            relative_path = str(path.relative_to(template_dir))
        except ValueError:
            relative_path = path.name
        templates.append(
            BodyTemplate(
                path=path,
                relative_path=relative_path,
                suffix=suffix,
                name_key=path.name.casefold(),
                relative_key=path_key(relative_path),
                stem_key=normalize_for_match(path.stem),
                cover_type=cover_type_for(path),
            )
        )
    return templates


def build_requirement_sources(paths: list[Path]) -> list[RequirementSource]:
    sources: list[RequirementSource] = []
    for path in paths:
        requirement_ids = extract_requirement_ids(path.name)
        if requirement_ids:
            sources.append(RequirementSource(path=path, requirement_ids=requirement_ids))
    return sources


def collect_requirement_ids(sources: list[RequirementSource]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for source in sources:
        for requirement_id in source.requirement_ids:
            key = requirement_id.upper()
            if key in seen:
                continue
            seen.add(key)
            result.append(key)
    return result


def build_cover_token_values(
    fields: dict[str, str],
    requirement_ids: list[str],
    standard_project_title: str = "",
) -> dict[str, str]:
    version = DEFAULT_DOCUMENT_VERSION
    revision_date = normalize_revision_date_input(
        fields.get("revision_date")
        or fields.get("revision_year")
        or fields.get("initial_revision_year")
    )
    return {
        "PROJECT_NAME": clean_token_value(fields.get("project_name")) or clean_token_value(standard_project_title),
        "DOCUMENT_TITLE": clean_token_value(fields.get("document_title")),
        "DOCUMENT_ID": clean_token_value(fields.get("document_id")),
        "REQUIREMENT_IDS": ", ".join(requirement_ids),
        "DOCUMENT_VERSION": version,
        "VERSION": version,
        "REVISION_DATE": revision_date,
        "AUTHOR": clean_token_value(
            fields.get("author")
            or fields.get("revision_author")
            or fields.get("initial_revision_author")
        ),
        "APPROVER": clean_token_value(
            fields.get("approver")
            or fields.get("approval_author")
            or fields.get("initial_revision_approval_author")
        ),
        "COMPANY_NAME": clean_token_value(fields.get("company_name")),
        "REVISION_REASON": clean_token_value(fields.get("revision_reason")),
        "REVISION_DETAIL": clean_token_value(fields.get("revision_detail")),
    }


def normalize_revision_date_input(value: str | None) -> str:
    text = clean_token_value(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{4}", text):
        return f"{text}-00-00"
    match = re.fullmatch(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return text


def clean_token_value(value: str | None) -> str:
    return str(value or "").strip()


def scan_artifact_inputs(
    source_dir: Path,
    standard_outputs: list[StandardOutput],
    body_templates: list[BodyTemplate],
    token_values: dict[str, str],
) -> list[TemplateBuildInput]:
    if not source_dir.exists() or not source_dir.is_dir():
        raise ValueError(f"산출물 폴더를 찾지 못했습니다: {source_dir}")

    items: list[TemplateBuildInput] = []
    for path in sorted(source_dir.rglob("*"), key=lambda item: str(item).casefold()):
        if not path.is_file() or is_noise_filename(path.name):
            continue
        if path.suffix.lower() not in SUPPORTED_TEMPLATE_SUFFIXES:
            continue

        try:
            relative_path = str(path.relative_to(source_dir))
        except ValueError:
            relative_path = path.name

        if should_ignore_artifact_relative_path(relative_path):
            continue

        identity = read_file_identity(path)
        input_has_cover = input_file_has_cover(path, identity)
        standard_output, standard_match_type = find_standard_output(path, relative_path, identity, standard_outputs)
        cover_type = cover_type_for(path)
        cover_template = cover_template_for(path)
        body_template, body_match_type = find_body_template(
            path,
            relative_path,
            identity,
            standard_output,
            body_templates,
        )
        item_token_values = token_values_for_item(token_values, identity, standard_output)
        use_template_cover = bool(input_has_cover and standard_output is not None and body_template is not None and cover_template is not None)
        output_relative_path = (
            make_output_relative_path(
                relative_path,
                cover_template,
                item_token_values,
            )
            if use_template_cover
            else relative_path
        )
        if cover_type == "hwp":
            output_relative_path = force_hwpx_relative_path(output_relative_path)
        if identity.error:
            status = "fallback"
            message = f"입력 파일 표지를 읽지 못해 입력 파일 복사본으로 결과를 생성합니다: {identity.error}"
        elif standard_output is None:
            status = "fallback"
            message = "문서관리표준에서 입력 파일과 맞는 산출물명을 찾지 못해 입력 파일 복사본으로 결과를 생성합니다."
        elif body_template is None:
            status = "fallback"
            message = "서버 본문 템플릿 파일을 찾지 못해 입력 파일 복사본으로 결과를 생성합니다."
        elif cover_template is None:
            status = "matched"
            message = "서버 표지 템플릿 파일이 없어 입력 파일 표지를 유지하고 값만 갱신합니다."
        elif not input_has_cover:
            status = "matched"
            message = "입력 파일에 표지가 없어 표지를 만들지 않고 입력 파일 복사본으로 결과를 생성합니다."
        else:
            status = "matched"
            message = "입력 파일명/문서명과 서버 본문 템플릿 파일을 매칭했습니다."

        items.append(
            TemplateBuildInput(
                path=path,
                relative_path=relative_path,
                standard_output=standard_output,
                standard_match_type=standard_match_type,
                cover_type=cover_type,
                cover_template=cover_template,
                body_template=body_template,
                body_match_type=body_match_type,
                output_relative_path=output_relative_path,
                document_title=identity.document_title,
                document_number=identity.document_number,
                project_title=identity.project_title,
                input_has_cover=input_has_cover,
                token_values=item_token_values,
                status=status,
                message=message,
            )
        )
    return items


def should_ignore_artifact_relative_path(relative_path: str) -> bool:
    parts = Path(relative_path).parts[:-1]
    return any(normalize_for_match(part) in IGNORED_UPLOAD_FOLDER_KEYS for part in parts)


def input_file_has_cover(path: Path, identity: object) -> bool:
    suffix = path.suffix.lower()
    if suffix in COVER_SUFFIXES_BY_TYPE["excel"]:
        return excel_file_has_named_cover_sheet(path)
    if getattr(identity, "error", ""):
        return False
    return any(
        clean_token_value(getattr(identity, field, ""))
        for field in ("project_title", "document_title", "preview_text")
    )


def excel_file_has_named_cover_sheet(path: Path) -> bool:
    try:
        from document_update.excel_ooxml import workbook_sheets

        with zipfile.ZipFile(path, "r") as workbook:
            return any("표지" in sheet.name for sheet in workbook_sheets(workbook))
    except Exception:
        return False


def find_standard_output(
    path: Path,
    relative_path: str,
    identity: object,
    outputs: list[StandardOutput],
) -> tuple[StandardOutput | None, str]:
    document_title_key = normalize_for_match(getattr(identity, "document_title", ""))
    file_key = normalize_for_match(path.stem)
    relative_key = normalize_for_match(relative_path)

    matchers = (
        ("cover_document_title", lambda output: document_title_matches_output(document_title_key, output)),
        ("filename_output_name", lambda output: output_name_matches_key(file_key, output)),
        ("path_output_name", lambda output: output_name_matches_key(relative_key, output)),
        ("filename_output_id", lambda output: output_id_matches_key(file_key, output)),
        ("path_output_id", lambda output: output_id_matches_key(relative_key, output)),
    )
    for match_type, predicate in matchers:
        for output in outputs:
            if predicate(output):
                return output, match_type
    return None, ""


def output_id_matches_key(value_key: str, output: StandardOutput) -> bool:
    if not value_key:
        return False
    output_id_key = normalize_for_match(output.output_id)
    return bool(output_id_key and (value_key == output_id_key or output_id_key in value_key))


def document_title_matches_output(document_title_key: str, output: StandardOutput) -> bool:
    if not document_title_key:
        return False
    return output_name_matches_key(document_title_key, output)


def output_name_matches_key(value_key: str, output: StandardOutput) -> bool:
    if not value_key:
        return False
    names = (output.output_name, *output.aliases)
    for name in names:
        name_key = normalize_for_match(name)
        if name_key and (value_key == name_key or name_key in value_key):
            return True
    return False


def find_body_template(
    path: Path,
    relative_path: str,
    identity: object,
    standard_output: StandardOutput | None,
    body_templates: list[BodyTemplate],
) -> tuple[BodyTemplate | None, str]:
    suffix = path.suffix.lower()
    relative_key = path_key(relative_path)
    name_key = path.name.casefold()
    stem_key = normalize_for_match(path.stem)
    document_title = normalize_for_match(getattr(identity, "document_title", ""))
    cover_type = cover_type_for(path)
    standard_id = normalize_for_match(standard_output.output_id if standard_output else "")
    standard_names = tuple(
        normalize_for_match(name)
        for name in ((standard_output.output_name, *standard_output.aliases) if standard_output else ())
        if normalize_for_match(name)
    )

    matchers = (
        ("same_relative_path", lambda item: item.suffix == suffix and item.relative_key == relative_key),
        ("same_filename", lambda item: item.suffix == suffix and item.name_key == name_key),
        ("standard_id", lambda item: bool(standard_id) and item.cover_type == cover_type and standard_id in item.stem_key),
        ("standard_output_name", lambda item: item.cover_type == cover_type and any(name == item.stem_key or name in item.stem_key for name in standard_names)),
        ("same_normalized_filename", lambda item: item.suffix == suffix and item.stem_key == stem_key),
        ("same_document_title", lambda item: bool(document_title) and item.cover_type == cover_type and item.stem_key == document_title),
        ("same_filename_family", lambda item: item.cover_type == cover_type and item.stem_key == stem_key),
    )
    for match_type, predicate in matchers:
        for template in body_templates:
            if predicate(template):
                return template, match_type
    return None, ""


def path_key(value: str) -> str:
    return str(value or "").replace("\\", "/").casefold()


def cover_type_for(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".pptx", ".pptm"}:
        return "ppt"
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return "excel"
    return "hwp"


def cover_template_for(path: Path) -> Path | None:
    cover_type = cover_type_for(path)
    typed_candidates = [
        candidate
        for candidate in sorted(COVER_TEMPLATE_DIR.iterdir(), key=lambda item: item.name.casefold())
        if candidate.is_file()
        and not is_noise_filename(candidate.name)
        and candidate.suffix.lower() in COVER_SUFFIXES_BY_TYPE[cover_type]
    ] if COVER_TEMPLATE_DIR.exists() else []

    for candidate in typed_candidates:
        if TOKEN_PATTERN.search(candidate.name):
            return candidate

    if typed_candidates:
        return typed_candidates[0]
    return None


def make_output_relative_path(
    artifact_relative_path: str,
    cover_template: Path | None,
    item_values: dict[str, str],
) -> str:
    parent = Path(artifact_relative_path).parent
    if cover_template is None or not TOKEN_PATTERN.search(cover_template.name):
        return artifact_relative_path

    rendered_name = TOKEN_PATTERN.sub(
        lambda match: item_values.get(match.group(1), match.group(0)),
        cover_template.name,
    )
    output_name = sanitize_output_filename(rendered_name, fallback=Path(artifact_relative_path).name)
    if str(parent) in {"", "."}:
        return output_name
    return str(parent / output_name)


def force_hwpx_relative_path(relative_path: str) -> str:
    path = Path(relative_path)
    return str(path.with_suffix(".hwpx"))


def token_values_for_item(
    base_token_values: dict[str, str],
    identity: object,
    standard_output: StandardOutput | None,
) -> dict[str, str]:
    values = dict(base_token_values)
    values["PROJECT_NAME"] = values.get("PROJECT_NAME") or clean_token_value(getattr(identity, "project_title", ""))
    values["DOCUMENT_TITLE"] = (
        values.get("DOCUMENT_TITLE")
        or clean_token_value(standard_output.output_name if standard_output else "")
        or clean_token_value(getattr(identity, "document_title", ""))
    )
    values["DOCUMENT_ID"] = (
        values.get("DOCUMENT_ID")
        or clean_token_value(standard_output.output_id if standard_output else "")
        or clean_token_value(getattr(identity, "document_number", ""))
    )
    values["DOCUMENT_VERSION"] = values.get("DOCUMENT_VERSION") or values.get("VERSION", "") or DEFAULT_DOCUMENT_VERSION
    values["VERSION"] = values.get("VERSION") or values.get("DOCUMENT_VERSION", "") or DEFAULT_DOCUMENT_VERSION
    return values


def sanitize_output_filename(name: str, *, fallback: str) -> str:
    path = Path(name)
    suffix = path.suffix
    stem = filesystem_safe_stem(path.stem)
    if not stem:
        return fallback
    return f"{stem}{suffix}" if suffix else stem


def copy_template_outputs(
    items: list[TemplateBuildInput],
    output_root: Path,
    *,
    work_root: Path | None = None,
) -> list[dict[str, object]]:
    output_dir = output_root / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    work_dir = work_root or output_root / "_template_work"
    work_dir.mkdir(parents=True, exist_ok=True)
    copied: list[dict[str, object]] = []
    used_targets: set[str] = set()
    try:
        for item in items:
            use_template_source = should_use_template_source(item)
            if use_template_source:
                source_path = prepare_body_template_for_output(item.body_template.path, item.cover_type, work_dir)
                source_mode = "server_template"
            else:
                source_path = prepare_input_file_for_output(item.path, item.cover_type, work_dir)
                source_mode = "input_copy"
            for variant in output_variants_for_item(item):
                target = unique_output_path(
                    output_target_path(output_dir, variant.output_relative_path, item.cover_type),
                    used_targets,
                )
                target.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source_path, target)
                if use_template_source:
                    cover_update_count = apply_cover_template_to_output(target, item, work_dir, variant.token_values)
                else:
                    cover_update_count = apply_existing_cover_updates_to_output(
                        target,
                        item,
                        work_dir,
                        variant.token_values,
                    )
                copied.append(
                    {
                        "name": target.name,
                        "path": target,
                        "cover_update_count": cover_update_count,
                        "requirement_id": variant.requirement_id,
                        "source_mode": source_mode,
                    }
                )
    finally:
        remove_template_work_dir(work_dir)
    return copied


def should_use_template_source(item: TemplateBuildInput) -> bool:
    return (
        item.status == "matched"
        and item.input_has_cover
        and item.body_template is not None
        and item.cover_template is not None
    )


def prepare_input_file_for_output(path: Path, cover_type: str, work_dir: Path) -> Path:
    if cover_type != "hwp":
        return path
    prepared_path, _converted = prepare_target_file(path.resolve(strict=False), work_dir / "input-converted")
    if not is_hwpx_zip(prepared_path):
        raise RuntimeError(f"입력 한글 파일을 HWPX로 변환하지 못했습니다: {path.name}")
    return prepared_path


def apply_existing_cover_updates_to_output(
    target: Path,
    item: TemplateBuildInput,
    work_dir: Path,
    token_values: dict[str, str],
) -> int:
    if not item.input_has_cover:
        return 0

    update_count = 0
    update_count += update_existing_project_title(target, item, work_dir, token_values)
    update_count += update_existing_revision_metadata(target, work_dir, token_values)
    return update_count


def update_existing_project_title(
    target: Path,
    item: TemplateBuildInput,
    work_dir: Path,
    token_values: dict[str, str],
) -> int:
    old_project_title = clean_token_value(item.project_title)
    new_project_title = clean_token_value(token_values.get("PROJECT_NAME"))
    if not values_differ(old_project_title, new_project_title):
        return 0

    update_dir = work_dir / "input-cover-project"
    update_dir.mkdir(parents=True, exist_ok=True)
    temp_output = update_dir / f"{target.stem}-{uuid4().hex}{target.suffix}"
    try:
        replace_count, updated_path = write_updated_project_title(
            target,
            old_project_title,
            new_project_title,
            output_path=temp_output,
        )
        if replace_count > 0 and updated_path.exists() and not same_path(updated_path, target):
            replace_file_with_fallback(updated_path, target)
        elif temp_output.exists():
            temp_output.unlink()
        return replace_count
    except Exception:
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                pass
        return 0


def update_existing_revision_metadata(
    target: Path,
    work_dir: Path,
    token_values: dict[str, str],
) -> int:
    revision_date = clean_token_value(token_values.get("REVISION_DATE"))
    author = clean_token_value(token_values.get("AUTHOR"))
    approver = clean_token_value(token_values.get("APPROVER"))
    if not (revision_date or author or approver):
        return 0

    metadata_result = update_metadata_in_document(
        target,
        author,
        revision_date,
        approver,
        work_dir / "input-cover-metadata",
        document_number="",
    )
    if metadata_result.status == "error":
        return 0
    return metadata_result.cover_update_count + metadata_result.revision_history_update_count


def values_differ(old_value: str | None, new_value: str | None) -> bool:
    if not old_value or not new_value:
        return False
    return normalize_for_match(old_value) != normalize_for_match(new_value)


def same_path(left: Path, right: Path) -> bool:
    return left.resolve(strict=False) == right.resolve(strict=False)


def replace_file_with_fallback(source: Path, target: Path) -> None:
    try:
        source.replace(target)
        return
    except OSError:
        shutil.copyfile(source, target)
        try:
            source.unlink()
        except OSError:
            pass


def remove_template_work_dir(work_dir: Path) -> None:
    for _attempt in range(5):
        if not work_dir.exists():
            return
        try:
            shutil.rmtree(work_dir)
            return
        except OSError:
            time.sleep(0.1)
    shutil.rmtree(work_dir, ignore_errors=True)


def output_variants_for_item(item: TemplateBuildInput) -> list[TemplateOutputVariant]:
    requirement_ids = requirement_ids_from_token_value(item.token_values.get("REQUIREMENT_IDS", ""))
    if len(requirement_ids) <= 1:
        requirement_id = requirement_ids[0] if requirement_ids else ""
        output_relative_path = ensure_requirement_id_in_relative_path(item.output_relative_path, requirement_id)
        if item.cover_type == "hwp":
            output_relative_path = force_hwpx_relative_path(output_relative_path)
        return [TemplateOutputVariant(requirement_id, output_relative_path, item.token_values)]

    variants: list[TemplateOutputVariant] = []
    for requirement_id in requirement_ids:
        token_values = dict(item.token_values)
        token_values["REQUIREMENT_IDS"] = requirement_id
        output_relative_path = (
            make_output_relative_path(item.relative_path, item.cover_template, token_values)
            if should_use_template_source(item)
            else item.output_relative_path
        )
        output_relative_path = ensure_requirement_id_in_relative_path(output_relative_path, requirement_id)
        if item.cover_type == "hwp":
            output_relative_path = force_hwpx_relative_path(output_relative_path)
        variants.append(TemplateOutputVariant(requirement_id, output_relative_path, token_values))
    return variants


def requirement_ids_from_token_value(value: str) -> list[str]:
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def ensure_requirement_id_in_relative_path(relative_path: str, requirement_id: str) -> str:
    if not requirement_id:
        return relative_path
    path = Path(relative_path)
    requirement_part = filesystem_safe_stem(requirement_id)
    if requirement_part.casefold() in path.stem.casefold():
        return relative_path
    return str(path.with_name(f"{path.stem}_{requirement_part}{path.suffix}"))


def output_target_path(output_dir: Path, output_relative_path: str, cover_type: str) -> Path:
    target = output_dir / Path(output_relative_path)
    if cover_type == "hwp":
        return target.with_suffix(".hwpx")
    return target


def prepare_body_template_for_output(path: Path, cover_type: str, work_dir: Path) -> Path:
    if cover_type != "hwp":
        return path
    prepared_path, _converted = prepare_target_file(path.resolve(strict=False), work_dir / "body-converted")
    if not is_hwpx_zip(prepared_path):
        raise RuntimeError(f"한글 본문 템플릿을 HWPX로 변환하지 못했습니다: {path.name}")
    return prepared_path


def apply_cover_template_to_output(
    target: Path,
    item: TemplateBuildInput,
    work_dir: Path,
    token_values: dict[str, str] | None = None,
) -> int:
    if item.cover_template is None:
        return 0
    values = token_values or item.token_values
    if item.cover_type == "excel":
        return apply_excel_cover_template(target, item.cover_template, values)
    if item.cover_type == "ppt":
        return apply_ppt_cover_template(target, item.cover_template, values)
    if item.cover_type == "hwp":
        return apply_hwpx_cover_template(target, item.cover_template, values, work_dir)
    return 0


def render_token_string(value: str, token_values: dict[str, str], *, xml: bool = False) -> str:
    def replace_token(match: re.Match[str]) -> str:
        replacement = token_values.get(match.group(1), "")
        return escape(replacement) if xml else replacement

    return TOKEN_PATTERN.sub(replace_token, value)


def render_token_cell_value(value: object, token_values: dict[str, str]) -> object:
    if isinstance(value, str):
        return render_token_string(value, token_values)
    return value


def apply_excel_cover_template(target: Path, cover_template: Path, token_values: dict[str, str]) -> int:
    from openpyxl import load_workbook

    cover_wb = load_workbook(cover_template)
    target_wb = load_workbook(target, keep_vba=target.suffix.lower() in {".xlsm", ".xltm"})
    try:
        cover_ws = select_excel_cover_sheet(cover_wb)
        old_cover_ws = find_excel_cover_sheet(target_wb)
        old_index = target_wb.worksheets.index(old_cover_ws) if old_cover_ws is not None else 0
        new_ws = target_wb.create_sheet("__template_cover__", old_index)
        changed_cells = copy_excel_cover_sheet(cover_ws, new_ws, token_values)
        if old_cover_ws is not None:
            target_wb.remove(old_cover_ws)
        new_ws.title = unique_excel_sheet_title(target_wb, cover_ws.title)
        target_wb.save(target)
        return changed_cells or 1
    finally:
        cover_wb.close()
        target_wb.close()


def select_excel_cover_sheet(workbook: object) -> object:
    return find_excel_cover_sheet(workbook) or workbook.worksheets[0]


def find_excel_cover_sheet(workbook: object) -> object | None:
    for sheet in workbook.worksheets:
        if "표지" in sheet.title:
            return sheet
    return None


def unique_excel_sheet_title(workbook: object, desired_title: str) -> str:
    title = desired_title or "표지"
    existing = {sheet.title for sheet in workbook.worksheets}
    if title not in existing:
        return title
    index = 2
    while f"{title}_{index}" in existing:
        index += 1
    return f"{title}_{index}"


def copy_excel_cover_sheet(source: object, target: object, token_values: dict[str, str]) -> int:
    changed_cells = 0
    target.sheet_properties = copy_object(source.sheet_properties)
    target.sheet_format = copy_object(source.sheet_format)
    target.page_margins = copy_object(source.page_margins)
    target.page_setup = copy_object(source.page_setup)
    target.print_options = copy_object(source.print_options)
    target.freeze_panes = source.freeze_panes

    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key] = copy_object(dimension)
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key] = copy_object(dimension)

    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))

    for row in source.iter_rows():
        for cell in row:
            target_cell = target[cell.coordinate]
            new_value = render_token_cell_value(cell.value, token_values)
            if new_value != cell.value:
                changed_cells += 1
            target_cell.value = new_value
            if cell.has_style:
                target_cell._style = copy_object(cell._style)
            if cell.number_format:
                target_cell.number_format = cell.number_format
            if cell.hyperlink:
                target_cell._hyperlink = copy_object(cell.hyperlink)
            if cell.comment:
                target_cell.comment = copy_object(cell.comment)

    for image in getattr(source, "_images", []):
        try:
            target.add_image(copy_object(image), image.anchor)
        except Exception:
            continue

    return changed_cells


def apply_ppt_cover_template(target: Path, cover_template: Path, token_values: dict[str, str]) -> int:
    if not zipfile.is_zipfile(target) or not zipfile.is_zipfile(cover_template):
        raise RuntimeError("PPT 표지는 PPTX/PPTM 압축 형식만 적용할 수 있습니다.")

    with zipfile.ZipFile(cover_template, "r") as cover_zip, zipfile.ZipFile(target, "r") as target_zip:
        cover_slide = render_xml_tokens(cover_zip.read("ppt/slides/slide1.xml"), token_values)
        cover_rels, extra_parts = build_ppt_cover_relationships(cover_zip, target_zip)

    replacements = {"ppt/slides/slide1.xml": cover_slide}
    if cover_rels is not None:
        replacements["ppt/slides/_rels/slide1.xml.rels"] = cover_rels
    rewrite_zip_file(target, replacements, extra_parts)
    return 1


def build_ppt_cover_relationships(
    cover_zip: zipfile.ZipFile,
    target_zip: zipfile.ZipFile,
) -> tuple[bytes | None, dict[str, bytes]]:
    rels_path = "ppt/slides/_rels/slide1.xml.rels"
    if rels_path not in cover_zip.namelist():
        return None, {}

    ET.register_namespace("", REL_NS)
    cover_root = ET.fromstring(cover_zip.read(rels_path))
    target_layout = target_slide_layout_relationship(target_zip)
    extra_parts: dict[str, bytes] = {}
    used_names = set(target_zip.namelist()) | set(cover_zip.namelist())

    for relationship in list(cover_root):
        rel_type = relationship.attrib.get("Type", "")
        if rel_type.endswith("/slideLayout"):
            cover_root.remove(relationship)
            continue

        target = relationship.attrib.get("Target", "")
        source_part = resolve_ppt_relationship_target("ppt/slides/slide1.xml", target)
        if source_part in cover_zip.namelist() and source_part.startswith("ppt/media/"):
            new_part = unique_zip_part_name("ppt/media", Path(source_part).name, used_names)
            used_names.add(new_part)
            extra_parts[new_part] = cover_zip.read(source_part)
            relationship.set("Target", f"../media/{Path(new_part).name}")

    if target_layout is not None:
        used_ids = {relationship.attrib.get("Id", "") for relationship in cover_root}
        layout = ET.Element(f"{{{REL_NS}}}Relationship", target_layout.attrib)
        layout.set("Id", unique_relationship_id(used_ids))
        cover_root.append(layout)

    return ET.tostring(cover_root, encoding="utf-8", xml_declaration=True), extra_parts


def target_slide_layout_relationship(target_zip: zipfile.ZipFile) -> ET.Element | None:
    rels_path = "ppt/slides/_rels/slide1.xml.rels"
    if rels_path not in target_zip.namelist():
        return None
    try:
        root = ET.fromstring(target_zip.read(rels_path))
    except ET.ParseError:
        return None
    for relationship in root:
        if relationship.attrib.get("Type", "").endswith("/slideLayout"):
            return relationship
    return None


def resolve_ppt_relationship_target(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return str(PurePosixPath(target.lstrip("/")))
    return posixpath.normpath(str(PurePosixPath(base_part).parent / target))


def unique_relationship_id(used_ids: set[str]) -> str:
    index = 1
    while f"rIdTemplateLayout{index}" in used_ids:
        index += 1
    return f"rIdTemplateLayout{index}"


def apply_hwpx_cover_template(
    target: Path,
    cover_template: Path,
    token_values: dict[str, str],
    work_dir: Path,
) -> int:
    # HWPX는 다른 문서의 section0.xml을 통째로 끼워 넣으면 한글의 내부 참조/보안
    # 검증에 걸릴 수 있다. 결과 문서 구조는 유지하고 텍스트/메타데이터만 갱신한다.
    if not is_hwpx_zip(target):
        raise RuntimeError("한글 결과 파일이 HWPX 형식이 아닙니다.")

    update_count = replace_hwpx_tokens_in_place(target, token_values)
    update_count += update_hwpx_cover_text_in_place(target, token_values, work_dir)
    return update_count


def replace_hwpx_tokens_in_place(target: Path, token_values: dict[str, str]) -> int:
    replacements: dict[str, bytes] = {}
    replace_count = 0
    with zipfile.ZipFile(target, "r") as target_zip:
        for name in target_zip.namelist():
            if not name.lower().endswith(".xml"):
                continue
            xml = target_zip.read(name).decode("utf-8", errors="ignore")
            rendered = render_token_string(xml, token_values, xml=True)
            if rendered == xml:
                continue
            replace_count += sum(1 for _match in TOKEN_PATTERN.finditer(xml))
            replacements[name] = rendered.encode("utf-8")
    if replacements:
        rewrite_zip_file(target, replacements, {})
    return replace_count


def update_hwpx_cover_text_in_place(target: Path, token_values: dict[str, str], work_dir: Path) -> int:
    update_count = 0
    document_id = clean_token_value(token_values.get("DOCUMENT_ID"))
    project_name = clean_token_value(token_values.get("PROJECT_NAME"))
    if document_id or project_name:
        update_dir = work_dir / "hwp-cover-text"
        update_dir.mkdir(parents=True, exist_ok=True)
        temp_output = update_dir / f"{target.stem}-{uuid4().hex}{target.suffix}"
        try:
            _old_number, _backup, project_count, document_count, updated_path = write_updated_document(
                target,
                new_document_number=document_id,
                old_project_title="",
                new_project_title=project_name,
                output_path=temp_output,
                allow_missing_document_number=True,
            )
            if updated_path.exists():
                shutil.move(str(updated_path), str(target))
                update_count += project_count + document_count
        except Exception:
            if temp_output.exists():
                try:
                    temp_output.unlink()
                except OSError:
                    pass

    revision_date = clean_token_value(token_values.get("REVISION_DATE"))
    author = clean_token_value(token_values.get("AUTHOR"))
    approver = clean_token_value(token_values.get("APPROVER"))
    if revision_date or author or approver or document_id:
        metadata_result = update_metadata_in_document(
            target,
            author,
            revision_date,
            approver,
            work_dir / "hwp-metadata",
            document_number=document_id,
        )
        if metadata_result.status != "error":
            update_count += metadata_result.cover_update_count + metadata_result.revision_history_update_count
    return update_count


def render_xml_tokens(data: bytes, token_values: dict[str, str]) -> bytes:
    xml = data.decode("utf-8", errors="ignore")
    return render_token_string(xml, token_values, xml=True).encode("utf-8")


def rewrite_zip_file(path: Path, replacements: dict[str, bytes], additions: dict[str, bytes]) -> None:
    temp_path = path.with_name(f"{path.stem}-cover-{uuid4().hex}{path.suffix}")
    written: set[str] = set()
    with zipfile.ZipFile(path, "r") as source_zip:
        with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as output_zip:
            for info in source_zip.infolist():
                data = replacements.get(info.filename, source_zip.read(info.filename))
                output_zip.writestr(info, data)
                written.add(info.filename)
            for name, data in additions.items():
                if name not in written:
                    output_zip.writestr(name, data)
                    written.add(name)
            for name, data in replacements.items():
                if name not in written:
                    output_zip.writestr(name, data)
                    written.add(name)
    try:
        temp_path.replace(path)
    except OSError:
        shutil.copyfile(temp_path, path)
        try:
            temp_path.unlink()
        except OSError:
            pass


def unique_zip_part_name(directory: str, filename: str, used_names: set[str]) -> str:
    stem = Path(filename).stem
    suffix = Path(filename).suffix
    index = 1
    while True:
        candidate = f"{directory}/cover_{index}_{stem}{suffix}"
        if candidate not in used_names:
            return candidate
        index += 1


def unique_output_path(path: Path, used_targets: set[str]) -> Path:
    candidate = path
    index = 2
    while str(candidate).casefold() in used_targets or candidate.exists():
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        index += 1
    used_targets.add(str(candidate).casefold())
    return candidate


def write_template_build_report(
    output_root: Path,
    *,
    category: str,
    standard_path: Path,
    standard_output_count: int,
    source_dir: Path,
    body_template_dir: Path,
    body_template_count: int,
    requirement_ids: list[str],
    requirement_sources: list[Path],
    token_values: dict[str, str],
    items: list[TemplateBuildInput],
) -> Path:
    report_path = output_root / "TEMPLATE_BUILD_REPORT.md"
    lines = [
        "# 템플릿 출력 리포트",
        "",
        f"- 템플릿 종류: {category}",
        f"- 문서관리표준: {standard_path.name}",
        f"- 표준 ID 수: {standard_output_count}",
        f"- 입력 산출물 폴더: {source_dir}",
        f"- 본문 템플릿 폴더: {body_template_dir}",
        f"- 입력 산출물 파일 수: {len(items)}",
        f"- 본문 템플릿 파일 수: {body_template_count}",
        f"- 요구사항 ID 수: {len(requirement_ids)}",
        "",
        "## 요구사항 ID",
        "",
    ]
    lines.extend(f"- {item}" for item in requirement_ids)
    if not requirement_ids:
        lines.append("- 없음")
    source_details = read_requirement_source_details(requirement_sources)
    if source_details:
        lines.extend(
            [
                "",
                "## 요구사항 ID 추출 근거",
                "",
                "| 상태 | 요구사항 ID | 원본 파일 | 매칭 텍스트 | 주변 문맥 | 사유 |",
                "|---|---|---|---|---|---|",
            ]
        )
        for detail in source_details:
            lines.append(
                "| "
                + " | ".join(
                    markdown_cell(detail.get(key, ""))
                    for key in ("status", "requirement_id", "source", "matched_text", "context", "ignore_reason")
                )
                + " |"
            )

    lines.extend(
        [
            "",
            "## 치환 토큰",
            "",
            "| 토큰 | 값 |",
            "|---|---|",
        ]
    )
    for token, value in token_values.items():
        lines.append(f"| `{{{{{token}}}}}` | {markdown_cell(value) or '-'} |")

    lines.extend(
        [
            "",
            "## 표지 매칭",
            "",
            "| 상태 | 입력 산출물 | 표준 ID | 표준 산출물명 | 출력 파일 | 결과 소스 | 입력 표지 | 본문 템플릿 | 매칭기준 | 표지유형 | 표지템플릿 | 문서번호 | 문서명 | 프로젝트명 | 메시지 |",
            "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|",
        ]
    )
    for item in items:
        lines.append(
            "| "
            + " | ".join(
                markdown_cell(value)
                for value in (
                    item.status,
                    item.relative_path,
                    item.standard_output.output_id if item.standard_output else "없음",
                    item.standard_output.output_name if item.standard_output else "없음",
                    ", ".join(variant.output_relative_path for variant in output_variants_for_item(item)),
                    "server_template" if should_use_template_source(item) else "input_copy",
                    "있음" if item.input_has_cover else "없음",
                    item.body_template.relative_path if item.body_template else "없음",
                    item.body_match_type or item.standard_match_type,
                    item.cover_type,
                    item.cover_template.name if item.cover_template else "없음",
                    item.document_number,
                    item.document_title,
                    item.project_title,
                    item.message,
                )
            )
            + " |"
        )

    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report_path


def read_requirement_source_details(requirement_sources: list[Path]) -> list[dict[str, str]]:
    details: list[dict[str, str]] = []
    for path in requirement_sources:
        if not path.is_file():
            continue
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        values: dict[str, str] = {}
        for line in text.splitlines():
            key, separator, value = line.partition(":")
            if not separator:
                continue
            key = key.strip()
            if key in {"source", "requirement_id", "ignored_requirement_id", "matched_text", "context", "ignore_reason"}:
                values[key] = value.strip()
        if values.get("ignored_requirement_id"):
            values["status"] = "제외"
            values["requirement_id"] = values["ignored_requirement_id"]
        elif values.get("requirement_id"):
            values["status"] = "사용"
        if values.get("requirement_id"):
            details.append(values)
    return details


def serialize_template_item(item: TemplateBuildInput, output_root: Path, *, applied: bool) -> dict[str, object]:
    variants = output_variants_for_item(item)
    output_paths = [
        str(output_target_path(output_root / "outputs", variant.output_relative_path, item.cover_type))
        for variant in variants
    ] if applied else []
    output_relative_paths = [variant.output_relative_path for variant in variants]
    return {
        "path": item.relative_path,
        "standard_output": {
            "id": item.standard_output.output_id if item.standard_output else "",
            "name": item.standard_output.output_name if item.standard_output else "",
            "match_type": item.standard_match_type,
        },
        "output_name": Path(output_relative_paths[0]).name if output_relative_paths else Path(item.output_relative_path).name,
        "output_relative_path": output_relative_paths[0] if output_relative_paths else item.output_relative_path,
        "output_relative_paths": output_relative_paths,
        "output_path": output_paths[0] if output_paths else "",
        "output_paths": output_paths,
        "status": item.status,
        "message": item.message,
        "cover_type": item.cover_type,
        "input_has_cover": item.input_has_cover,
        "output_source": "server_template" if should_use_template_source(item) else "input_copy",
        "cover_template": item.cover_template.name if item.cover_template else "",
        "body_template": item.body_template.relative_path if item.body_template else "",
        "body_match_type": item.body_match_type,
        "identity": {
            "project_title": item.project_title,
            "document_title": item.document_title,
            "document_number": item.document_number,
        },
    }


def markdown_cell(value: object) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")
