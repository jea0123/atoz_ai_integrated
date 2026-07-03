from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
import re
import shutil
import time
from uuid import uuid4
import zipfile
from xml.sax.saxutils import escape, unescape

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .document_number import get_cell_text, replace_cell_text
from .excel_ooxml import (
    EXCEL_DOCUMENT_SUFFIXES,
    CellInfo,
    col_to_name,
    iter_cells,
    read_shared_strings,
    replace_or_insert_cell_xml,
    rewrite_excel_part_for_modified_workbook,
    workbook_sheets,
)
from .header_metadata import (
    has_unlabeled_metadata_signal as shared_has_unlabeled_metadata_signal,
    unlabeled_author_value_ok,
    unlabeled_header_metadata_indexes as shared_unlabeled_header_metadata_indexes,
    unlabeled_header_slot_is_clean,
)
from .hwpx_text import (
    editable_hwpx_part_scope,
    extract_document_text,
    is_hwpx_zip,
    strip_hwpx_line_seg_arrays,
)
from .patterns import (
    CELL_PATTERN,
    NUMBER_OUTPUT_ID_PATTERN_TEXT,
    OUTPUT_ID_PATTERN,
    ROW_PATTERN,
    split_output_id_and_name,
)
from . import ppt_ooxml
from .runtime_conversion import prepare_target_file
from output_file_check.file_noise import is_noise_filename


SUPPORTED_METADATA_SUFFIXES = {".hwp", ".hwpx", *EXCEL_DOCUMENT_SUFFIXES, *ppt_ooxml.PPT_DOCUMENT_SUFFIXES}
IGNORED_FOLDER_NAMES = {"bak", "backup", "font", "KRDS_UIUX", "__pycache__"}
SCHEDULE_SHEET_NAME = "Schedule"
WBS_START_COL = 16
WBS_AUTHOR_COL = 28
WBS_OUTPUT_COL = 29
WBS_TASK_COLS = (5, 6, 7, 8, 9, 10)
DOCUMENT_VERSION_VALUE = "v0.1"
UNLABELED_HEADER_VERSION_VALUE = DOCUMENT_VERSION_VALUE
DOCUMENT_NUMBER_LABELS = {"\ubb38\uc11c\ubc88\ud638", "\ubb38\uc11c \ubc88\ud638", "\ubb38 \uc11c \ubc88 \ud638"}
REVISION_REASON_VALUE = "\uc81c\uc815"
REVISION_DETAIL_VALUE = "\ucd5c\ucd08 \uc791\uc131"
REVISION_REASON_HEADER_LABELS = {"\uac1c\uc815\uc0ac\uc720", "\ubcc0\uacbd\uc0ac\uc720"}
REVISION_DETAIL_HEADER_LABELS = {
    "\uac1c\uc815\ub0b4\uc5ed",
    "\uac1c\uc815\ub0b4\uc6a9",
    "\ubcc0\uacbd\ub0b4\uc5ed",
    "\ubcc0\uacbd\ub0b4\uc6a9",
}
EXCEL_HEADER_METADATA_SCAN_MAX_ROW = 8
DATE_LABELS = {"개정일자"}
AUTHOR_LABELS = {"작성자", "작성 자", "작 성 자"}
VERSION_LABELS = {"문서버전", "문 서 버 전", "Version"}
REVISION_DATE_HEADER_LABELS = {"개정일자", "변경일", "변경일자"}
APPROVAL_LABELS = {"승인", "승인자"}
LABEL_LIKE_VALUES = {
    "문서번호",
    "문서버전",
    "개정일자",
    "변경일",
    "변경일자",
    "작성자",
    "작성 자",
    "작 성 자",
    "버전",
    "개정사유",
    "개정내역",
    "승인",
    "승인자",
}
VERSION_PATTERN = re.compile(r"^[vV]?\d+(?:\.\d+)*$")
NUMBER_OUTPUT_ID_PATTERN = re.compile(NUMBER_OUTPUT_ID_PATTERN_TEXT)
REQUIREMENT_ID_PATTERN = re.compile(r"(?<![A-Z0-9])SFR-(?:[A-Z0-9]+-)*\d+(?![A-Z0-9])", re.IGNORECASE)
DATE_VALUE_PATTERN = re.compile(r"^\d{4}(?:[-./]\d{1,2}[-./]\d{1,2}|\s+\d{1,2}\s+\d{1,2})$")
DRAWING_TEXT_PATTERN = re.compile(r"<(?P<tag>(?:\w+:)?t)\b[^>]*>(?P<text>.*?)</(?P=tag)>", re.DOTALL)
HWPX_CARET_POSITION_PATTERN = re.compile(r"<(?:\w+:)?CaretPosition\b[^>]*/>", re.DOTALL)


@dataclass(frozen=True)
class WbsMetadata:
    output_name: str
    author: str
    revision_date: str
    wbs: str
    task: str
    row: int
    requirement_id: str = ""


@dataclass
class DocumentMetadata:
    author: str = ""
    revision_date: str = ""
    revision_author: str = ""
    revision_history_date: str = ""


@dataclass
class MetadataTarget:
    path: Path
    relative_path: str
    status: str
    message: str
    output_name: str = ""
    author: str = ""
    revision_date: str = ""
    current: DocumentMetadata = field(default_factory=DocumentMetadata)
    candidates: list[WbsMetadata] = field(default_factory=list)


@dataclass
class MetadataWriteResult:
    status: str
    old_path: Path
    new_path: Path | None = None
    backup_path: Path | None = None
    converted_to_hwpx: bool = False
    cover_update_count: int = 0
    revision_history_update_count: int = 0
    error: str = ""


def normalize_key(value: str) -> str:
    return re.sub(r"[\s_\-(){}\[\]·.,/]+", "", value or "").casefold()


def normalize_label(value: str) -> str:
    return re.sub(r"\s+", "", value or "")


def clean_text(value: object) -> str:
    return str(value or "").strip()


def split_output_names(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    names: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"\s*(?:[,;\n&＆]|(?:\band\b))\s*", text, flags=re.IGNORECASE):
        raw = part.strip()
        _output_id, output_name = split_output_id_and_name(raw)
        for name in (raw, output_name):
            key = normalize_key(name)
            if not key or key in seen:
                continue
            names.append(name)
            seen.add(key)
    return names


def format_revision_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return datetime.combine(value, time()).strftime("%Y-%m-%d")

    text = clean_text(value)
    if not text:
        return ""

    for pattern in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], pattern).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text[:10]


def read_wbs_metadata(wbs_path: Path) -> list[WbsMetadata]:
    workbook = load_workbook(wbs_path, read_only=True, data_only=True, keep_vba=True)
    if SCHEDULE_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError("WBS 파일에서 Schedule 시트를 찾지 못했습니다.")

    sheet = workbook[SCHEDULE_SHEET_NAME]
    records: list[WbsMetadata] = []
    current_requirement_id = ""
    for row_index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        row_requirement_id = extract_requirement_id_from_values(row)
        if row_requirement_id:
            current_requirement_id = row_requirement_id

        author = clean_text(row[WBS_AUTHOR_COL - 1] if len(row) >= WBS_AUTHOR_COL else "")
        revision_date = format_revision_date(row[WBS_START_COL - 1] if len(row) >= WBS_START_COL else "")
        output_names = split_output_names(row[WBS_OUTPUT_COL - 1] if len(row) >= WBS_OUTPUT_COL else "")
        if not author or not revision_date or not output_names:
            continue

        wbs = clean_text(row[3] if len(row) >= 4 else "")
        task = next(
            (clean_text(row[column - 1]) for column in reversed(WBS_TASK_COLS) if len(row) >= column and clean_text(row[column - 1])),
            "",
        )
        for output_name in output_names:
            records.append(
                WbsMetadata(
                    output_name=output_name,
                    author=author,
                    revision_date=revision_date,
                    wbs=wbs,
                    task=task,
                    row=row_index,
                    requirement_id=current_requirement_id,
                )
            )
    return records


def extract_requirement_id_from_values(values: object) -> str:
    if isinstance(values, (str, Path)):
        text = str(values)
    else:
        try:
            text = " ".join(clean_text(value) for value in values if value is not None)
        except TypeError:
            text = clean_text(values)

    match = REQUIREMENT_ID_PATTERN.search(text)
    return match.group(0).upper() if match else ""


def should_scan_document(path: Path) -> bool:
    if path.suffix.lower() not in SUPPORTED_METADATA_SUFFIXES:
        return False
    if is_noise_filename(path.name):
        return False
    if any(part in IGNORED_FOLDER_NAMES for part in path.parts):
        return False
    if "WBS" in path.name.upper():
        return False
    return True


def collect_metadata_documents(root: Path) -> list[Path]:
    return sorted(
        [path for path in root.rglob("*") if path.is_file() and should_scan_document(path)],
        key=lambda item: str(item).casefold(),
    )


def records_matching_path(path: Path, records: list[WbsMetadata]) -> list[WbsMetadata]:
    haystack = normalize_key(" ".join(path.parts[-5:]))
    matches = [record for record in records if normalize_key(record.output_name) and normalize_key(record.output_name) in haystack]
    if not matches:
        return []

    best_length = max(len(normalize_key(record.output_name)) for record in matches)
    output_matches = [record for record in matches if len(normalize_key(record.output_name)) == best_length]

    requirement_id = extract_requirement_id_from_values(path.name)
    if requirement_id:
        return [
            record
            for record in output_matches
            if record.requirement_id.casefold() == requirement_id.casefold()
        ]

    return output_matches


def choose_record(
    path: Path,
    candidates: list[WbsMetadata],
    current: DocumentMetadata | None = None,
) -> tuple[WbsMetadata | None, str, str]:
    if not candidates:
        return None, "unmatched", "WBS 산출물명과 파일/폴더명이 매칭되지 않았습니다."

    if not extract_requirement_id_from_values(path.name):
        return earliest_record(candidates), "matched", ""

    path_text = normalize_key(str(path))
    author_matches = [record for record in candidates if normalize_key(record.author) in path_text]
    if len({(item.author, item.output_name) for item in author_matches}) == 1:
        return earliest_record(author_matches), "matched", ""

    current_author = normalize_key(current.author if current else "")
    if current_author:
        current_author_matches = [record for record in candidates if normalize_key(record.author) == current_author]
        if len({(item.author, item.output_name) for item in current_author_matches}) == 1:
            return earliest_record(current_author_matches), "matched", ""

    distinct = {(item.author, item.revision_date, item.output_name) for item in candidates}
    if len(distinct) == 1:
        return candidates[0], "matched", ""

    if len({(item.author, item.output_name) for item in candidates}) == 1:
        return earliest_record(candidates), "matched", ""

    return None, "ambiguous", "같은 산출물명에 서로 다른 WBS 담당/시작일 후보가 있습니다."


def earliest_record(records: list[WbsMetadata]) -> WbsMetadata:
    return sorted(records, key=lambda item: (item.revision_date, item.row))[0]


def inspect_document_metadata(path: Path) -> DocumentMetadata:
    suffix = path.suffix.lower()
    try:
        if suffix in EXCEL_DOCUMENT_SUFFIXES:
            return inspect_excel_metadata(path)
        if suffix in ppt_ooxml.PPT_DOCUMENT_SUFFIXES:
            return inspect_ppt_metadata(path)
        if is_hwpx_zip(path):
            return inspect_hwpx_metadata(path)
        return inspect_text_metadata(path)
    except Exception:
        return DocumentMetadata()


def inspect_text_metadata(path: Path) -> DocumentMetadata:
    text = extract_document_text(path)[:6000]
    metadata = DocumentMetadata()
    author_match = re.search(r"<\s*(?:작성자|작\s*성\s*자)\s*>\s*<\s*([^<>]+)\s*>", text)
    date_match = re.search(r"<\s*개정일자\s*>\s*<\s*([^<>]+)\s*>", text)
    if author_match:
        metadata.author = author_match.group(1).strip()
    if date_match:
        metadata.revision_date = date_match.group(1).strip()
    return metadata


def inspect_hwpx_metadata(path: Path) -> DocumentMetadata:
    metadata = DocumentMetadata()
    with zipfile.ZipFile(path, "r") as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".xml"):
                continue
            xml = zf.read(name).decode("utf-8", errors="ignore")
            if not metadata.revision_date:
                metadata.revision_date = find_label_value_in_xml(xml, DATE_LABELS)
            if not metadata.author:
                metadata.author = find_label_value_in_xml(xml, AUTHOR_LABELS)
            header_date, header_author = find_header_metadata_values_in_xml(xml)
            if header_date and not metadata.revision_date:
                metadata.revision_date = header_date
            if header_author and not metadata.author:
                metadata.author = header_author
            revision_date, revision_author = find_revision_history_values_in_xml(xml)
            if revision_date and not metadata.revision_history_date:
                metadata.revision_history_date = revision_date
            if revision_author and not metadata.revision_author:
                metadata.revision_author = revision_author
    return metadata


def find_label_value_in_xml(xml: str, labels: set[str]) -> str:
    normalized_labels = {normalize_label(label) for label in labels}
    for row_match in ROW_PATTERN.finditer(xml):
        cells = list(CELL_PATTERN.finditer(row_match.group(0)))
        if len(cells) > 4:
            continue
        for cell_index, cell_match in enumerate(cells[:-1]):
            label = normalize_label(get_cell_text(cell_match.group(0)).strip())
            if label not in normalized_labels:
                continue
            value = get_cell_text(cells[cell_index + 1].group(0)).strip()
            if value and normalize_label(value) not in {normalize_label(item) for item in LABEL_LIKE_VALUES}:
                return value
    return ""


def find_header_metadata_values_in_xml(xml: str) -> tuple[str, str]:
    for row_match in ROW_PATTERN.finditer(xml):
        cells = list(CELL_PATTERN.finditer(row_match.group(0)))
        if len(cells) < 4:
            continue
        values = [get_cell_text(cell.group(0)).strip() for cell in cells]
        indexes = unlabeled_header_metadata_indexes(values)
        if indexes is None:
            continue
        _code_index, _version_index, date_index, author_index = indexes
        return values[date_index], values[author_index]
    return "", ""


def unlabeled_header_metadata_indexes(values: list[str]) -> tuple[int, int, int, int] | None:
    return shared_unlabeled_header_metadata_indexes(
        values,
        clean_text=clean_text,
        normalize_label=normalize_label,
        label_like_values=LABEL_LIKE_VALUES,
        version_pattern=VERSION_PATTERN,
        date_pattern=DATE_VALUE_PATTERN,
        author_value_ok=looks_like_unlabeled_author_value,
    )


def has_unlabeled_metadata_signal(version_value: str, date_value: str, author_value: str) -> bool:
    return shared_has_unlabeled_metadata_signal(
        version_value,
        date_value,
        author_value,
        clean_text=clean_text,
    )


def looks_like_unlabeled_document_code_slot(value: str) -> bool:
    return unlabeled_header_slot_is_clean(
        value,
        "code",
        clean_text=clean_text,
        normalize_label=normalize_label,
        label_like_values=LABEL_LIKE_VALUES,
        version_pattern=VERSION_PATTERN,
        date_pattern=DATE_VALUE_PATTERN,
    )


def looks_like_unlabeled_version_slot(value: str) -> bool:
    return unlabeled_header_slot_is_clean(
        value,
        "version",
        clean_text=clean_text,
        normalize_label=normalize_label,
        label_like_values=LABEL_LIKE_VALUES,
        version_pattern=VERSION_PATTERN,
        date_pattern=DATE_VALUE_PATTERN,
    )


def looks_like_unlabeled_date_slot(value: str) -> bool:
    return unlabeled_header_slot_is_clean(
        value,
        "date",
        clean_text=clean_text,
        normalize_label=normalize_label,
        label_like_values=LABEL_LIKE_VALUES,
        version_pattern=VERSION_PATTERN,
        date_pattern=DATE_VALUE_PATTERN,
    )


def looks_like_unlabeled_author_slot(value: str) -> bool:
    return unlabeled_header_slot_is_clean(
        value,
        "author",
        clean_text=clean_text,
        normalize_label=normalize_label,
        label_like_values=LABEL_LIKE_VALUES,
        version_pattern=VERSION_PATTERN,
        date_pattern=DATE_VALUE_PATTERN,
        author_value_ok=looks_like_unlabeled_author_value,
    )


def find_revision_history_values_in_xml(xml: str) -> tuple[str, str]:
    for _row_match, cells, header_map in iter_revision_history_data_rows(xml):
        date_idx = header_map.get("date")
        author_idx = header_map.get("author")
        if date_idx is None or author_idx is None:
            continue
        return get_cell_text(cells[date_idx].group(0)).strip(), get_cell_text(cells[author_idx].group(0)).strip()
    return "", ""


def inspect_excel_metadata(path: Path) -> DocumentMetadata:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.lower() in {".xlsm", ".xltm"})
    metadata = DocumentMetadata()
    for sheet in workbook.worksheets:
        if not metadata.revision_date:
            metadata.revision_date = find_label_value_in_sheet(sheet, DATE_LABELS)
        if not metadata.author:
            metadata.author = find_label_value_in_sheet(sheet, AUTHOR_LABELS)
    for sheet in workbook.worksheets:
        if normalize_key(sheet.title).startswith(normalize_key("개정이력")):
            metadata.revision_history_date, metadata.revision_author = find_revision_history_values_in_sheet(sheet)
            break
    return metadata


def inspect_ppt_metadata(path: Path) -> DocumentMetadata:
    metadata = DocumentMetadata()
    if path.suffix.lower() not in ppt_ooxml.PPT_OOXML_SUFFIXES:
        return metadata

    with zipfile.ZipFile(path, "r") as zf:
        for name in zf.namelist():
            if not is_ppt_slide_xml(name):
                continue
            xml = zf.read(name).decode("utf-8", errors="ignore")
            if not metadata.revision_date:
                metadata.revision_date = find_ppt_label_value_in_xml(xml, DATE_LABELS)
            if not metadata.author:
                metadata.author = find_ppt_label_value_in_xml(xml, AUTHOR_LABELS)
            revision_date, revision_author = find_ppt_revision_history_values_in_xml(xml)
            if revision_date and not metadata.revision_history_date:
                metadata.revision_history_date = revision_date
            if revision_author and not metadata.revision_author:
                metadata.revision_author = revision_author
            if metadata.revision_date and metadata.author and metadata.revision_history_date and metadata.revision_author:
                break
    return metadata


def is_ppt_slide_xml(name: str) -> bool:
    lowered = name.lower()
    return lowered.startswith("ppt/slides/slide") and lowered.endswith(".xml")


def find_ppt_label_value_in_xml(xml: str, labels: set[str]) -> str:
    normalized_labels = {normalize_label(label) for label in labels}
    label_values = {normalize_label(item) for item in LABEL_LIKE_VALUES}
    for row_match in ppt_ooxml.TABLE_ROW_PATTERN.finditer(xml):
        cells = list(ppt_ooxml.TABLE_CELL_PATTERN.finditer(row_match.group(0)))
        if len(cells) > 4:
            continue
        for cell_index, cell_match in enumerate(cells[:-1]):
            label = normalize_label(ppt_ooxml.xml_fragment_text(cell_match.group(0)))
            if label not in normalized_labels:
                continue
            value = ppt_ooxml.xml_fragment_text(cells[cell_index + 1].group(0)).strip()
            if value and normalize_label(value) not in label_values:
                return value
    return ""


def find_ppt_revision_history_values_in_xml(xml: str) -> tuple[str, str]:
    for _row_match, cells, header_map in iter_ppt_revision_history_data_rows(xml):
        date_idx = header_map.get("date")
        author_idx = header_map.get("author")
        if date_idx is None or author_idx is None:
            continue
        return (
            ppt_ooxml.xml_fragment_text(cells[date_idx].group(0)).strip(),
            ppt_ooxml.xml_fragment_text(cells[author_idx].group(0)).strip(),
        )
    return "", ""


def find_label_value_in_sheet(sheet: Worksheet, labels: set[str]) -> str:
    normalized_labels = {normalize_label(label) for label in labels}
    for row in sheet.iter_rows():
        for index, cell in enumerate(row[:-1]):
            if normalize_label(clean_text(cell.value)) not in normalized_labels:
                continue
            target = row[index + 1]
            value = clean_text(target.value)
            if value and normalize_label(value) not in {normalize_label(item) for item in LABEL_LIKE_VALUES}:
                return value
    return ""


def find_revision_history_values_in_sheet(sheet: Worksheet) -> tuple[str, str]:
    for row in sheet.iter_rows():
        labels = [normalize_label(clean_text(cell.value)) for cell in row]
        header_map = revision_header_map(labels)
        if not header_map:
            continue
        date_idx = header_map["date"]
        author_idx = header_map["author"]
        for data_row in sheet.iter_rows(min_row=row[0].row + 1, max_col=max(date_idx, author_idx) + 1):
            version = clean_text(data_row[0].value)
            if VERSION_PATTERN.fullmatch(version):
                return clean_text(data_row[date_idx].value), clean_text(data_row[author_idx].value)
    return "", ""


def build_metadata_targets(
    folder_root: Path,
    wbs_records: list[WbsMetadata],
    approval_author: str = "",
) -> list[MetadataTarget]:
    targets: list[MetadataTarget] = []
    for path in collect_metadata_documents(folder_root):
        relative_path = str(path.relative_to(folder_root))
        candidates = records_matching_path(path, wbs_records)
        current = inspect_document_metadata(path)
        record, status, message = choose_record(path, candidates, current)
        if (
            record
            and status == "matched"
            and path.suffix.lower() in ppt_ooxml.PPT_DOCUMENT_SUFFIXES
            and not has_updatable_ppt_metadata(path, record.author, record.revision_date, approval_author)
        ):
            status = "no_change"
            message = "문서 내부에서 수정할 표지/개정이력 위치를 찾지 못했습니다."
        targets.append(
            MetadataTarget(
                path=path,
                relative_path=relative_path,
                status=status,
                message=message,
                output_name=record.output_name if record else (candidates[0].output_name if candidates else ""),
                author=record.author if record else "",
                revision_date=record.revision_date if record else "",
                current=current,
                candidates=candidates,
            )
        )
    return targets


def has_updatable_ppt_metadata(path: Path, author: str, revision_date: str, approval_author: str) -> bool:
    if path.suffix.lower() not in ppt_ooxml.PPT_OOXML_SUFFIXES:
        return False

    try:
        with zipfile.ZipFile(path, "r") as zf:
            for name in zf.namelist():
                if not is_ppt_slide_xml(name):
                    continue
                xml = zf.read(name).decode("utf-8", errors="ignore")
                _xml, cover_count = update_ppt_label_cells_xml(
                    xml,
                    {
                        **{label: revision_date for label in DATE_LABELS},
                        **{label: author for label in AUTHOR_LABELS},
                        **{label: approval_author for label in APPROVAL_LABELS},
                    },
                )
                _xml, revision_count = update_ppt_revision_history_xml(
                    xml,
                    revision_date,
                    author,
                    approval_author,
                )
                if cover_count or revision_count:
                    return True
    except Exception:
        return False
    return False


def update_metadata_in_document(
    file_path: Path,
    author: str,
    revision_date: str,
    approval_author: str,
    temp_dir: Path,
    document_number: str = "",
) -> MetadataWriteResult:
    try:
        target_file, converted_to_hwpx = prepare_target_file(file_path, temp_dir)
        output_path = target_file
        if is_hwpx_zip(target_file):
            result = write_updated_hwpx_metadata(target_file, author, revision_date, approval_author, document_number)
        elif target_file.suffix.lower() in EXCEL_DOCUMENT_SUFFIXES:
            result = write_updated_excel_metadata(target_file, author, revision_date, approval_author, document_number)
        elif target_file.suffix.lower() in ppt_ooxml.PPT_DOCUMENT_SUFFIXES:
            result = write_updated_ppt_metadata(target_file, author, revision_date, approval_author, document_number)
        else:
            raise RuntimeError("지원하지 않는 문서 형식입니다.")

        if converted_to_hwpx:
            output_path = file_path.with_suffix(".hwpx")
            shutil.move(str(target_file), str(output_path))
            if output_path != file_path and file_path.exists():
                file_path.unlink()

        no_change_ppt = target_file.suffix.lower() in ppt_ooxml.PPT_DOCUMENT_SUFFIXES and not sum(result)
        return MetadataWriteResult(
            status="skipped" if no_change_ppt else "updated",
            old_path=file_path,
            new_path=output_path,
            backup_path=None,
            converted_to_hwpx=converted_to_hwpx,
            cover_update_count=result[0],
            revision_history_update_count=result[1],
            error=(
                "문서 내부에서 수정할 표지/개정이력 위치를 찾지 못했습니다."
                if no_change_ppt
                else ""
            ),
        )
    except Exception as exc:
        return MetadataWriteResult(
            status="error",
            old_path=file_path,
            backup_path=None,
            error=str(exc),
        )


def metadata_temp_path(path: Path) -> Path:
    return path.with_name(f"{path.stem}_metadata_{uuid4().hex}{path.suffix}")


def safe_unlink(path: Path) -> None:
    try:
        path.unlink()
    except OSError:
        pass


def replace_with_retry(source: Path, target: Path, *, attempts: int = 10, delay_seconds: float = 0.2) -> None:
    last_error: OSError | None = None
    for _ in range(attempts):
        try:
            source.replace(target)
            return
        except OSError as exc:
            last_error = exc
            try:
                shutil.copyfile(source, target)
                safe_unlink(source)
                return
            except OSError as copy_exc:
                last_error = copy_exc
            time.sleep(delay_seconds)
    if last_error is not None:
        raise last_error


def write_updated_excel_metadata(
    path: Path,
    author: str,
    revision_date: str,
    approval_author: str,
    document_number: str = "",
) -> tuple[int, int]:
    revision_date = normalize_revision_date_for_write(revision_date)
    cover_count = 0
    revision_count = 0
    temp_path = metadata_temp_path(path)
    try:
        with zipfile.ZipFile(path, "r") as zin:
            shared_strings = read_shared_strings(zin)
            sheet_map = {sheet.path: sheet.name for sheet in workbook_sheets(zin)}
            updates_by_sheet: dict[str, tuple[bytes, int, int]] = {}

            for sheet_path, sheet_name in sheet_map.items():
                original_xml = zin.read(sheet_path).decode("utf-8", errors="ignore")
                is_revision_history_sheet = normalize_key(sheet_name).startswith(normalize_key("개정이력"))
                label_count = 0
                updated_xml = original_xml
                if not is_revision_history_sheet:
                    label_values = {
                        **{label: revision_date for label in DATE_LABELS},
                        **{label: author for label in AUTHOR_LABELS},
                        **{label: DOCUMENT_VERSION_VALUE for label in VERSION_LABELS},
                    }
                    if document_number:
                        label_values.update({label: document_number for label in DOCUMENT_NUMBER_LABELS})
                    updated_xml, label_count = update_excel_label_cells_xml(
                        original_xml,
                        shared_strings,
                        label_values,
                    )
                    header_count = 0
                    updated_xml, header_count = update_unlabeled_excel_header_metadata_xml(
                        updated_xml,
                        shared_strings,
                        revision_date,
                        author,
                        document_number,
                    )
                    label_count += header_count
                sheet_revision_count = 0
                if is_revision_history_sheet:
                    updated_xml, sheet_revision_count = update_excel_revision_history_xml(
                        updated_xml,
                        shared_strings,
                        revision_date,
                        author,
                        approval_author,
                    )
                view_xml = set_excel_sheet_selected_state(
                    updated_xml,
                    selected=sheet_path == next(iter(sheet_map), ""),
                )
                if view_xml != updated_xml:
                    updated_xml = view_xml
                if label_count or sheet_revision_count:
                    updates_by_sheet[sheet_path] = (
                        updated_xml.encode("utf-8"),
                        label_count,
                        sheet_revision_count,
                    )
                    cover_count += label_count
                    revision_count += sheet_revision_count
                elif updated_xml != original_xml:
                    updates_by_sheet[sheet_path] = (
                        updated_xml.encode("utf-8"),
                        0,
                        0,
                    )

            for item in zin.infolist():
                if not item.filename.startswith("xl/drawings/") or not item.filename.endswith(".xml"):
                    continue
                original_xml = zin.read(item.filename).decode("utf-8", errors="ignore")
                updated_xml, drawing_count = update_excel_drawing_metadata_xml(
                    original_xml,
                    revision_date,
                    author,
                )
                if drawing_count:
                    updates_by_sheet[item.filename] = (
                        updated_xml.encode("utf-8"),
                        drawing_count,
                        0,
                    )
                    cover_count += drawing_count

            workbook_modified = bool(updates_by_sheet)
            with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename in updates_by_sheet:
                        data = updates_by_sheet[item.filename][0]
                    elif item.filename == "xl/workbook.xml":
                        data = set_excel_first_sheet_active(data.decode("utf-8", errors="ignore")).encode("utf-8")
                    if workbook_modified:
                        rewritten = rewrite_excel_part_for_modified_workbook(item.filename, data)
                        if rewritten is None:
                            continue
                        data = rewritten
                    zout.writestr(item, data)
        replace_with_retry(temp_path, path)
    except Exception:
        if temp_path.exists():
            safe_unlink(temp_path)
        raise
    return cover_count, revision_count


def write_updated_ppt_metadata(
    path: Path,
    author: str,
    revision_date: str,
    approval_author: str,
    document_number: str = "",
) -> tuple[int, int]:
    if path.suffix.lower() not in ppt_ooxml.PPT_OOXML_SUFFIXES:
        return 0, 0

    revision_date = normalize_revision_date_for_write(revision_date)
    cover_count = 0
    revision_count = 0
    temp_path = metadata_temp_path(path)
    try:
        with zipfile.ZipFile(path, "r") as zin:
            with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if is_ppt_slide_xml(item.filename):
                        xml = data.decode("utf-8", errors="ignore")
                        changed_count = 0
                        label_values = {
                            **{label: revision_date for label in DATE_LABELS},
                            **{label: author for label in AUTHOR_LABELS},
                            **{label: approval_author for label in APPROVAL_LABELS},
                            **{label: DOCUMENT_VERSION_VALUE for label in VERSION_LABELS},
                        }
                        if document_number:
                            label_values.update({label: document_number for label in DOCUMENT_NUMBER_LABELS})
                        xml, count = update_ppt_label_cells_xml(
                            xml,
                            label_values,
                        )
                        cover_count += count
                        changed_count += count
                        xml, count = update_ppt_revision_history_xml(
                            xml,
                            revision_date,
                            author,
                            approval_author,
                        )
                        revision_count += count
                        changed_count += count
                        if changed_count:
                            data = xml.encode("utf-8")
                    zout.writestr(item, data)
        if cover_count or revision_count:
            replace_with_retry(temp_path, path)
        else:
            temp_path.unlink(missing_ok=True)
    except Exception:
        if temp_path.exists():
            safe_unlink(temp_path)
        raise
    return cover_count, revision_count


def update_ppt_label_cells_xml(
    xml: str,
    values_by_label: dict[str, str],
) -> tuple[str, int]:
    normalized_values = {normalize_label(label): value for label, value in values_by_label.items()}
    label_values = {normalize_label(item) for item in LABEL_LIKE_VALUES}
    pieces: list[str] = []
    last = 0
    count = 0

    for row_match in ppt_ooxml.TABLE_ROW_PATTERN.finditer(xml):
        row_xml = row_match.group(0)
        cells = list(ppt_ooxml.TABLE_CELL_PATTERN.finditer(row_xml))
        updated_row = row_xml
        offset = 0
        row_changed = False
        for cell_index, cell_match in enumerate(cells[:-1]):
            label = normalize_label(ppt_ooxml.xml_fragment_text(cell_match.group(0)))
            if label not in normalized_values:
                continue
            target_cell = cells[cell_index + 1]
            target_text = normalize_label(ppt_ooxml.xml_fragment_text(target_cell.group(0)))
            if target_text in label_values:
                continue
            updated_cell, _old_value = ppt_ooxml.replace_cell_text(
                target_cell.group(0),
                normalized_values[label],
            )
            start = target_cell.start() + offset
            end = target_cell.end() + offset
            updated_row = updated_row[:start] + updated_cell + updated_row[end:]
            offset += len(updated_cell) - len(target_cell.group(0))
            row_changed = True
            count += 1

        if row_changed:
            pieces.append(xml[last:row_match.start()])
            pieces.append(updated_row)
            last = row_match.end()

    if not pieces:
        return xml, 0
    pieces.append(xml[last:])
    return "".join(pieces), count


def iter_ppt_revision_history_data_rows(xml: str):
    rows = list(ppt_ooxml.TABLE_ROW_PATTERN.finditer(xml))
    for header_index, row_match in enumerate(rows):
        cells = list(ppt_ooxml.TABLE_CELL_PATTERN.finditer(row_match.group(0)))
        labels = [normalize_label(ppt_ooxml.xml_fragment_text(cell.group(0))) for cell in cells]
        header_map = revision_header_map(labels)
        if not header_map:
            continue
        for data_row_match in rows[header_index + 1:]:
            data_cells = list(ppt_ooxml.TABLE_CELL_PATTERN.finditer(data_row_match.group(0)))
            if is_ppt_revision_history_data_cells(data_cells, header_map):
                yield data_row_match, data_cells, header_map
                break


def is_ppt_revision_history_data_cells(cells: list[re.Match[str]], header_map: dict[str, int]) -> bool:
    if len(cells) <= max(header_map.values()):
        return False
    labels = [normalize_label(ppt_ooxml.xml_fragment_text(cell.group(0))) for cell in cells]
    return not revision_header_map(labels)


def update_ppt_revision_history_xml(
    xml: str,
    revision_date: str,
    author: str,
    approval_author: str,
) -> tuple[str, int]:
    revision_date = normalize_revision_date_for_write(revision_date)
    rows = list(ppt_ooxml.TABLE_ROW_PATTERN.finditer(xml))
    for header_index, row_match in enumerate(rows):
        cells = list(ppt_ooxml.TABLE_CELL_PATTERN.finditer(row_match.group(0)))
        labels = [normalize_label(ppt_ooxml.xml_fragment_text(cell.group(0))) for cell in cells]
        header_map = revision_header_map(labels)
        if not header_map:
            continue

        pieces: list[str] = []
        last = 0
        count = 0
        found_first_data_row = False
        for data_row_match in rows[header_index + 1:]:
            data_cells = list(ppt_ooxml.TABLE_CELL_PATTERN.finditer(data_row_match.group(0)))
            if not is_ppt_revision_history_data_cells(data_cells, header_map):
                continue

            if found_first_data_row:
                updated_row, row_count = clear_ppt_revision_history_data_row(data_row_match.group(0), data_cells)
            else:
                updates = [
                    (header_map["version"], "0.1"),
                    (header_map["date"], revision_date),
                    (header_map["reason"], REVISION_REASON_VALUE) if "reason" in header_map else None,
                    (header_map["detail"], REVISION_DETAIL_VALUE) if "detail" in header_map else None,
                    (header_map["author"], author),
                ]
                if "approval" in header_map:
                    updates.append((header_map["approval"], approval_author))
                updates = [item for item in updates if item is not None]
                updated_row, row_count = update_ppt_revision_history_first_row(
                    data_row_match.group(0),
                    data_cells,
                    updates,
                )
                found_first_data_row = True

            if row_count:
                pieces.append(xml[last:data_row_match.start()])
                pieces.append(updated_row)
                last = data_row_match.end()
                count += row_count

        if not count:
            return xml, 0
        pieces.append(xml[last:])
        return "".join(pieces), count
    return xml, 0


def update_ppt_revision_history_first_row(
    row_xml: str,
    cells: list[re.Match[str]],
    updates: list[tuple[int, str]],
) -> tuple[str, int]:
    offset = 0
    updated_row = row_xml
    count = 0
    for cell_index, new_text in updates:
        cell = cells[cell_index]
        updated_cell, _old_value = ppt_ooxml.replace_cell_text(cell.group(0), new_text)
        start = cell.start() + offset
        end = cell.end() + offset
        updated_row = updated_row[:start] + updated_cell + updated_row[end:]
        offset += len(updated_cell) - len(cell.group(0))
        count += 1
    return updated_row, count


def clear_ppt_revision_history_data_row(
    row_xml: str,
    cells: list[re.Match[str]],
) -> tuple[str, int]:
    offset = 0
    updated_row = row_xml
    count = 0
    for cell in cells:
        if not ppt_ooxml.xml_fragment_text(cell.group(0)).strip():
            continue
        updated_cell, _old_value = ppt_ooxml.replace_cell_text(cell.group(0), "")
        start = cell.start() + offset
        end = cell.end() + offset
        updated_row = updated_row[:start] + updated_cell + updated_row[end:]
        offset += len(updated_cell) - len(cell.group(0))
        count += 1
    return updated_row, count


def set_excel_first_sheet_active(workbook_xml: str) -> str:
    def update_workbook_view(match: re.Match[str]) -> str:
        tag = match.group(0)
        tag = upsert_xml_attribute(tag, "activeTab", "0")
        tag = upsert_xml_attribute(tag, "firstSheet", "0")
        return tag

    updated_xml, count = re.subn(
        r"<(?:\w+:)?workbookView\b[^>]*/?>",
        update_workbook_view,
        workbook_xml,
        count=1,
    )
    if count:
        return updated_xml

    insert_at = workbook_xml.find(">")
    if insert_at < 0:
        return workbook_xml
    return (
        workbook_xml[: insert_at + 1]
        + '<bookViews><workbookView activeTab="0" firstSheet="0"/></bookViews>'
        + workbook_xml[insert_at + 1:]
    )


def set_excel_sheet_selected_state(sheet_xml: str, selected: bool) -> str:
    def update_sheet_view(match: re.Match[str]) -> str:
        tag = match.group(0)
        if selected:
            return upsert_xml_attribute(tag, "tabSelected", "1")
        return remove_xml_attribute(tag, "tabSelected")

    updated_xml, count = re.subn(
        r"<(?:\w+:)?sheetView\b[^>]*/?>",
        update_sheet_view,
        sheet_xml,
        count=1,
    )
    if count:
        return updated_xml

    insert_at = sheet_xml.find(">")
    if insert_at < 0:
        return sheet_xml
    tab_selected = ' tabSelected="1"' if selected else ""
    return (
        sheet_xml[: insert_at + 1]
        + f'<sheetViews><sheetView workbookViewId="0"{tab_selected}/></sheetViews>'
        + sheet_xml[insert_at + 1:]
    )


def upsert_xml_attribute(tag: str, name: str, value: str) -> str:
    updated, count = re.subn(rf'\s{name}="[^"]*"', f' {name}="{value}"', tag, count=1)
    if count:
        return updated
    insert_at = -2 if tag.endswith("/>") else -1
    return f"{tag[:insert_at]} {name}=\"{value}\"{tag[insert_at:]}"


def remove_xml_attribute(tag: str, name: str) -> str:
    return re.sub(rf'\s{name}="[^"]*"', "", tag, count=1)


def update_excel_label_cells_xml(
    sheet_xml: str,
    shared_strings: list[str],
    values_by_label: dict[str, str],
) -> tuple[str, int]:
    import xml.etree.ElementTree as ET

    root = ET.fromstring(sheet_xml)
    cells = iter_cells(root, shared_strings)
    cells_by_position = {(cell.row, cell.col): cell for cell in cells}
    normalized_values = {normalize_label(label): value for label, value in values_by_label.items()}
    label_values = {normalize_label(item) for item in LABEL_LIKE_VALUES}
    updates: dict[str, tuple[int, int, str]] = {}

    for info in cells:
        label = normalize_label(clean_text(info.text))
        if label not in normalized_values:
            continue
        target_col = info.col + 1
        target = cells_by_position.get((info.row, target_col))
        if target and normalize_label(clean_text(target.text)) in label_values:
            continue
        target_ref = f"{col_to_name(target_col)}{info.row}"
        updates[target_ref] = (info.row, target_col, normalized_values[label])

    updated_xml = sheet_xml
    count = 0
    for cell_ref, (row, col, value) in updates.items():
        next_xml = replace_or_insert_cell_xml(updated_xml, cell_ref, row, col, value)
        if next_xml == updated_xml:
            continue
        updated_xml = next_xml
        count += 1
    return updated_xml, count


def update_unlabeled_excel_header_metadata_xml(
    sheet_xml: str,
    shared_strings: list[str],
    revision_date: str,
    author: str,
    document_number: str = "",
) -> tuple[str, int]:
    import xml.etree.ElementTree as ET

    revision_date = normalize_revision_date_for_write(revision_date)
    root = ET.fromstring(sheet_xml)
    rows: dict[int, list[CellInfo]] = {}
    for cell in iter_cells(root, shared_strings):
        if 1 <= cell.row <= EXCEL_HEADER_METADATA_SCAN_MAX_ROW:
            rows.setdefault(cell.row, []).append(cell)

    updated_xml = sheet_xml
    count = 0
    metadata_labels = {
        normalize_label(label)
        for label in DATE_LABELS | AUTHOR_LABELS | VERSION_LABELS | APPROVAL_LABELS | LABEL_LIKE_VALUES
    }

    for row_index in sorted(rows):
        row_cells = sorted(rows[row_index], key=lambda item: item.col)
        row_values = [clean_text(cell.text) for cell in row_cells]
        normalized_values = {normalize_label(value) for value in row_values if value}
        if normalized_values & metadata_labels:
            continue

        updates = unlabeled_excel_header_metadata_updates(row_cells, revision_date, author, document_number)
        if not updates:
            continue

        for cell_ref, row, col, value in updates:
            next_xml = replace_or_insert_cell_xml(updated_xml, cell_ref, row, col, value)
            if next_xml == updated_xml:
                continue
            updated_xml = next_xml
            count += 1
        break

    return updated_xml, count


def unlabeled_excel_header_metadata_updates(
    row_cells: list[CellInfo],
    revision_date: str,
    author: str,
    document_number: str = "",
) -> list[tuple[str, int, int, str]]:
    by_col = {cell.col: cell for cell in row_cells}
    for version_cell in row_cells:
        version_value = clean_text(version_cell.text)
        if not VERSION_PATTERN.fullmatch(version_value):
            continue
        date_col = version_cell.col + 1
        author_col = version_cell.col + 2
        if not unlabeled_excel_slot_is_clean(by_col.get(version_cell.col - 1), "code"):
            continue
        if not unlabeled_excel_slot_is_clean(by_col.get(date_col), "date"):
            continue
        if not unlabeled_excel_slot_is_clean(by_col.get(author_col), "author"):
            continue
        updates = [
            (version_cell.ref, version_cell.row, version_cell.col, UNLABELED_HEADER_VERSION_VALUE),
            (
                f"{col_to_name(date_col)}{version_cell.row}",
                version_cell.row,
                date_col,
                format_date_like_existing(revision_date, by_col.get(date_col).text if by_col.get(date_col) else ""),
            ),
            (f"{col_to_name(author_col)}{version_cell.row}", version_cell.row, author_col, author),
        ]
        if document_number and version_cell.col > 1:
            code_col = version_cell.col - 1
            updates.insert(0, (f"{col_to_name(code_col)}{version_cell.row}", version_cell.row, code_col, document_number))
        return updates

    for date_cell in row_cells:
        date_value = clean_text(date_cell.text)
        if not DATE_VALUE_PATTERN.fullmatch(date_value):
            continue
        version_col = date_cell.col - 1
        author_col = date_cell.col + 1
        if version_col < 1:
            continue
        if not unlabeled_excel_slot_is_clean(by_col.get(version_col - 1), "code"):
            continue
        if not unlabeled_excel_slot_is_clean(by_col.get(version_col), "version"):
            continue
        if not unlabeled_excel_slot_is_clean(by_col.get(author_col), "author"):
            continue
        updates = [
            (f"{col_to_name(version_col)}{date_cell.row}", date_cell.row, version_col, UNLABELED_HEADER_VERSION_VALUE),
            (date_cell.ref, date_cell.row, date_cell.col, format_date_like_existing(revision_date, date_cell.text)),
            (f"{col_to_name(author_col)}{date_cell.row}", date_cell.row, author_col, author),
        ]
        if document_number and version_col > 1:
            code_col = version_col - 1
            updates.insert(0, (f"{col_to_name(code_col)}{date_cell.row}", date_cell.row, code_col, document_number))
        return updates

    return []


def unlabeled_excel_slot_is_clean(cell: CellInfo | None, slot: str) -> bool:
    if cell is None:
        return True
    value = clean_text(cell.text)
    if slot == "code":
        return looks_like_unlabeled_document_code_slot(value)
    if slot == "version":
        return looks_like_unlabeled_version_slot(value)
    if slot == "date":
        return looks_like_unlabeled_date_slot(value)
    if slot == "author":
        return looks_like_unlabeled_author_slot(value)
    return False


def update_excel_drawing_metadata_xml(xml: str, revision_date: str, author: str) -> tuple[str, int]:
    text_matches = list(DRAWING_TEXT_PATTERN.finditer(xml))
    if not text_matches:
        return xml, 0

    texts = [unescape(match.group("text")) for match in text_matches]
    combined = "".join(texts)
    if not has_embedded_output_id(combined):
        return xml, 0

    spans: list[tuple[int, int]] = []
    cursor = 0
    for text in texts:
        start = cursor
        cursor += len(text)
        spans.append((start, cursor))

    date_match = select_excel_drawing_metadata_date(combined)
    if date_match is None:
        return xml, 0
    version_match = select_excel_drawing_metadata_version(combined, date_match.start())

    date_start, date_end = date_match.span()
    date_run_indexes = [
        index
        for index, (start, end) in enumerate(spans)
        if start < date_end and end > date_start
    ]
    if not date_run_indexes:
        return xml, 0

    author_run_index = next(
        (
            index
            for index, (start, _end) in enumerate(spans)
            if start >= date_end and looks_like_unlabeled_author_value(texts[index])
        ),
        None,
    )
    if author_run_index is None:
        return xml, 0

    new_texts: dict[int, str] = {}
    if version_match is not None:
        add_drawing_text_span_update(new_texts, texts, spans, version_match.span(), UNLABELED_HEADER_VERSION_VALUE)
    add_drawing_text_span_update(new_texts, texts, spans, (date_start, date_end), revision_date)
    new_texts[author_run_index] = author

    pieces: list[str] = []
    last = 0
    for index, match in enumerate(text_matches):
        if index not in new_texts:
            continue
        pieces.append(xml[last:match.start("text")])
        pieces.append(escape(new_texts[index]))
        last = match.end("text")
    pieces.append(xml[last:])
    return "".join(pieces), len(new_texts)


def add_drawing_text_span_update(
    new_texts: dict[int, str],
    texts: list[str],
    spans: list[tuple[int, int]],
    target_span: tuple[int, int],
    replacement: str,
) -> None:
    target_start, target_end = target_span
    run_indexes = [
        index
        for index, (start, end) in enumerate(spans)
        if start < target_end and end > target_start
    ]
    for index in run_indexes:
        run_start, run_end = spans[index]
        text = texts[index]
        prefix = text[: max(0, target_start - run_start)] if index == run_indexes[0] else ""
        suffix = text[max(0, target_end - run_start):] if index == run_indexes[-1] and target_end < run_end else ""
        new_texts[index] = f"{prefix}{replacement if index == run_indexes[0] else ''}{suffix}"


def select_excel_drawing_metadata_version(text: str, before_index: int) -> re.Match[str] | None:
    prefix = text[:before_index]
    candidates = [
        match
        for match in re.finditer(r"[vV]?\d+(?:\.\d+)+", prefix)
        if has_embedded_output_id(prefix[max(0, match.start() - 200):match.start()])
    ]
    return candidates[-1] if candidates else None


def select_excel_drawing_metadata_date(text: str) -> re.Match[str] | None:
    for match in re.finditer(r"\d{4}[-./]\d{1,2}[-./]\d{1,2}", text):
        prefix = text[:match.start()]
        if not has_embedded_output_id(prefix[-200:]):
            continue
        if not re.search(r"[vV]?\d+(?:\.\d+)+\s*$", prefix):
            continue
        return match
    return None


def has_embedded_output_id(text: str) -> bool:
    return bool(
        OUTPUT_ID_PATTERN.search(text)
        or NUMBER_OUTPUT_ID_PATTERN.search(text)
        or re.search(r"(?:MFDS-\d{3,6}|[A-Za-z]{2,10}(?:-[A-Za-z0-9]{1,12})*-\d{2})", text)
    )


def looks_like_unlabeled_author_value(value: str) -> bool:
    text = clean_text(value)
    if not unlabeled_author_value_ok(
        text,
        clean_text=clean_text,
        normalize_label=normalize_label,
        label_like_values=LABEL_LIKE_VALUES,
        version_pattern=VERSION_PATTERN,
        date_pattern=DATE_VALUE_PATTERN,
    ):
        return False
    if OUTPUT_ID_PATTERN.search(text) or NUMBER_OUTPUT_ID_PATTERN.search(text):
        return False
    if re.search(r"\d", text):
        return False
    return 2 <= len(text) <= 20


def update_excel_revision_history_xml(
    sheet_xml: str,
    shared_strings: list[str],
    revision_date: str,
    author: str,
    approval_author: str,
) -> tuple[str, int]:
    import xml.etree.ElementTree as ET

    revision_date = normalize_revision_date_for_write(revision_date)
    root = ET.fromstring(sheet_xml)
    cells = iter_cells(root, shared_strings)
    rows: dict[int, list[CellInfo]] = {}
    for cell in cells:
        rows.setdefault(cell.row, []).append(cell)

    for row_index in sorted(rows):
        row_cells = sorted(rows[row_index], key=lambda item: item.col)
        labels = [normalize_label(clean_text(cell.text)) for cell in row_cells]
        header_by_key = revision_header_map(labels)
        if not header_by_key:
            continue
        header_by_col = {
            key: row_cells[index].col
            for key, index in header_by_key.items()
        }
        first_data_row_index: int | None = None
        updated_xml = sheet_xml
        count = 0
        for data_row_index in sorted(index for index in rows if index > row_index):
            data_cells = {cell.col: cell for cell in rows[data_row_index]}
            if not is_excel_revision_history_data_row(data_cells, header_by_col):
                continue
            if first_data_row_index is None:
                first_data_row_index = data_row_index
                updates = {
                    "version": "0.1",
                    "date": revision_date,
                    "reason": REVISION_REASON_VALUE,
                    "detail": REVISION_DETAIL_VALUE,
                    "author": author,
                    "approval": approval_author,
                }
                for key, value in updates.items():
                    col = header_by_col.get(key)
                    if col is None:
                        continue
                    next_xml = replace_or_insert_cell_xml(
                        updated_xml,
                        f"{col_to_name(col)}{data_row_index}",
                        data_row_index,
                        col,
                        value,
                    )
                    if next_xml == updated_xml:
                        continue
                    updated_xml = next_xml
                    count += 1
                continue

            min_col = min(header_by_col.values())
            max_col = max(header_by_col.values())
            for cell in sorted(data_cells.values(), key=lambda item: item.col):
                if min_col <= cell.col <= max_col and clean_text(cell.text):
                    next_xml = replace_or_insert_cell_xml(
                        updated_xml,
                        cell.ref,
                        data_row_index,
                        cell.col,
                        "",
                    )
                    if next_xml == updated_xml:
                        continue
                    updated_xml = next_xml
                    count += 1
        return updated_xml, count
    return sheet_xml, 0


def is_excel_revision_history_data_row(data_cells: dict[int, CellInfo], header_by_col: dict[str, int]) -> bool:
    min_col = min(header_by_col.values())
    max_col = max(header_by_col.values())
    return any(min_col <= col <= max_col for col in data_cells)


def revision_header_map(labels: list[str]) -> dict[str, int]:
    normalized_versions = {normalize_label(label) for label in VERSION_LABELS} | {normalize_label("버전")}
    normalized_revision_dates = {normalize_label(label) for label in REVISION_DATE_HEADER_LABELS}
    normalized_authors = {normalize_label(label) for label in AUTHOR_LABELS}
    normalized_reasons = {normalize_label(label) for label in REVISION_REASON_HEADER_LABELS}
    normalized_details = {normalize_label(label) for label in REVISION_DETAIL_HEADER_LABELS}
    normalized_approvals = {normalize_label(label) for label in APPROVAL_LABELS}
    result: dict[str, int] = {}
    for index, label in enumerate(labels):
        if normalized_label_matches(label, normalized_versions):
            result["version"] = index
        elif normalized_label_matches(label, normalized_revision_dates):
            result["date"] = index
        elif normalized_label_matches(label, normalized_reasons):
            result["reason"] = index
        elif normalized_label_matches(label, normalized_details):
            result["detail"] = index
        elif (
            "version" in result
            and "date" not in result
            and index == result["version"] + 1
            and VERSION_PATTERN.fullmatch(label)
        ):
            result["date"] = index
        elif normalized_label_matches(label, normalized_authors):
            result["author"] = index
        elif normalized_label_matches(label, normalized_approvals):
            result["approval"] = index
    required = {"version", "date", "author"}
    return result if required.issubset(result) else {}


def normalized_label_matches(label: str, expected_labels: set[str]) -> bool:
    if not label:
        return False
    return any(expected and (label == expected or expected in label) for expected in expected_labels)


def write_updated_hwpx_metadata(
    path: Path,
    author: str,
    revision_date: str,
    approval_author: str,
    document_number: str = "",
) -> tuple[int, int]:
    revision_date = normalize_revision_date_for_write(revision_date)
    temp_path = metadata_temp_path(path)
    cover_count = 0
    revision_count = 0
    try:
        with zipfile.ZipFile(path, "r") as zin:
            with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.lower().endswith(".xml"):
                        xml = data.decode("utf-8", errors="ignore")
                        editable_xml, preserved_xml = editable_hwpx_part_scope(item.filename, xml)
                        if not editable_xml:
                            zout.writestr(item, data)
                            continue
                        changed_count = 0
                        editable_xml, count = update_label_right_rows_in_xml(editable_xml, DATE_LABELS, revision_date)
                        cover_count += count
                        changed_count += count
                        editable_xml, count = update_label_right_rows_in_xml(editable_xml, AUTHOR_LABELS, author)
                        cover_count += count
                        changed_count += count
                        editable_xml, count = update_label_right_rows_in_xml(editable_xml, VERSION_LABELS, DOCUMENT_VERSION_VALUE)
                        cover_count += count
                        changed_count += count
                        if document_number:
                            editable_xml, count = update_label_right_rows_in_xml(editable_xml, DOCUMENT_NUMBER_LABELS, document_number)
                            cover_count += count
                            changed_count += count
                        editable_xml, count = update_unlabeled_metadata_rows_xml(editable_xml, revision_date, author, document_number)
                        cover_count += count
                        changed_count += count
                        editable_xml, count = update_revision_history_xml(editable_xml, revision_date, author, approval_author)
                        revision_count += count
                        changed_count += count
                        if changed_count:
                            editable_xml, _line_seg_count = strip_hwpx_line_seg_arrays(editable_xml)
                            data = (editable_xml + preserved_xml).encode("utf-8")
                        elif item.filename.lower() == "settings.xml":
                            xml, view_count = reset_hwpx_open_position(xml)
                            if view_count:
                                data = xml.encode("utf-8")
                    zout.writestr(item, data)
        if cover_count or revision_count:
            replace_with_retry(temp_path, path)
        else:
            temp_path.unlink(missing_ok=True)
    except Exception:
        if temp_path.exists():
            safe_unlink(temp_path)
        raise
    return cover_count, revision_count


def update_label_right_rows_in_xml(xml: str, labels: set[str], new_text: str) -> tuple[str, int]:
    normalized_labels = {normalize_label(label) for label in labels}
    pieces: list[str] = []
    last = 0
    count = 0

    for row_match in ROW_PATTERN.finditer(xml):
        row_xml = row_match.group(0)
        cells = list(CELL_PATTERN.finditer(row_xml))
        if not cells or len(cells) > 4:
            continue
        updated_row = row_xml
        offset = 0
        changed = False
        for cell_index, cell_match in enumerate(cells[:-1]):
            label = normalize_label(get_cell_text(cell_match.group(0)).strip())
            if label not in normalized_labels:
                continue
            target_cell = cells[cell_index + 1]
            target_text = normalize_label(get_cell_text(target_cell.group(0)).strip())
            if target_text in {normalize_label(item) for item in LABEL_LIKE_VALUES}:
                continue
            updated_cell, _old_value = replace_cell_text(target_cell.group(0), new_text)
            start = target_cell.start() + offset
            end = target_cell.end() + offset
            updated_row = updated_row[:start] + updated_cell + updated_row[end:]
            offset += len(updated_cell) - len(target_cell.group(0))
            changed = True
            count += 1
        if changed:
            pieces.append(xml[last:row_match.start()])
            pieces.append(updated_row)
            last = row_match.end()

    if not pieces:
        return xml, 0
    pieces.append(xml[last:])
    return "".join(pieces), count


def reset_hwpx_open_position(xml: str) -> tuple[str, int]:
    return HWPX_CARET_POSITION_PATTERN.subn("", xml)


def update_unlabeled_metadata_rows_xml(
    xml: str,
    revision_date: str,
    author: str,
    document_number: str = "",
) -> tuple[str, int]:
    pieces: list[str] = []
    last = 0
    count = 0

    for row_match in ROW_PATTERN.finditer(xml):
        updated_row, row_count = update_unlabeled_metadata_row(
            row_match.group(0),
            revision_date,
            author,
            document_number,
        )
        if not row_count:
            continue
        pieces.append(xml[last:row_match.start()])
        pieces.append(updated_row)
        last = row_match.end()
        count += row_count

    if not pieces:
        return xml, 0
    pieces.append(xml[last:])
    return "".join(pieces), count


def update_unlabeled_metadata_row(
    row_xml: str,
    revision_date: str,
    author: str,
    document_number: str = "",
) -> tuple[str, int]:
    revision_date = normalize_revision_date_for_write(revision_date)
    cells = list(CELL_PATTERN.finditer(row_xml))
    if len(cells) < 4:
        return row_xml, 0

    values = [get_cell_text(cell.group(0)).strip() for cell in cells]
    indexes = unlabeled_header_metadata_indexes(values)
    if indexes is None:
        return row_xml, 0
    code_index, version_index, date_index, author_index = indexes

    updates: list[tuple[int, str]] = [
        (version_index, UNLABELED_HEADER_VERSION_VALUE),
        (date_index, format_date_like_existing(revision_date, values[date_index])),
        (author_index, author),
    ]
    if document_number:
        updates.insert(0, (code_index, document_number))
    updated_row = row_xml
    offset = 0
    count = 0
    for cell_index, new_text in updates:
        cell = cells[cell_index]
        old_text = get_cell_text(cell.group(0)).strip()
        if old_text == new_text:
            continue
        updated_cell, _old_value = replace_cell_text(cell.group(0), new_text)
        start = cell.start() + offset
        end = cell.end() + offset
        updated_row = updated_row[:start] + updated_cell + updated_row[end:]
        offset += len(updated_cell) - len(cell.group(0))
        count += 1
    return updated_row, count


def format_date_like_existing(revision_date: str, existing_date: str) -> str:
    del existing_date
    return normalize_revision_date_for_write(revision_date)


def normalize_revision_date_for_write(revision_date: str) -> str:
    text = clean_text(str(revision_date or ""))
    if not text:
        return ""
    if text.isdigit() and len(text) == 4:
        return f"{text}-00-00"
    date_match = re.fullmatch(r"(\d{4})(?:[-./\s]+(\d{1,2})[-./\s]+(\d{1,2}))?", text)
    if date_match:
        year = date_match.group(1)
        month = date_match.group(2) or "00"
        day = date_match.group(3) or "00"
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return re.sub(r"[-./\s]+", "-", text)


def is_revision_history_data_cells(cells: list[re.Match[str]], header_map: dict[str, int]) -> bool:
    if len(cells) <= max(header_map.values()):
        return False
    labels = [normalize_label(get_cell_text(cell.group(0)).strip()) for cell in cells]
    return not revision_header_map(labels)


def iter_revision_history_data_rows(xml: str):
    rows = list(ROW_PATTERN.finditer(xml))
    for header_index, row_match in enumerate(rows):
        cells = list(CELL_PATTERN.finditer(row_match.group(0)))
        labels = [normalize_label(get_cell_text(cell.group(0)).strip()) for cell in cells]
        header_map = revision_header_map(labels)
        if not header_map:
            continue
        for data_row_match in rows[header_index + 1:]:
            data_cells = list(CELL_PATTERN.finditer(data_row_match.group(0)))
            if is_revision_history_data_cells(data_cells, header_map):
                yield data_row_match, data_cells, header_map
                break


def update_revision_history_xml(xml: str, revision_date: str, author: str, approval_author: str) -> tuple[str, int]:
    revision_date = normalize_revision_date_for_write(revision_date)
    rows = list(ROW_PATTERN.finditer(xml))
    for header_index, row_match in enumerate(rows):
        cells = list(CELL_PATTERN.finditer(row_match.group(0)))
        labels = [normalize_label(get_cell_text(cell.group(0)).strip()) for cell in cells]
        header_map = revision_header_map(labels)
        if not header_map:
            continue

        pieces: list[str] = []
        last = 0
        count = 0
        found_first_data_row = False
        for data_row_match in rows[header_index + 1:]:
            data_cells = list(CELL_PATTERN.finditer(data_row_match.group(0)))
            if not is_revision_history_data_cells(data_cells, header_map):
                continue

            if found_first_data_row:
                updated_row, row_count = clear_revision_history_data_row(data_row_match.group(0), data_cells)
            else:
                updates = [
                    (header_map["version"], "0.1"),
                    (header_map["date"], revision_date),
                    (header_map["reason"], REVISION_REASON_VALUE) if "reason" in header_map else None,
                    (header_map["detail"], REVISION_DETAIL_VALUE) if "detail" in header_map else None,
                    (header_map["author"], author),
                ]
                if "approval" in header_map:
                    updates.append((header_map["approval"], approval_author))
                updates = [item for item in updates if item is not None]
                updated_row, row_count = update_revision_history_first_row(
                    data_row_match.group(0),
                    data_cells,
                    updates,
                )
                found_first_data_row = True

            if row_count:
                pieces.append(xml[last:data_row_match.start()])
                pieces.append(updated_row)
                last = data_row_match.end()
                count += row_count

        if not count:
            return xml, 0
        pieces.append(xml[last:])
        return "".join(pieces), count
    return xml, 0


def update_revision_history_first_row(
    row_xml: str,
    cells: list[re.Match[str]],
    updates: list[tuple[int, str]],
) -> tuple[str, int]:
    offset = 0
    updated_row = row_xml
    count = 0
    for cell_index, new_text in updates:
        cell = cells[cell_index]
        updated_cell, _old_value = replace_cell_text(cell.group(0), new_text)
        start = cell.start() + offset
        end = cell.end() + offset
        updated_row = updated_row[:start] + updated_cell + updated_row[end:]
        offset += len(updated_cell) - len(cell.group(0))
        count += 1
    return updated_row, count


def clear_revision_history_data_row(row_xml: str, cells: list[re.Match[str]]) -> tuple[str, int]:
    offset = 0
    updated_row = row_xml
    count = 0
    for cell in cells:
        if not get_cell_text(cell.group(0)).strip():
            continue
        updated_cell, _old_value = replace_cell_text(cell.group(0), "")
        start = cell.start() + offset
        end = cell.end() + offset
        updated_row = updated_row[:start] + updated_cell + updated_row[end:]
        offset += len(updated_cell) - len(cell.group(0))
        count += 1
    return updated_row, count
