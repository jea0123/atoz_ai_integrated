import os
from pathlib import Path
import tempfile
import fitz
import re
import copy
import openpyxl
from typing import Callable
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
from document_update.hwpx_text import extract_document_text
from document_update.runtime_conversion import prepare_target_file

SCREEN_ID_PATTERN = re.compile(r"\bUI-[A-Z0-9]+(?:-[A-Z0-9]+)+\b", re.IGNORECASE)


def _check_cancel(cancel_check: Callable[[], None] | None) -> None:
  if cancel_check:
    cancel_check()

def derive_requirement_id_from_screen_id(screen_id):
  match = re.match(r"^UI-([A-Z0-9]+)-(\d{3})(?:-|$)", str(screen_id or "").strip(), re.IGNORECASE)
  if not match:
    return ""
  return f"SFR-{match.group(1).upper()}-{match.group(2)}"

def extract_text_from_pdf(pdf_path):
  if not os.path.exists(pdf_path):
    raise FileNotFoundError(f"사용자인터페이스설계서 파일을 찾을 수 없습니다: {pdf_path}")
  
  print(f"\n*** 사용자인터페이스설계서에서 텍스트 추출 중: {pdf_path}")
  suffix = Path(pdf_path).suffix.lower()
  if suffix == ".hwp":
    with tempfile.TemporaryDirectory(prefix="qa-ui-hwp-") as temp:
      converted_path, _converted = prepare_target_file(Path(pdf_path), Path(temp))
      return extract_document_text(converted_path).strip()
  if suffix == ".hwpx":
    return extract_document_text(Path(pdf_path)).strip()
  if suffix != ".pdf":
    raise ValueError("사용자인터페이스설계서는 PDF, HWP 또는 HWPX 파일만 지원합니다.")

  text_content = ""
  with fitz.open(pdf_path) as doc:
    for page_num in range(len(doc)):
      page = doc.load_page(page_num)
      text_content += page.get_text()

  return text_content.strip()

def extract_req_mapping_from_pdf(pdf_path):
  """사용자인터페이스설계서 PDF/HWPX에서 {화면ID: 요구사항ID} 매핑을 만든다."""
  text = extract_text_from_pdf(pdf_path)
  mapping = {}

  normalized = re.sub(r'[ \t]+', ' ', text)

  pair_pattern = re.compile(
    r'요구\s*사항\s*ID\s*'
    r'(SFR-[A-Z0-9]+(?:-[A-Z0-9]+)*)'
    r'[\s\S]{0,500}?'
    r'화면\s*ID\s*'
    r'(UI-[A-Z0-9]+(?:-[A-Z0-9]+)*)',
    re.IGNORECASE,
  )

  for match in pair_pattern.finditer(normalized):
    req_id = match.group(1).strip()
    screen_id = match.group(2).strip()
    mapping[screen_id] = req_id

  for match in SCREEN_ID_PATTERN.finditer(normalized):
    screen_id = match.group(0).strip()
    derived_req_id = derive_requirement_id_from_screen_id(screen_id)
    if derived_req_id:
      mapping.setdefault(screen_id, derived_req_id)

  return mapping

def extract_unit_test_from_excel(xlsx_path):
  if not os.path.exists(xlsx_path):
    raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {xlsx_path}")
  
  print(f"단위시험 케이스에서 데이터 추출 중: {xlsx_path}")
  wb = openpyxl.load_workbook(xlsx_path, data_only=True)

  unit_test_data = []

  for sheet in wb.worksheets:
    header_row_idx = None
    headers = []
    common_info = {}

    # 1. 상단 정보 추출 및 헤더(항목명)가 있는 행 찾기
    first_row_val = sheet.cell(row=1, column=1).value
    if first_row_val:
      common_info["타이틀"] = str(first_row_val).strip()

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
      if not any(row):
        continue

      # 셀 값들의 띄어쓰기와 줄바꿈을 제거하여 검색용 리스트 생성
      search_row = [str(cell).replace(" ", "").replace("\n", "") for cell in row if cell is not None]

      # '순서'(또는 '순번')과 '예상'이라는 단어가 같은 행에 모두 존재하면 헤더
      has_seq = any('순번' in val or '순서' in val for val in search_row)
      has_result = any('예상' in val or '결과' in val for val in search_row)

      if has_seq and has_result:
        header_row_idx = row_idx
        headers = [str(cell).strip().replace(" ", "_") if cell else f"Col_{i}" for i, cell in enumerate(row)]
        break
      else:
        # 헤더가 아니라면 상단의 메타 정보 행이므로 '사전조건', '화면ID' 등을 찾아 common_info에 저장
        for c_idx, cell in enumerate(row):
          if isinstance(cell, str):
            clean_label = cell.replace(" ", "")
            key_map = {
              "단위시험ID": "단위시험_ID",
              "단위시험명": "단위시험_명",
              "사전조건" : "사전조건",
              "화면ID": "화면_ID"
            }
            if clean_label in key_map:
              # 바로 오른쪽 칸에 있는 값을 가져옴
              val = row[c_idx + 1] if c_idx + 1 < len(row) and row[c_idx + 1] is not None else ""
              common_info[key_map[clean_label]] = str(val).strip()

    if not header_row_idx:
      continue

    # 2. 헤더 다음 행부터 실제 데이터 읽기
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
      if not any(row):
        continue

      # 헤더와 값을 1:1로 매핑하여 딕셔너리 생성
      row_data = {}
      for i, cell in enumerate(row):
        if i < len(headers):
          row_data[headers[i]] = str(cell).strip() if cell is not None else ""

      # 3. 추출해 둔 상단 메타 정보를 개별 스텝 데이터에 합침
      row_data.update(common_info)

      unit_test_data.append(row_data)

  return unit_test_data

def fill_req_mapping_from_screen_ids(req_mapping, unit_test_data):
  mapping = dict(req_mapping or {})
  for data in unit_test_data or []:
    screen_id = str(data.get("화면_ID", "")).strip()
    if not screen_id or screen_id in mapping:
      continue
    derived_req_id = derive_requirement_id_from_screen_id(screen_id)
    if derived_req_id:
      mapping[screen_id] = derived_req_id
  return mapping

def build_scenario_id(screen_id):
  screen_id = str(screen_id or "").strip()
  if screen_id.startswith("UI-"):
    return "AT-" + screen_id[3:]
  if screen_id == "UI":
    return "AT"
  return screen_id

def extract_scenario_case_name(title, unit_test_name):
  if title and " - " in title:
    return title.rsplit(" - ", 1)[-1].strip()
  if unit_test_name:
    return unit_test_name.strip()
  return ""

def normalize_label(value):
  return re.sub(r"\s+", "", str(value or ""))

def extract_cover_author_from_workbook(workbook_path):
  try:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
  except Exception:
    return ""

  try:
    cover_sheets = [ws for ws in wb.worksheets if "표지" in ws.title]
    for ws in cover_sheets:
      for row in ws.iter_rows():
        cells = list(row)
        for index, cell in enumerate(cells):
          if normalize_label(cell.value) != "작성자":
            continue
          for next_cell in cells[index + 1:]:
            if next_cell.value not in (None, ""):
              return str(next_cell.value).strip()

    for ws in wb.worksheets:
      for row in ws.iter_rows():
        cells = list(row)
        for cell in cells:
          if normalize_label(cell.value) != "작성자":
            continue
          for next_row in range(cell.row + 1, min(cell.row + 6, ws.max_row) + 1):
            value = ws.cell(row=next_row, column=cell.column).value
            if value not in (None, ""):
              return str(value).strip()
  finally:
    wb.close()

  return ""

def build_test_scenarios_from_unit_tests(unit_test_data, author=""):
  scenarios = []
  sequence_by_scenario = {}
  scenario_name_by_id = {}

  if unit_test_data:
    print(f"[통합시험 시나리오] 단위시험 데이터 {len(unit_test_data)}건 변환 중...")

  for row in unit_test_data:
    screen_id = row.get("화면_ID", "")
    scenario_id = build_scenario_id(screen_id)
    sequence_by_scenario[scenario_id] = sequence_by_scenario.get(scenario_id, 0) + 1

    scenario_name = scenario_name_by_id.get(scenario_id)
    if scenario_name is None:
      scenario_name = row.get("단위시험_명", "") or extract_scenario_case_name(row.get("타이틀", ""), "")
      scenario_name_by_id[scenario_id] = scenario_name

    scenarios.append({
      "시스템": "",
      "작성자": author,
      "테스트_기간": "",
      "시나리오ID": scenario_id,
      "시나리오명": scenario_name,
      "요구사항_ID": row.get("요구사항_ID", ""),
      "케이스명": extract_scenario_case_name(row.get("타이틀", ""), row.get("단위시험_명", "")),
      "순번": sequence_by_scenario[scenario_id],
      "업무처리내용": "",
      "시험항목": row.get("테스트_케이스", ""),
      "사전조건": row.get("사전조건", ""),
      "입력자료": row.get("테스트_데이터", ""),
      "예상결과": row.get("예상_결과", ""),
      "화면ID": row.get("화면_ID", "")
    })

  return scenarios
  
def safe_sheet_title(value, fallback="미분류"):
  title = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(value or "").strip()) or fallback
  return title[:31]

def unique_sheet_title(wb, value, fallback="미분류"):
  base = safe_sheet_title(value, fallback)
  title = base
  counter = 2
  while title in wb.sheetnames:
    suffix = f"_{counter}"
    title = f"{base[:31 - len(suffix)]}{suffix}"
    counter += 1
  return title

def group_test_scenarios_by_requirement(test_scenarios):
  grouped = {}
  for ts in test_scenarios:
    req_id = str(ts.get("요구사항_ID") or "").strip()
    if not req_id:
      req_id = derive_requirement_id_from_screen_id(ts.get("화면ID", "")) or "미분류"

    scenario_id = str(ts.get("시나리오ID") or "미분류").strip() or "미분류"
    grouped.setdefault(req_id, {})
    grouped[req_id].setdefault(scenario_id, [])
    grouped[req_id][scenario_id].append(ts)
  return grouped

REPEATED_SCENARIO_BLOCK_START_ROW = 3
TEMPLATE_DATA_START_ROW = 7

def get_effective_template_block_end_row(source_ws):
  end_row = TEMPLATE_DATA_START_ROW

  for row in source_ws.iter_rows():
    if any(cell.value not in (None, "") for cell in row):
      end_row = max(end_row, row[0].row)

  for merged_range in source_ws.merged_cells.ranges:
    end_row = max(end_row, merged_range.max_row)

  return end_row

def copy_worksheet_block(source_ws, target_ws, row_offset=0, source_start_row=1, source_end_row=None):
  source_end_row = source_end_row or get_effective_template_block_end_row(source_ws)

  for key, dim in source_ws.column_dimensions.items():
    target_ws.column_dimensions[key].min = dim.min
    target_ws.column_dimensions[key].max = dim.max
    target_ws.column_dimensions[key].width = dim.width
    target_ws.column_dimensions[key].hidden = dim.hidden

  for key, dim in source_ws.row_dimensions.items():
    if key < source_start_row or key > source_end_row:
      continue
    target_key = key + row_offset
    target_ws.row_dimensions[target_key].height = dim.height
    target_ws.row_dimensions[target_key].hidden = dim.hidden

  existing_ranges = {str(m_range) for m_range in target_ws.merged_cells.ranges}
  for m_range in source_ws.merged_cells.ranges:
    min_col, min_row, max_col, max_row = range_boundaries(str(m_range))
    if min_row < source_start_row or max_row > source_end_row:
      continue
    target_range = (
      f"{get_column_letter(min_col)}{min_row + row_offset}:"
      f"{get_column_letter(max_col)}{max_row + row_offset}"
    )
    if target_range not in existing_ranges:
      target_ws.merge_cells(target_range)
      existing_ranges.add(target_range)

  for row in source_ws.iter_rows():
    for cell in row:
      if cell.row < source_start_row or cell.row > source_end_row:
        continue
      new_cell = target_ws.cell(
        row=cell.row + row_offset,
        column=cell.column,
        value=cell.value,
      )
      if cell.has_style:
        new_cell.font = copy.copy(cell.font)
        new_cell.border = copy.copy(cell.border)
        new_cell.fill = copy.copy(cell.fill)
        new_cell.number_format = copy.copy(cell.number_format)
        new_cell.protection = copy.copy(cell.protection)
        new_cell.alignment = copy.copy(cell.alignment)

def ensure_row_merge(ws, row, start_column, end_column):
  target_range = (
    f"{get_column_letter(start_column)}{row}:"
    f"{get_column_letter(end_column)}{row}"
  )
  if target_range not in {str(m_range) for m_range in ws.merged_cells.ranges}:
    ws.merge_cells(
      start_row=row,
      start_column=start_column,
      end_row=row,
      end_column=end_column,
    )

def set_cell_value(ws, row, column, value):
  cell = ws.cell(row=row, column=column)
  if cell.__class__.__name__ == "MergedCell":
    for merged_range in ws.merged_cells.ranges:
      if cell.coordinate in merged_range:
        cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        break
  cell.value = value
  return cell

def apply_data_row_style(ws, source_ws, row, max_column, border, center_align, center_columns):
  template_row = min(7, source_ws.max_row)

  for col in range(1, max_column + 1):
    cell = ws.cell(row=row, column=col)
    template_cell = source_ws.cell(row=template_row, column=min(col, source_ws.max_column))
    if template_cell.has_style:
      cell.font = copy.copy(template_cell.font)
      cell.fill = copy.copy(template_cell.fill)
      cell.number_format = copy.copy(template_cell.number_format)
      cell.protection = copy.copy(template_cell.protection)
    cell.border = copy.copy(template_cell.border) if template_cell.has_style else border
    if col in center_columns:
      cell.alignment = center_align

def write_scenario_block(ws, source_ws, group, row_offset, max_column, border, center_align, left_align, center_columns):
  common = group[0]
  if row_offset == 0:
    set_cell_value(ws, 2, 2, common.get("시스템", ""))
    set_cell_value(ws, 2, 6, common.get("작성자", ""))
  set_cell_value(ws, row_offset + 4, 2, common.get("시나리오ID", ""))
  set_cell_value(ws, row_offset + 4, 6, common.get("시나리오명", ""))
  set_cell_value(ws, row_offset + 4, 9, common.get("요구사항_ID", ""))

  for i, ts in enumerate(group):
    row_number = row_offset + 7 + i
    ensure_row_merge(ws, row_number, 1, 2)
    apply_data_row_style(ws, source_ws, row_number, max_column, border, center_align, center_columns)

    set_cell_value(ws, row_number, 1, ts.get("케이스명", ""))
    data_mapping = [
      (3, i + 1, center_align),
      (4, ts.get("업무처리내용", ""), left_align),
      (5, ts.get("시험항목", ""), left_align),
      (6, ts.get("사전조건", ""), left_align),
      (7, "", left_align),
      (8, ts.get("예상결과", ""), left_align),
      (9, ts.get("화면ID", ""), left_align),
    ]

    for col, val, align in data_mapping:
      cell = set_cell_value(ws, row_number, col, str(val) if val else "")
      cell.alignment = align

def save_test_scenarios_to_excel(test_scenarios, base_workbook_path, output_dir: Path, scenario_sheet_form_path, base_filename="generated_TS"):
  if not test_scenarios:
    print("저장할 데이터가 없습니다.")
    return
  
  base_workbook_path = Path(base_workbook_path)
  output_dir = Path(output_dir)
  scenario_sheet_form_path = Path(scenario_sheet_form_path)
  
  if not base_workbook_path.exists():
    print(f"업로드된 기존 시나리오 파일을 찾을 수 없습니다: {base_workbook_path}")
    return
  
  if not scenario_sheet_form_path.exists():
    print(f"서버에 표준 양식 파일({scenario_sheet_form_path})이 없습니다.")
    return
  
  output_dir.mkdir(parents=True, exist_ok=True)
  
  # 파일명 자동 증가
  counter = 1
  while True:
    output_filename = f"{base_filename}_{counter}.xlsx"
    full_path = output_dir / output_filename
    if not full_path.exists():
      break
    counter += 1

  print(f"기존 파일 로드 중: {base_workbook_path}")
  wb = openpyxl.load_workbook(base_workbook_path)

  print(f"외부 표준 양식 로드 중: {scenario_sheet_form_path}")
  form_wb = openpyxl.load_workbook(scenario_sheet_form_path)
  source_ws = form_wb.active

  # 1. '개정이력'과 '작성방법' 인덱스 찾기
  sheet_names = wb.sheetnames
  start_idx = -1
  end_idx = -1

  for i, name in enumerate(sheet_names):
    if "개정이력" in name:
      start_idx = i
    elif "작성방법" in name:
      end_idx = i

  # 3. 기존 시나리오 시트 삭제 (개정이력과 작성방법 사이)
  if start_idx != -1 and end_idx != -1 and start_idx < end_idx:
      sheets_to_remove = sheet_names[start_idx + 1 : end_idx]
      for sheet_name in sheets_to_remove:
        del wb[sheet_name]
      insert_idx = start_idx + 1
  else:
      insert_idx = 1

  border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
  center_align = Alignment(horizontal="center", vertical="center")
  left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

  grouped = group_test_scenarios_by_requirement(test_scenarios)
  max_scenario_column = max(source_ws.max_column, 9)
  template_block_end_row = get_effective_template_block_end_row(source_ws)

  # 4. 요구사항ID 시트 안에 시나리오 블록을 반복하여 새 데이터 채우기
  for req_id, scenario_groups in grouped.items():
    sheet_title = unique_sheet_title(wb, req_id)
    new_ws = wb.create_sheet(title=sheet_title, index=insert_idx)
    insert_idx += 1

    next_start_row = 1
    is_first_block = True
    for group in scenario_groups.values():
      source_start_row = 1 if is_first_block else REPEATED_SCENARIO_BLOCK_START_ROW
      row_offset = next_start_row - source_start_row
      copy_worksheet_block(source_ws, new_ws, row_offset, source_start_row, template_block_end_row)
      write_scenario_block(
        new_ws,
        source_ws,
        group,
        row_offset,
        max_scenario_column,
        border,
        center_align,
        left_align,
        center_columns={1, 2, 3},
      )
      block_last_row = row_offset + max(template_block_end_row, 6 + len(group))
      next_start_row = block_last_row + 1
      is_first_block = False

  form_wb.close()
  wb.save(full_path)
  print(f"통합시험 시나리오 저장 완료: {full_path.resolve()}")
  return full_path

def find_workbook_section_indices(wb):
  start_idx = -1
  end_idx = -1
  for i, name in enumerate(wb.sheetnames):
    if "개정이력" in name:
      start_idx = i
    elif "작성방법" in name:
      end_idx = i
  return start_idx, end_idx

def save_integration_test_results_to_excel(test_scenarios, base_workbook_path, output_dir: Path, result_sheet_form_path, base_filename="generated_TR"):
  if not test_scenarios:
    print("저장할 데이터가 없습니다.")
    return

  base_workbook_path = Path(base_workbook_path)
  output_dir = Path(output_dir)
  result_sheet_form_path = Path(result_sheet_form_path)

  if not base_workbook_path.exists():
    print(f"업로드된 기존 통합시험 결과서 파일을 찾을 수 없습니다: {base_workbook_path}")
    return

  if not result_sheet_form_path.exists():
    print(f"서버에 통합시험 결과서 표준 양식 파일({result_sheet_form_path})이 없습니다.")
    return

  output_dir.mkdir(parents=True, exist_ok=True)

  counter = 1
  while True:
    output_filename = f"{base_filename}_{counter}.xlsx"
    full_path = output_dir / output_filename
    if not full_path.exists():
      break
    counter += 1

  print(f"기존 통합시험 결과서 파일 로드 중: {base_workbook_path}")
  wb = openpyxl.load_workbook(base_workbook_path)

  print(f"외부 통합시험 결과서 표준 양식 로드 중: {result_sheet_form_path}")
  form_wb = openpyxl.load_workbook(result_sheet_form_path)
  source_ws = form_wb.active

  start_idx, end_idx = find_workbook_section_indices(wb)
  sheet_names = wb.sheetnames
  if start_idx != -1 and end_idx != -1 and start_idx < end_idx:
    sheets_to_remove = sheet_names[start_idx + 1 : end_idx]
    for sheet_name in sheets_to_remove:
      del wb[sheet_name]
    insert_idx = start_idx + 1
  else:
    insert_idx = 1

  border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
  center_align = Alignment(horizontal="center", vertical="center")
  left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
  max_result_column = max(source_ws.max_column, 13)
  template_block_end_row = get_effective_template_block_end_row(source_ws)

  grouped = group_test_scenarios_by_requirement(test_scenarios)

  for req_id, scenario_groups in grouped.items():
    sheet_title = unique_sheet_title(wb, req_id)
    new_ws = wb.create_sheet(title=sheet_title, index=insert_idx)
    insert_idx += 1

    next_start_row = 1
    is_first_block = True
    for group in scenario_groups.values():
      source_start_row = 1 if is_first_block else REPEATED_SCENARIO_BLOCK_START_ROW
      row_offset = next_start_row - source_start_row
      copy_worksheet_block(source_ws, new_ws, row_offset, source_start_row, template_block_end_row)
      write_scenario_block(
        new_ws,
        source_ws,
        group,
        row_offset,
        max_result_column,
        border,
        center_align,
        left_align,
        center_columns={1, 2, 3},
      )
      block_last_row = row_offset + max(template_block_end_row, 6 + len(group))
      next_start_row = block_last_row + 1
      is_first_block = False

  form_wb.close()
  wb.save(full_path)
  print(f"통합시험 결과서 저장 완료: {full_path.resolve()}")
  return full_path

def generate_test_scenarios(
    template_xlsx_path: Path,
    tc_xlsx_path: Path,
    ui_pdf_path: Path | None,
    output_dir: Path,
    form_path: Path,
    req_mapping: dict[str, str] | None = None,
    unit_test_data: list[dict] | None = None,
    log_progress: bool = True,
    cancel_check: Callable[[], None] | None = None,
) -> dict:
    _check_cancel(cancel_check)
    template_xlsx_path = Path(template_xlsx_path)
    tc_xlsx_path = Path(tc_xlsx_path)
    output_dir = Path(output_dir)
    form_path = Path(form_path)

    output_dir.mkdir(parents=True, exist_ok=True)

    cover_author = extract_cover_author_from_workbook(template_xlsx_path)
    _check_cancel(cancel_check)
    _check_cancel(cancel_check)
    req_mapping = dict(req_mapping or {})
    if not req_mapping and ui_pdf_path:
        ui_pdf_path = Path(ui_pdf_path)
        if ui_pdf_path.exists():
            req_mapping = extract_req_mapping_from_pdf(ui_pdf_path)
            _check_cancel(cancel_check)

    if unit_test_data is None:
        unit_test_data = extract_unit_test_from_excel(tc_xlsx_path)
    else:
        unit_test_data = [dict(data) for data in unit_test_data]
    _check_cancel(cancel_check)
    if not unit_test_data:
        return {
            "ok": False,
            "error": "단위시험 케이스 엑셀에서 데이터를 찾지 못했습니다.",
            "files": [],
        }

    req_mapping = fill_req_mapping_from_screen_ids(req_mapping, unit_test_data)
    missing_req_screen_ids = sorted({
        str(data.get("화면_ID", "")).strip()
        for data in unit_test_data
        if str(data.get("화면_ID", "")).strip()
        and str(data.get("화면_ID", "")).strip() not in req_mapping
    })
    if missing_req_screen_ids:
        return {
            "ok": False,
            "error": "사용자인터페이스설계서에서 화면 ID에 해당하는 요구사항 ID를 찾지 못했습니다.",
            "missing_screen_ids": missing_req_screen_ids[:10],
            "missing_screen_count": len(missing_req_screen_ids),
            "files": [],
        }

    for data in unit_test_data:
        screen_id = data.get("화면_ID", "").strip()
        data["요구사항_ID"] = req_mapping.get(screen_id, "")

    ts_result = build_test_scenarios_from_unit_tests(unit_test_data, author=cover_author)
    _check_cancel(cancel_check)

    if not ts_result:
        return {
            "ok": False,
            "error": "통합시험 시나리오로 변환할 수 있는 단위시험 케이스 데이터가 없습니다.",
            "files": [],
        }

    excel_path = save_test_scenarios_to_excel(
        ts_result,
        base_workbook_path=template_xlsx_path,
        output_dir=output_dir,
        scenario_sheet_form_path=form_path,
    )
    _check_cancel(cancel_check)

    files = []
    if excel_path:
        files.append({
            "kind": "xlsx",
            "path": str(excel_path),
            "name": excel_path.name,
        })

    return {
        "ok": True,
        "count": len(ts_result),
        "files": files,
    }

def generate_integration_test_results(
    template_xlsx_path: Path,
    tc_xlsx_path: Path,
    ui_pdf_path: Path | None,
    output_dir: Path,
    form_path: Path,
    req_mapping: dict[str, str] | None = None,
    unit_test_data: list[dict] | None = None,
    cancel_check: Callable[[], None] | None = None,
) -> dict:
    _check_cancel(cancel_check)
    template_xlsx_path = Path(template_xlsx_path)
    tc_xlsx_path = Path(tc_xlsx_path)
    output_dir = Path(output_dir)
    form_path = Path(form_path)

    output_dir.mkdir(parents=True, exist_ok=True)

    cover_author = extract_cover_author_from_workbook(template_xlsx_path)
    req_mapping = dict(req_mapping or {})
    if not req_mapping and ui_pdf_path:
        ui_pdf_path = Path(ui_pdf_path)
        if ui_pdf_path.exists():
            req_mapping = extract_req_mapping_from_pdf(ui_pdf_path)
            _check_cancel(cancel_check)

    if unit_test_data is None:
        unit_test_data = extract_unit_test_from_excel(tc_xlsx_path)
    else:
        unit_test_data = [dict(data) for data in unit_test_data]
    _check_cancel(cancel_check)
    if not unit_test_data:
        return {
            "ok": False,
            "error": "단위시험 케이스 엑셀에서 데이터를 찾지 못했습니다.",
            "files": [],
        }

    req_mapping = fill_req_mapping_from_screen_ids(req_mapping, unit_test_data)
    missing_req_screen_ids = sorted({
        str(data.get("화면_ID", "")).strip()
        for data in unit_test_data
        if str(data.get("화면_ID", "")).strip()
        and str(data.get("화면_ID", "")).strip() not in req_mapping
    })
    if missing_req_screen_ids:
        return {
            "ok": False,
            "error": "사용자인터페이스설계서에서 화면 ID에 해당하는 요구사항 ID를 찾지 못했습니다.",
            "missing_screen_ids": missing_req_screen_ids[:10],
            "missing_screen_count": len(missing_req_screen_ids),
            "files": [],
        }

    for data in unit_test_data:
        _check_cancel(cancel_check)
        _check_cancel(cancel_check)
        screen_id = data.get("화면_ID", "").strip()
        data["요구사항_ID"] = req_mapping.get(screen_id, "")

    result_rows = build_test_scenarios_from_unit_tests(unit_test_data, author=cover_author)
    _check_cancel(cancel_check)
    if not result_rows:
        return {
            "ok": False,
            "error": "통합시험 결과서로 변환할 수 있는 단위시험 케이스 데이터가 없습니다.",
            "files": [],
        }

    excel_path = save_integration_test_results_to_excel(
        result_rows,
        base_workbook_path=template_xlsx_path,
        output_dir=output_dir,
        result_sheet_form_path=form_path,
    )
    _check_cancel(cancel_check)

    files = []
    if excel_path:
        files.append({
            "kind": "xlsx",
            "path": str(excel_path),
            "name": excel_path.name,
        })

    return {
        "ok": True,
        "count": len(result_rows),
        "files": files,
    }
