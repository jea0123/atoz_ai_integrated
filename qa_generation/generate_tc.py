import json
import os
import re
import time
import ctypes
import zipfile
import tempfile
from typing import Callable
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pathlib import Path
import time
from document_update.hwpx_text import extract_document_text
from document_update.runtime_conversion import prepare_target_file

try:
  import requests
except ImportError:
  class _RequestsUnavailableExceptions:
    class Timeout(Exception):
      pass

    class RequestException(Exception):
      pass

  class _RequestsUnavailable:
    exceptions = _RequestsUnavailableExceptions

    @staticmethod
    def post(*_args, **_kwargs):
      raise _RequestsUnavailableExceptions.RequestException("requests 패키지가 설치되어 있지 않습니다.")

  requests = _RequestsUnavailable()

try:
  import fitz
except ImportError:
  fitz = None

try:
  from pyhwpx import Hwp
  import pythoncom
  HWP_AVAILABLE = True
except ImportError:
  HWP_AVAILABLE = False

REQUIRED_TC_KEYS = [
  "단위시험_ID",
  "화면명",
  "수행자",
  "단위시험_명",
  "수행_일자",
  "사전조건",
  "화면_ID",
  "순서",
  "테스트_케이스",
  "테스트_데이터",
  "예상_결과",
  "수행_결과",
]

CIRCLED_NUMBERS = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"


class TestCaseGenerationError(RuntimeError):
  pass


SIZE_ORDER = {
  "small": 0,
  "medium": 1,
  "large": 2,
}

SIZE_LABELS = {
  "small": "소량",
  "medium": "보통",
  "large": "대량",
}


def classify_document_size(screen_count):
  try:
    count = int(screen_count or 0)
  except (TypeError, ValueError):
    count = 0
  if count <= 5:
    return "small"
  if count <= 10:
    return "medium"
  return "large"


def extract_cover_author_from_document(document_path):
  try:
    text = extract_document_text(Path(document_path))
  except Exception:
    return ""

  lines = [line.strip() for line in text.splitlines() if line.strip()]
  for index, line in enumerate(lines[:80]):
    if re.sub(r"\s+", "", line) != "작성자":
      continue
    for value in lines[index + 1:index + 5]:
      normalized = re.sub(r"\s+", "", value)
      if normalized and normalized not in {"작성자", "승인자"}:
        return value
  return ""

def extract_text_from_pdf(pdf_path):
  document_path = Path(pdf_path)
  if not document_path.exists():
    raise FileNotFoundError(f"사용자인터페이스설계서 파일을 찾을 수 없습니다: {document_path}")

  print(f"\n*** 사용자인터페이스설계서에서 텍스트 추출 중: {document_path}")
  suffix = document_path.suffix.lower()
  if suffix not in {".pdf", ".hwp", ".hwpx"}:
    raise ValueError("사용자인터페이스설계서는 PDF, HWP 또는 HWPX 파일만 지원합니다.")

  if suffix == ".hwp":
    with tempfile.TemporaryDirectory(prefix="qa-ui-hwp-") as temp:
      converted_path, _converted = prepare_target_file(document_path, Path(temp))
      return extract_document_text(converted_path).strip()

  if suffix == ".hwpx":
    return extract_document_text(document_path).strip()

  if fitz is not None:
    try:
      text_content = ""
      with fitz.open(document_path) as doc:
        for page_num in range(len(doc)):
          page = doc.load_page(page_num)
          page_text = page.get_text()
          text_content += page_text
      if text_content.strip():
        return text_content.strip()
    except Exception:
      pass

  return extract_document_text(document_path).strip()


def extract_section_title(section_heading):
  raw_text = str(section_heading or "")
  if not re.match(r"\s*4\s*\.\s*\d+\s*\.", raw_text, re.DOTALL):
    return ""
  title_source = re.sub(r"^\s*4\s*\.\s*\d+\s*\.\s*", "", raw_text, count=1, flags=re.DOTALL)
  title = ""
  metadata_labels = ("요구사항ID", "화면ID", "화면명", "화면설명", "메뉴경로", "개발구분", "개인정보등급", "화면구성", "처리흐름")
  for line in title_source.splitlines():
    candidate = clean_extracted_text_value(line)
    if not candidate or candidate.isdigit():
      continue
    if any(candidate.startswith(label) for label in metadata_labels):
      return ""
    title = candidate
    break
  title = re.sub(r"\s+\d{1,4}$", "", title).strip()
  return title


def extract_block_section_title(block_text):
  match = re.search(r"(?m)^\s*4\s*\.\s*\d+\s*\.\s*[^\n]*(?:\n\s*[^\n]+)?", str(block_text or ""))
  return extract_section_title(match.group(0)) if match else ""


def clean_extracted_text_value(value):
  text = str(value or "")
  text = re.sub(r"<[^>]+>", " ", text)
  text = re.sub(r"\s+", " ", text).strip()
  return text


def extract_screen_blocks(extracted_text):
  screen_pattern = re.compile(r"\bUI-[A-Z0-9]+(?:-[A-Z0-9]+)+\b")
  section_pattern = re.compile(r"(?m)^\s*4\s*\.\s*\d+\s*\.\s*[^\n]*(?:\n\s*[^\n]+)?")
  main_section_match = re.search(r"(?m)^\s*4\.\s*화면/보고서\s*정의", extracted_text)
  search_offset = main_section_match.start() if main_section_match else 0
  target_text = extracted_text[search_offset:]
  section_matches = list(section_pattern.finditer(target_text))

  blocks_by_screen_id = {}

  def screen_block_score(block_text):
    keyword_score = sum(
      1 for keyword in ["요구사항ID", "화면ID", "화면명", "화면설명", "메뉴경로", "화면구성", "처리흐름", "이벤트 정의"]
      if keyword in block_text
    )
    flow_score = 100 if re.search(r"처리\s*흐름|처리흐름", block_text) else 0
    event_score = 100 if re.search(r"이벤트\s*정의", block_text) else 0
    step_score = (
      len(extract_event_definition_steps(block_text)) * 12
      + len(extract_process_flow_steps(block_text)) * 10
    )
    length_score = min(len(block_text) // 500, 20)
    return flow_score + event_score + step_score + keyword_score + length_score

  if section_matches:
    print(f"화면 정의 제목 기준 분할 시도: {len(section_matches)}개 제목 발견")
    if main_section_match:
      print("'4. 화면/보고서 정의' 본문 이후부터 화면 제목 분할을 적용합니다.")
    else:
      print("'4. 화면/보고서 정의' 본문 시작점을 찾지 못해 전체 텍스트에서 화면 제목 분할을 적용합니다.")

    def add_screen_candidate(screen_id, candidate_text, section_title=""):
      candidate = {
        "screen_id": screen_id,
        "unit_test_id": screen_id.replace("UI", "UT", 1),
        "section_title": section_title,
        "text": candidate_text,
        "score": screen_block_score(candidate_text)
      }

      existing = blocks_by_screen_id.get(screen_id)
      if existing and existing["score"] >= candidate["score"]:
        print(f"[screen split skip] duplicate screen_id: {screen_id} (existing score {existing['score']} >= candidate score {candidate['score']})")
        return

      if existing:
        print(f"[screen split replace] duplicate screen_id: {screen_id} (existing score {existing['score']} < candidate score {candidate['score']})")
        if not candidate.get("section_title") and existing.get("section_title"):
          candidate["section_title"] = existing["section_title"]

      blocks_by_screen_id[screen_id] = candidate

    for idx, section_match in enumerate(section_matches):
      start = section_match.start()
      end = section_matches[idx + 1].start() if idx + 1 < len(section_matches) else len(target_text)
      block_text = target_text[start:end].strip()
      section_title = extract_section_title(section_match.group(0))
      screen_matches = list(screen_pattern.finditer(block_text))

      if not screen_matches:
        print(f"[화면 분할 제외] 화면ID 없음: {section_match.group(0).strip()}")
        continue

      screen_boundaries = []
      for screen_match in screen_matches:
        screen_id = screen_match.group(0)
        if screen_boundaries and screen_boundaries[-1].group(0) == screen_id:
          continue
        screen_boundaries.append(screen_match)

      if len(screen_boundaries) == 1:
        add_screen_candidate(screen_boundaries[0].group(0), block_text, section_title)
        continue

      print(f"[screen split] {section_match.group(0).strip()} contains {len(screen_boundaries)} screen IDs")
      for screen_idx, screen_match in enumerate(screen_boundaries):
        screen_id = screen_match.group(0)
        screen_start = 0 if screen_idx == 0 else screen_match.start()
        screen_end = (
          screen_boundaries[screen_idx + 1].start()
          if screen_idx + 1 < len(screen_boundaries)
          else len(block_text)
        )
        screen_section_title = section_title if screen_idx == 0 else ""
        add_screen_candidate(screen_id, block_text[screen_start:screen_end].strip(), screen_section_title)
      continue

    if blocks_by_screen_id:
      return [
        {key: value for key, value in block.items() if key != "score"}
        for block in blocks_by_screen_id.values()
      ]

  print("화면 정의 제목 기준 분할 실패. 화면ID 위치 기준 fallback을 사용합니다.")
  matches = list(screen_pattern.finditer(target_text))

  if not matches:
    return []

  blocks = []
  screen_order = []
  for match in matches:
    screen_id = match.group(0)
    if screen_id not in screen_order:
      screen_order.append(screen_id)

  for screen_id in screen_order:
    screen_matches = [match for match in matches if match.group(0) == screen_id]
    parts = []

    for match in screen_matches:
      next_match_start = len(target_text)
      for candidate in matches:
        if candidate.start() > match.start() and candidate.group(0) != screen_id:
          next_match_start = candidate.start()
          break

      start = match.start()
      end = min(len(target_text), next_match_start)
      parts.append(target_text[start:end].strip())

    merged_block = "\n\n".join(dict.fromkeys(part for part in parts if part))
    blocks.append({
      "screen_id": screen_id,
      "unit_test_id": screen_id.replace("UI", "UT", 1),
      "section_title": "",
      "text": merged_block
    })

  return blocks

def extract_process_flow_steps(screen_text):
  flow_match = re.search(r"처리\s*흐름|처리흐름", screen_text)
  if not flow_match:
    return []

  target_text = screen_text[flow_match.end():]
  next_section = re.search(
    r"이벤트\s*정의|입력데이터\s*검증|입력\s*데이터|보안기능|에러처리|세션통제|개인정보|4\.\d+",
    target_text
  )
  if next_section:
    target_text = target_text[:next_section.start()]

  steps = []
  circled_pattern = re.compile(rf"([{CIRCLED_NUMBERS}])\s*")
  circled_matches = list(circled_pattern.finditer(target_text))

  if circled_matches:
    for idx, match in enumerate(circled_matches):
      start = match.end()
      end = circled_matches[idx + 1].start() if idx + 1 < len(circled_matches) else len(target_text)
      text = target_text[start:end].strip()
      text = re.sub(r"\s+", " ", text).strip()
      if text:
        steps.append({
          "순서": CIRCLED_NUMBERS.index(match.group(1)) + 1,
          "내용": text
        })
    return steps

  numbered_pattern = re.compile(r"(?m)^\s*(\d{1,2})[\.\)]\s+")
  numbered_matches = list(numbered_pattern.finditer(target_text))

  for idx, match in enumerate(numbered_matches):
    start = match.end()
    end = numbered_matches[idx + 1].start() if idx + 1 < len(numbered_matches) else len(target_text)
    text = target_text[start:end].strip()
    text = re.sub(r"\s+", " ", text).strip()
    if text:
      steps.append({
        "순서": int(match.group(1)),
        "내용": text
      })

  return steps

def extract_event_definition_steps(screen_text):
  event_match = re.search(r"이벤트\s*정의", screen_text)
  if not event_match:
    return []

  target_text = screen_text[event_match.end():]
  next_section = re.search(
    r"입력데이터\s*검증|입력\s*데이터|보안기능|에러처리|세션통제|개인정보|처리\s*흐름|처리흐름|4\.\d+",
    target_text
  )
  if next_section:
    target_text = target_text[:next_section.start()]

  lines = [
    line.strip()
    for line in target_text.splitlines()
  ]
  lines = [
    line for line in lines
    if line and not re.fullmatch(r"(순번|이벤트명|처리설명|순번 이벤트명 처리설명)", re.sub(r"\s+", " ", line).strip())
  ]
  normalized_text = "\n".join(lines)
  row_pattern = re.compile(r"(?m)^\s*(\d{1,2})[\.\)]?\s+")
  row_matches = list(row_pattern.finditer(normalized_text))

  steps = []
  for idx, match in enumerate(row_matches):
    start = match.end()
    end = row_matches[idx + 1].start() if idx + 1 < len(row_matches) else len(normalized_text)
    row_text = normalized_text[start:end].strip()
    if not row_text:
      continue
    event_name, description = split_event_definition_row(row_text)
    steps.append({
      "순서": int(match.group(1)),
      "이벤트명": event_name,
      "처리설명": description,
      "내용": re.sub(r"\s+", " ", row_text).strip(),
    })

  return steps


def split_event_definition_row(row_text):
  raw_text = str(row_text or "").strip()
  text = re.sub(r"\s+", " ", raw_text).strip()
  if not text:
    return "", ""

  parts = re.split(r"\s{2,}", raw_text, maxsplit=1)
  if len(parts) == 2:
    return re.sub(r"\s+", " ", parts[0]).strip(), re.sub(r"\s+", " ", parts[1]).strip()

  event_name_pattern = re.compile(
    r"^(.+?(?:버튼\s*클릭|행\s*선택|건\s*더블클릭|더블클릭|클릭|입력|선택|저장|검색|조회|다운로드|닫기|초기화|삭제|수정|등록|이동))\s+(.+)$"
  )
  match = event_name_pattern.match(text)
  if match:
    return match.group(1).strip(), match.group(2).strip()

  return text, ""


def is_user_action_step(step_text):
  text = str(step_text or "").strip()
  if not text:
    return False

  action_keywords = (
    "클릭", "입력", "선택", "저장", "검색", "조회", "다운로드", "업로드",
    "등록", "수정", "삭제", "초기화", "이동", "로그인", "체크", "해제",
  )
  result_keywords = (
    "확인", "표시", "조회된다", "조회된다.", "호출", "검출", "보여", "나타",
    "출력", "완료", "성공", "실패", "저장된다", "변경된다",
  )
  has_action = any(keyword in text for keyword in action_keywords)
  if not has_action:
    return False
  if any(keyword in text for keyword in ("클릭", "입력", "선택", "저장", "검색", "다운로드", "업로드", "등록", "수정", "삭제", "초기화")):
    return True
  return not any(keyword in text for keyword in result_keywords)


def extract_action_flow_steps(screen_text):
  return [
    step for step in extract_process_flow_steps(screen_text)
    if is_user_action_step(step.get("내용", ""))
  ]


def get_action_type(text):
  value = str(text or "")
  action_groups = [
    ("search", ("검색", "조회")),
    ("save", ("저장",)),
    ("select", ("행 선택", "선택", "클릭", "더블클릭")),
    ("input", ("입력",)),
    ("download", ("다운로드",)),
    ("upload", ("업로드",)),
    ("reset", ("초기화",)),
    ("delete", ("삭제",)),
    ("edit", ("수정", "변경")),
    ("move", ("이동",)),
    ("login", ("로그인",)),
  ]
  return {
    action_type
    for action_type, keywords in action_groups
    if any(keyword in value for keyword in keywords)
  }


def normalize_action_text(text):
  value = str(text or "")
  value = re.sub(r"\[[^\]]+\]", " ", value)
  value = re.sub(r"\([^)]*\)", " ", value)
  value = re.sub(r"[^0-9A-Za-z가-힣]+", " ", value)
  tokens = [
    token for token in value.split()
    if token not in {
      "버튼", "클릭", "선택", "입력", "저장", "검색", "조회", "다운로드", "업로드",
      "등록", "수정", "삭제", "초기화", "이동", "한다", "한다.", "후", "전",
      "목록", "결과", "중", "하나", "행", "화면", "팝업",
    }
  ]
  return "".join(tokens)


def is_similar_action(event_step, flow_step):
  event_text = f"{event_step.get('이벤트명', '')} {event_step.get('처리설명', '')}"
  flow_text = str(flow_step.get("내용", ""))
  event_actions = get_action_type(event_text)
  flow_actions = get_action_type(flow_text)

  if event_actions and flow_actions and not event_actions.intersection(flow_actions):
    return False

  event_key = normalize_action_text(event_text)
  flow_key = normalize_action_text(flow_text)
  if event_key and flow_key and (event_key in flow_key or flow_key in event_key):
    return True

  if "행" in event_text and any(keyword in flow_text for keyword in ("행", "목록 중", "하나의")):
    return True

  return bool(event_actions and flow_actions and event_actions.intersection(flow_actions) and event_key and flow_key and event_key[:2] == flow_key[:2])


def merge_user_action_candidates(event_steps, action_flow_steps):
  candidates = []
  for step in event_steps:
    candidates.append({
      "source": "event",
      "순서": step.get("순서", len(candidates) + 1),
      "내용": f"{step.get('이벤트명', '')} - {step.get('처리설명', '')}".strip(" -"),
      "event_step": step,
      "flow_step": None,
    })

  for flow_step in action_flow_steps:
    if any(is_similar_action(candidate.get("event_step") or {}, flow_step) for candidate in candidates if candidate.get("event_step")):
      continue
    candidates.append({
      "source": "flow",
      "순서": flow_step.get("순서", len(candidates) + 1),
      "내용": flow_step.get("내용", ""),
      "event_step": None,
      "flow_step": flow_step,
    })

  return candidates


def get_test_case_generation_basis(block_text):
  event_steps = extract_event_definition_steps(block_text)
  flow_steps = extract_process_flow_steps(block_text)
  action_flow_steps = [
    step for step in flow_steps
    if is_user_action_step(step.get("내용", ""))
  ]
  action_candidates = merge_user_action_candidates(event_steps, action_flow_steps)

  if event_steps and flow_steps:
    basis = "event_flow_compare"
  elif event_steps:
    basis = "event"
  else:
    basis = "flow"
  expected_steps = len(action_candidates) or len(flow_steps)

  return {
    "basis": basis,
    "expected_steps": expected_steps,
    "event_steps": event_steps,
    "flow_steps": flow_steps,
    "action_flow_steps": action_flow_steps,
    "action_candidates": action_candidates,
  }


def parse_llm_json(response_text):
  try:
    return json.loads(response_text)
  except json.JSONDecodeError:
    start = response_text.find("{")
    end = response_text.rfind("}")
    if start != -1 and end != -1 and start < end:
      return json.loads(response_text[start:end + 1])
    raise


def extract_field_value(block_text, field_name):
  pattern = re.compile(
    rf"{re.escape(field_name)}\s*[:：]?\s*(.+?)(?=\s+(?:요구사항ID|화면ID|화면명|화면설명|메뉴경로|개발구분|개인정보등급|화면구성|처리흐름|이벤트\s*정의)\b|$)",
    re.DOTALL,
  )
  match = pattern.search(block_text)
  if not match:
    return ""
  return clean_extracted_text_value(match.group(1))


def is_bad_generated_title(value, screen_id, unit_test_id):
  text = str(value or "").strip()
  if not text:
    return True
  if text in {str(screen_id or "").strip(), str(unit_test_id or "").strip()}:
    return True
  screen_parts = [part for part in str(screen_id or "").split("-") if part]
  unit_parts = [part for part in str(unit_test_id or "").split("-") if part]
  return text in set(screen_parts + unit_parts)


def strip_leading_sequence_marker(value):
  text = str(value or "").strip()
  if not text:
    return ""
  text = re.sub(r"^\s*\d{1,3}\s*[\.\)]\s*", "", text).strip()
  circled_pattern = re.compile(rf"^\s*[{re.escape(CIRCLED_NUMBERS)}]\s*")
  text = circled_pattern.sub("", text).strip()
  return text


def normalize_test_cases(parsed_json, screen_id, unit_test_id, fallback_screen_name="", unit_test_title=""):
  if not isinstance(parsed_json, dict):
    raise ValueError("응답이 JSON 객체가 아닙니다.")

  if "test_cases" not in parsed_json:
    raise ValueError("최상위 키 'test_cases'가 없습니다.")

  test_cases = parsed_json["test_cases"]

  if isinstance(test_cases, dict):
    test_cases = [test_cases]

  if not isinstance(test_cases, list):
    raise ValueError("'test_cases'가 배열이 아닙니다.")

  normalized = []
  for idx, case in enumerate(test_cases, 1):
    if not isinstance(case, dict):
      raise ValueError(f"{idx}번째 test case가 객체가 아닙니다.")

    fixed = {}
    for key in REQUIRED_TC_KEYS:
      fixed[key] = case.get(key, "")

    fixed["화면_ID"] = screen_id or fixed.get("화면_ID", "")
    fixed["단위시험_ID"] = unit_test_id or fixed.get("단위시험_ID", "")
    fixed["테스트_데이터"] = ""
    fixed["테스트_케이스"] = strip_leading_sequence_marker(fixed.get("테스트_케이스", ""))
    fixed["예상_결과"] = strip_leading_sequence_marker(fixed.get("예상_결과", ""))
    if fallback_screen_name:
      if is_bad_generated_title(fixed.get("화면명"), screen_id, unit_test_id):
        fixed["화면명"] = fallback_screen_name

    title = unit_test_title or fallback_screen_name
    if title:
      fixed["단위시험_제목"] = title
      fixed["단위시험_명"] = title

    if not fixed["순서"]:
      fixed["순서"] = idx

    normalized.append(fixed)

  return normalized

def get_missing_sequences(test_cases, expected_steps):
  if not expected_steps:
    return []

  existing = set()
  for case in test_cases:
    try:
      existing.add(int(case.get("순서", 0)))
    except (TypeError, ValueError):
      continue

  return [seq for seq in range(1, expected_steps + 1) if seq not in existing]

def sort_and_dedupe_cases(test_cases):
  seen = set()
  result = []

  def sequence_value(case):
    try:
      return int(case.get("순서", 0))
    except (TypeError, ValueError):
      return 0

  for case in sorted(test_cases, key=sequence_value):
    seq = sequence_value(case)
    if seq in seen:
      continue
    seen.add(seq)
    result.append(case)

  return result


def attach_block_original_index(test_cases, original_index):
  for case in test_cases:
    case["__block_original_index"] = original_index
  return test_cases


def restore_original_case_order(test_cases):
  def sort_key(case):
    try:
      block_index = int(case.get("__block_original_index", 0))
    except (TypeError, ValueError):
      block_index = 0
    try:
      sequence = int(case.get("순서", 0))
    except (TypeError, ValueError):
      sequence = 0
    return block_index, sequence

  ordered = sorted(test_cases, key=sort_key)
  for case in ordered:
    case.pop("__block_original_index", None)
  return ordered


def call_ollama(ollama_url, model_name, system_prompt, user_prompt, num_predict=1024, timeout=60):
  if not ollama_url:
    raise ValueError("OLLAMA_URL 값이 비어 있습니다.")

  payload = {
    "model": model_name,
    "messages": [
      {"role": "system", "content": system_prompt},
      {"role": "user", "content": user_prompt}
    ],
    "format": "json",
    "stream": False,
    "options": {
      "num_predict": num_predict
    }
  }

  response = requests.post(ollama_url, json=payload, timeout=timeout)
  response.raise_for_status()
  result_data = response.json()
  return result_data.get("message", {}).get("content", "")

def get_llm_limits(expected_steps, block_text):
  if expected_steps >= 9 or len(block_text) >= 3000:
    return 4096, 180
  if expected_steps >= 4 or len(block_text) >= 1500:
    return 2048, 90
  return 1024, 60


def classify_block_size(expected_steps, block_text):
  if expected_steps >= 9 or len(block_text) >= 3000:
    return "large"
  if expected_steps >= 4 or len(block_text) >= 1500:
    return "medium"
  return "small"


def block_processing_key(block):
  basis_info = get_test_case_generation_basis(block["text"])
  block_size = classify_block_size(int(basis_info.get("expected_steps") or 0), block["text"])
  return (
    SIZE_ORDER.get(block_size, SIZE_ORDER["large"]),
    int(block.get("_original_index") or 0),
  )

def build_retry_prompt(screen_id, unit_test_id, expected_steps, block_text, bad_response, error_message):
  basis_info = get_test_case_generation_basis(block_text)
  event_steps = basis_info["event_steps"]
  flow_steps = basis_info["flow_steps"]
  action_flow_steps = basis_info["action_flow_steps"]
  action_candidates = basis_info["action_candidates"]
  basis_label = {
    "event": "이벤트 정의",
    "event_flow_compare": "이벤트 정의 + 처리흐름 비교",
    "flow": "처리흐름",
  }.get(str(basis_info.get("basis")), "설계서")
  event_steps_text = "\n".join(
    f"{step['순서']}. {step.get('이벤트명', '')} - {step.get('처리설명', '')}".strip(" -")
    for step in event_steps
  )
  action_flow_steps_text = "\n".join(
    f"{step['순서']}. {step['내용']}" for step in action_flow_steps
  )
  flow_steps_text = "\n".join(
    f"{step['순서']}. {step['내용']}" for step in flow_steps
  )
  action_candidates_text = "\n".join(
    f"{idx}. {candidate.get('내용', '')}" for idx, candidate in enumerate(action_candidates, 1)
  )
  return f"""
  이전 응답은 사용할 수 없습니다.

  [오류]
  {error_message}

  [잘못된 이전 응답]
  {bad_response}

  아래 화면 정의를 다시 분석해 순수 JSON 객체만 출력하세요.
  최상위 키는 반드시 "test_cases" 하나만 사용하세요.

  [대상 화면ID]
  {screen_id}

  [대상 단위시험_ID]
  {unit_test_id}

  [주 작성 기준]
  {basis_label}

  [기준 단계 수]
  {expected_steps if expected_steps else "추출 안 됨"}

  [이벤트 정의 목록]
  {event_steps_text if event_steps_text else "이벤트 정의 목록을 추출하지 못했습니다."}

  [처리흐름 중 사용자 조작 단계]
  {action_flow_steps_text if action_flow_steps_text else "처리흐름에서 사용자 조작 단계를 별도로 추출하지 못했습니다."}

  [이벤트 정의와 처리흐름을 비교해 구성한 사용자 조작 후보]
  {action_candidates_text if action_candidates_text else "사용자 조작 후보를 추출하지 못했습니다."}

  [전체 처리흐름 - 예상 결과 보강 참고]
  {flow_steps_text if flow_steps_text else "처리흐름 목록을 추출하지 못했습니다."}

  [작성 기준]
  - 이벤트 정의와 처리흐름을 항상 비교해 최종 사용자 조작 목록을 구성하세요.
  - 이벤트 정의와 처리흐름이 같은 조작을 말하면 하나의 test case로 합치세요.
  - 처리흐름에만 있는 사용자 조작은 누락하지 말고 test case로 보완하세요.
  - 이벤트명과 처리흐름의 사용자 조작은 "테스트_케이스"로 작성하세요.
  - 시스템 반응, 조회 결과, 팝업 호출, 확인/검증 내용은 "예상_결과"로 작성하세요.
  - "테스트_데이터"는 출력하지 마세요. 시스템에서 빈 값으로 채웁니다.
  - 화면_ID와 단위시험_ID는 아래 값 그대로 사용하세요.

  [출력 JSON]
  {{
    "test_cases": [
      {{
        "단위시험_ID": "{unit_test_id}",
        "화면명": "",
        "단위시험_명": "",
        "사전조건": "",
        "화면_ID": "{screen_id}",
        "순서": 1,
        "테스트_케이스": "",
        "예상_결과": ""
      }}
    ]
  }}

  [설계서 화면 정의 블록]
  {block_text}
  """

def _check_cancel(cancel_check: Callable[[], None] | None) -> None:
  if cancel_check:
    cancel_check()


def _report_progress(progress_callback: Callable[[str], None] | None, message: str) -> None:
  if progress_callback:
    progress_callback(message)


def _report_block_status(block_status_callback: Callable[[dict], None] | None, update: dict) -> None:
  if block_status_callback:
    block_status_callback(update)


def build_test_cases_from_text(
    extracted_text,
    model_name,
    ollama_url,
    screen_blocks=None,
    cancel_check: Callable[[], None] | None = None,
    progress_callback: Callable[[str], None] | None = None,
    block_status_callback: Callable[[dict], None] | None = None,
):
  print(f"\nAI 추론 중... ({model_name})")

  system_prompt = """
  당신은 UI 설계서의 화면 정의를 단위시험 케이스로 정리하는 QA 엔지니어입니다.

  이벤트 정의와 처리흐름을 항상 비교해 사용자 조작 단위로 test_cases를 작성하세요.
  이벤트 정의와 처리흐름이 같은 조작을 말하면 하나로 합치고, 처리흐름에만 있는 사용자 조작은 보완하세요.
  시스템 반응, 조회 결과, 팝업 호출처럼 사용자가 직접 수행하지 않는 항목은 별도 케이스로 만들지 말고 "예상_결과"에 반영하세요.
  처리흐름의 확인/검증/조회 결과는 테스트 케이스가 아니라 관련 이벤트의 "예상_결과"에 반영하세요.

  작성 필드:
  - 화면명
  - 단위시험_명
  - 사전조건
  - 순서
  - 테스트_케이스
  - 예상_결과

  화면_ID, 단위시험_ID, 테스트_데이터, 수행자, 수행_일자, 수행_결과는 시스템에서 보정합니다.
  "테스트_데이터" 필드는 출력하지 마세요.
  반드시 마크다운 없이 순수 JSON 객체만 출력하고, 최상위 키는 "test_cases"만 사용하세요.
  """

  screen_blocks = screen_blocks if screen_blocks is not None else extract_screen_blocks(extracted_text)
  if not screen_blocks:
    print("화면ID를 찾지 못해 전체 설계서를 한 번에 처리합니다.")
    screen_blocks = [{
      "screen_id": "",
      "unit_test_id": "",
      "text": extracted_text
    }]

  screen_blocks = [
    {**block, "_original_index": index}
    for index, block in enumerate(screen_blocks)
  ]
  screen_blocks = sorted(screen_blocks, key=block_processing_key)

  print(f"\n추출된 고유 화면ID 수: {len(screen_blocks)}")
  for idx, block in enumerate(screen_blocks, 1):
    basis_info = get_test_case_generation_basis(block["text"])
    expected_steps = int(basis_info.get("expected_steps") or 0)
    num_predict, request_timeout = get_llm_limits(expected_steps, block["text"])
    block_size = classify_block_size(expected_steps, block["text"])
    original_index = int(block.get("_original_index", idx - 1))
    _report_block_status(block_status_callback, {
      "status": "queued",
      "screen_id": block.get("screen_id", ""),
      "unit_test_id": block.get("unit_test_id", ""),
      "original_index": original_index,
      "display_index": original_index + 1,
      "processing_index": idx,
      "expected_steps": expected_steps,
      "block_size": block_size,
      "num_predict": num_predict,
      "timeout": request_timeout,
      "generated_count": 0,
      "error": "",
    })
    print(
      f"[화면 분할] {idx}/{len(screen_blocks)} screen_id={block['screen_id']} "
      f"basis={basis_info.get('basis')} expected_steps={expected_steps} text_len={len(block['text'])}"
    )
    for step in basis_info["event_steps"]:
      print(f"  - 이벤트 정의 {step['순서']}: {step.get('이벤트명', '')} | {step.get('처리설명', '')}")
    for step in basis_info["flow_steps"]:
      print(f"  - 처리흐름 {step['순서']}: {step['내용']}")

  _report_progress(progress_callback, f"단위시험케이스 AI 생성 시작 | 화면 블록 {len(screen_blocks)}개")
  all_test_cases = []

  for idx, block in enumerate(screen_blocks, 1):
    _check_cancel(cancel_check)
    basis_info = get_test_case_generation_basis(block["text"])
    event_steps = basis_info["event_steps"]
    flow_steps = basis_info["flow_steps"]
    action_flow_steps = basis_info["action_flow_steps"]
    action_candidates = basis_info["action_candidates"]
    expected_steps = int(basis_info.get("expected_steps") or 0)
    screen_id = block["screen_id"]
    unit_test_id = block["unit_test_id"]
    fallback_screen_name = extract_field_value(block["text"], "화면명")
    unit_test_title = str(block.get("section_title") or "").strip() or extract_block_section_title(block["text"]) or fallback_screen_name
    response_text = ""
    num_predict, request_timeout = get_llm_limits(expected_steps, block["text"])
    block_size = classify_block_size(expected_steps, block["text"])
    original_index = int(block.get("_original_index", idx - 1))
    block_status_base = {
      "screen_id": screen_id,
      "unit_test_id": unit_test_id,
      "original_index": original_index,
      "display_index": original_index + 1,
      "processing_index": idx,
      "expected_steps": expected_steps,
      "block_size": block_size,
      "num_predict": num_predict,
      "timeout": request_timeout,
    }
    basis_label = {
      "event": "이벤트 정의",
      "event_flow_compare": "이벤트 정의 + 처리흐름 비교",
      "flow": "처리흐름",
    }.get(str(basis_info.get("basis")), "설계서")
    event_steps_text = "\n".join(
      f"{step['순서']}. {step.get('이벤트명', '')} - {step.get('처리설명', '')}".strip(" -")
      for step in event_steps
    )
    action_flow_steps_text = "\n".join(
      f"{step['순서']}. {step['내용']}" for step in action_flow_steps
    )
    flow_steps_text = "\n".join(
      f"{step['순서']}. {step['내용']}" for step in flow_steps
    )
    action_candidates_text = "\n".join(
      f"{candidate_idx}. {candidate.get('내용', '')}" for candidate_idx, candidate in enumerate(action_candidates, 1)
    )

    user_prompt = f"""
    아래 화면 정의 블록으로 단위시험 케이스를 작성하세요.

    [대상 화면ID - 참고용]
    {screen_id}

    [대상 단위시험_ID - 참고용]
    {unit_test_id}

    [주 작성 기준]
    {basis_label}

    [코드가 추출한 이벤트 정의 목록]
    {event_steps_text if event_steps_text else "이벤트 정의 목록을 추출하지 못했습니다."}

    [처리흐름 중 사용자 조작 단계]
    {action_flow_steps_text if action_flow_steps_text else "처리흐름에서 사용자 조작 단계를 별도로 추출하지 못했습니다."}

    [이벤트 정의와 처리흐름을 비교해 구성한 사용자 조작 후보]
    {action_candidates_text if action_candidates_text else "사용자 조작 후보를 추출하지 못했습니다."}

    [코드가 추출한 전체 처리흐름 목록 - 예상 결과 보강 참고]
    {flow_steps_text if flow_steps_text else "처리흐름 목록을 추출하지 못했습니다. 설계서 화면 정의 블록에서 직접 판단하세요."}

    [작성 기준]
    - 이벤트 정의와 처리흐름을 항상 비교해 최종 사용자 조작 목록을 구성하세요.
    - 이벤트 정의와 처리흐름이 같은 조작을 말하면 하나의 test case로 합치세요.
    - 처리흐름에만 있는 사용자 조작은 누락하지 말고 test case로 보완하세요.
    - 이벤트명과 처리흐름의 사용자 조작은 "테스트_케이스"에 작성하세요.
    - 조회됨, 표시됨, 팝업 호출, 확인/검증 내용 같은 시스템 반응은 "예상_결과"에 작성하세요.
    - 처리흐름의 확인/검증 단계만 별도 테스트 케이스로 만들지 말고 관련 조작의 예상 결과에 합치세요.
    - "테스트_데이터"는 출력하지 마세요. 시스템에서 빈 값으로 채웁니다.
    - 순수 JSON 객체만 출력하세요.

    [JSON 형식]
    {{"test_cases":[{{"화면명":"","단위시험_명":"","사전조건":"","순서":1,"테스트_케이스":"","예상_결과":""}}]}}

    [설계서 화면 정의 블록]:
    {block["text"]}
    """

    try:
      print(f"\n[AI 호출] {idx}/{len(screen_blocks)} {screen_id} 처리 중...")
      print(f"[AI 호출 설정] {screen_id}: expected_steps={expected_steps} text_len={len(block['text'])} num_predict={num_predict} timeout={request_timeout}s")
      _check_cancel(cancel_check)
      _report_progress(
        progress_callback,
        f"단위시험케이스 AI 호출 중 | {idx}/{len(screen_blocks)} 화면={screen_id or '-'} num_predict={num_predict} timeout={request_timeout}s",
      )
      _report_block_status(block_status_callback, {
        **block_status_base,
        "status": "running",
        "generated_count": 0,
        "error": "",
      })
      response_text = call_ollama(ollama_url, model_name, system_prompt, user_prompt, num_predict=num_predict, timeout=request_timeout)
      _check_cancel(cancel_check)
      parsed_json = parse_llm_json(response_text)
      normalized_cases = normalize_test_cases(parsed_json, screen_id, unit_test_id, fallback_screen_name, unit_test_title)

      normalized_cases = sort_and_dedupe_cases(normalized_cases)

      print(f"[완료] {screen_id}: 생성 {len(normalized_cases)}개")

      _report_progress(
        progress_callback,
        f"단위시험케이스 AI 응답 완료 | {idx}/{len(screen_blocks)} 화면={screen_id or '-'} 생성 {len(normalized_cases)}건",
      )
      _report_block_status(block_status_callback, {
        **block_status_base,
        "status": "updated",
        "generated_count": len(normalized_cases),
        "error": "",
      })
      all_test_cases.extend(attach_block_original_index(normalized_cases, block.get("_original_index", idx - 1)))
    
    except requests.exceptions.Timeout:
      message = f"단위시험케이스 AI 응답 시간 초과 | 화면={screen_id or '-'} | timeout={request_timeout}s"
      print(message)
      print("AI 응답 시간 초과로 전체 테스트 케이스 생성을 중단합니다.")
      _report_progress(progress_callback, message)
      _report_block_status(block_status_callback, {
        **block_status_base,
        "status": "error",
        "generated_count": 0,
        "error": message,
      })
      raise TestCaseGenerationError(message)
    except requests.exceptions.RequestException as e:
      message = f"단위시험케이스 AI 호출 실패 | 화면={screen_id or '-'} | {e}"
      print(message)
      print("AI 호출 실패로 전체 테스트 케이스 생성을 중단합니다.")
      _report_progress(progress_callback, message)
      _report_block_status(block_status_callback, {
        **block_status_base,
        "status": "error",
        "generated_count": 0,
        "error": message,
      })
      raise TestCaseGenerationError(message)
    except (json.JSONDecodeError, ValueError) as e:
      bad_response = response_text if "response_text" in locals() else ""
      print(f"[응답 검증 실패] {screen_id}: {e}")
      print(f"Raw Output: {bad_response}")

      retry_prompt = build_retry_prompt(
        screen_id=screen_id,
        unit_test_id=unit_test_id,
        expected_steps=expected_steps,
        block_text=block["text"],
        bad_response=bad_response,
        error_message=str(e)
      )

      try:
        print(f"[AI 재호출] {screen_id} 응답 형식 보정 중...")
        _check_cancel(cancel_check)
        _report_progress(
          progress_callback,
          f"단위시험케이스 AI 응답 보정 호출 중 | 화면={screen_id or '-'} timeout={request_timeout}s",
        )
        retry_text = call_ollama(ollama_url, model_name, system_prompt, retry_prompt, num_predict=num_predict, timeout=request_timeout)
        _check_cancel(cancel_check)
        retry_json = parse_llm_json(retry_text)
        normalized_cases = normalize_test_cases(retry_json, screen_id, unit_test_id, fallback_screen_name, unit_test_title)

        normalized_cases = sort_and_dedupe_cases(normalized_cases)
        print(f"[재호출 완료] {screen_id}: 생성 {len(normalized_cases)}개")
        _report_progress(
          progress_callback,
          f"단위시험케이스 AI 응답 보정 완료 | 화면={screen_id or '-'} 생성 {len(normalized_cases)}건",
        )
        _report_block_status(block_status_callback, {
          **block_status_base,
          "status": "updated",
          "generated_count": len(normalized_cases),
          "error": "",
        })
        all_test_cases.extend(attach_block_original_index(normalized_cases, block.get("_original_index", idx - 1)))
      except Exception as retry_error:
        message = f"단위시험케이스 AI 응답 보정 실패 | 화면={screen_id or '-'} | {retry_error}"
        print(f"[재호출 실패] {screen_id}: {retry_error}")
        print("응답 보정 실패로 전체 테스트 케이스 생성을 중단합니다.")
        _report_progress(progress_callback, message)
        _report_block_status(block_status_callback, {
          **block_status_base,
          "status": "error",
          "generated_count": 0,
          "error": message,
        })
        raise TestCaseGenerationError(message)

  all_test_cases = restore_original_case_order(all_test_cases)
  print(f"\n전체 생성 테스트 케이스 행 수: {len(all_test_cases)}")
  return all_test_cases
  
def save_test_cases_to_excel(test_cases, output_dir: Path, base_filename="generated_TC", performer=""):
  if not test_cases:
    print("저장할 데이터가 없습니다.")
    return
  
  output_dir = Path(output_dir)
  output_dir.mkdir(parents=True, exist_ok=True)
  
  # 파일명 자동 증가
  counter = 1
  while True:
    output_filename = f"{base_filename}_{counter}.xlsx"
    full_path = output_dir / output_filename
    if not full_path.exists():
      break
    counter += 1

  print(f"\n엑셀 생성 중...")

  wb = Workbook()
  is_first_group = True

  # 스타일 정의
  thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
  header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
  center_align = Alignment(horizontal="center", vertical="center")
  left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

  # 단위시험 ID를 기준으로 그룹화
  grouped_tc = {}
  for tc in test_cases:
    if not isinstance(tc, dict):
      continue
    tc_id = tc.get("단위시험_ID", "미분류")
    if tc_id not in grouped_tc:
      grouped_tc[tc_id] = []
    grouped_tc[tc_id].append(tc)

  for group_idx, (tc_id, group) in enumerate(grouped_tc.items(), 1):
    if not group:
      continue

    # 시트 생성/선택 로직
    if is_first_group:
      ws = wb.active
      ws.title = str(tc_id)[:31]
      is_first_group = False
    else:
      ws = wb.create_sheet(title=str(tc_id)[:31])

    current_row = 1
    common = group[0]
    display_title = common.get("단위시험_제목") or common.get("단위시험_명") or common.get("화면명", "")

    # -- 타이틀 --
    title = f"{group_idx}. {tc_id} - {display_title}"
    ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    current_row += 1

    # -- 상단 공통 영역 --
    # 1행: [단위시험]
    ws.cell(row=current_row, column=1, value="단위시험").font = Font(bold=True)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)

    # 2행: 단위시험 ID / 수행자
    r2 = current_row + 1
    ws.cell(row=r2, column=1, value="단위시험 ID")
    ws.cell(row=r2, column=2, value=common.get("단위시험_ID", ""))
    ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=3)
    ws.cell(row=r2, column=4, value="수행자")
    ws.cell(row=r2, column=5, value=performer)

    # 3행: 단위시험 명 / 수행 일자
    r3 = current_row + 2
    ws.cell(row=r3, column=1, value="단위시험 명")
    ws.cell(row=r3, column=2, value=common.get("단위시험_명", ""))
    ws.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=3)
    ws.cell(row=r3, column=4, value="수행 일자")
    ws.cell(row=r3, column=5, value="")

    # 4행: 사전조건 / 화면 ID
    r4 = current_row + 3
    ws.cell(row=r4, column=1, value="사전조건")
    ws.cell(row=r4, column=2, value=common.get("사전조건", ""))
    ws.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=3)
    ws.cell(row=r4, column=4, value="화면 ID")
    ws.cell(row=r4, column=5, value=common.get("화면_ID", ""))

    # 상단 공통 영역 스타일 일괄 적용
    for row_idx in range(current_row, current_row + 4):
      for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if col_idx == 1 or col_idx == 4 or row_idx == current_row:
          cell.fill = header_fill
          cell.alignment = center_align
        else:
          cell.alignment = left_align

    # -- 하단 테스트 스텝 영역 --
    # 5행: 테이블 헤더
    header_row = current_row + 4
    headers = ["순서", "테스트 케이스", "테스트 데이터", "예상 결과", "수행 결과"]
    for h_idx, header in enumerate(headers):
      cell = ws.cell(row=header_row, column=h_idx + 1, value=header)
      cell.alignment = center_align
      cell.fill = header_fill
      cell.font = Font(bold=True)
      cell.border = thin_border

    # 6행~: 실제 그룹의 테스트 스텝 데이터
    data_start_row = current_row + 5
    for i, tc in enumerate(group):
      r = data_start_row + i
      ws.cell(row=r, column=1, value=str(i + 1)).alignment = center_align
      ws.cell(row=r, column=2, value=str(tc.get("테스트_케이스", ""))).alignment = left_align
      ws.cell(row=r, column=3, value="")
      ws.cell(row=r, column=4, value=str(tc.get("예상_결과", ""))).alignment = left_align
      ws.cell(row=r, column=5, value=str(tc.get("수행_결과", ""))).alignment = left_align

      # 해당 행에 테두리 적용
      for col_num in range(1, 6):
        ws.cell(row=r, column=col_num).border = thin_border

    # 컬럼 너비 세밀 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20

  wb.save(full_path)
  print(f"엑셀 저장 완료: {full_path.resolve()}")
  return full_path

def show_hwpx_headers_on_test_pages(hwpx_path):
  """생성된 HWPX에서 표지 이후 페이지의 머리말/꼬리말 숨김 설정만 해제합니다."""
  if not hwpx_path or not os.path.exists(hwpx_path):
    return

  output_dir = os.path.dirname(os.path.abspath(hwpx_path))
  temp_path = os.path.join(output_dir, f".tmp_{os.path.basename(hwpx_path)}")
  changed = False
  seen_cover_page_hiding = False

  with zipfile.ZipFile(hwpx_path, "r") as src, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as dst:
    for item in src.infolist():
      data = src.read(item.filename)

      if item.filename.startswith("Contents/section") and item.filename.endswith(".xml"):
        xml = data.decode("utf-8")

        def reveal_header_after_cover(match):
          nonlocal changed, seen_cover_page_hiding
          page_hiding = match.group(0)

          if not seen_cover_page_hiding:
            seen_cover_page_hiding = True
            return page_hiding

          updated = page_hiding.replace('hideHeader="1"', 'hideHeader="0"')
          updated = updated.replace('hideFooter="1"', 'hideFooter="0"')
          if updated != page_hiding:
            changed = True
          return updated

        xml = re.sub(
          r'<hp:pageHiding\b[^>]*/>',
          reveal_header_after_cover,
          xml,
          flags=re.S,
        )
        data = xml.encode("utf-8")

      dst.writestr(item, data)

  if changed:
    os.replace(temp_path, hwpx_path)
  elif os.path.exists(temp_path):
    os.remove(temp_path)

def move_blank_page_breaks_to_test_titles(hwpx_path):
  """단위시험 제목 앞의 빈 페이지 나누기 문단을 제거하고 제목 문단이 새 페이지를 시작하게 합니다."""
  if not hwpx_path or not os.path.exists(hwpx_path):
    return

  output_dir = os.path.dirname(os.path.abspath(hwpx_path))
  temp_path = os.path.join(output_dir, f".tmp_pagebreak_{os.path.basename(hwpx_path)}")
  changed = False
  paragraph_pattern = re.compile(r'<hp:p\b[\s\S]*?</hp:p>')

  def paragraph_text(paragraph):
    return " ".join(re.findall(r'<hp:t(?:\b[^>]*)?>(.*?)</hp:t>', paragraph)).strip()

  def is_blank_page_break(paragraph):
    return 'pageBreak="1"' in paragraph and not paragraph_text(paragraph)

  def is_test_title(paragraph):
    text = paragraph_text(paragraph)
    return text.startswith("UT-") and " - " in text

  with zipfile.ZipFile(hwpx_path, "r") as src, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as dst:
    for item in src.infolist():
      data = src.read(item.filename)

      if item.filename.startswith("Contents/section") and item.filename.endswith(".xml"):
        xml = data.decode("utf-8")
        parts = []
        last_end = 0
        paragraphs = list(paragraph_pattern.finditer(xml))
        idx = 0

        while idx < len(paragraphs):
          match = paragraphs[idx]
          parts.append(xml[last_end:match.start()])

          current = match.group(0)
          next_match = paragraphs[idx + 1] if idx + 1 < len(paragraphs) else None

          if next_match and is_blank_page_break(current) and is_test_title(next_match.group(0)):
            next_paragraph = next_match.group(0).replace('pageBreak="0"', 'pageBreak="1"', 1)
            parts.append(xml[match.end():next_match.start()])
            parts.append(next_paragraph)
            last_end = next_match.end()
            idx += 2
            changed = True
            continue

          parts.append(current)
          last_end = match.end()
          idx += 1

        parts.append(xml[last_end:])
        xml = "".join(parts)
        data = xml.encode("utf-8")

      dst.writestr(item, data)

  if changed:
    os.replace(temp_path, hwpx_path)
  elif os.path.exists(temp_path):
    os.remove(temp_path)

def save_test_cases_to_hwpx(test_cases, temp_path, output_filename, performer="", clear_execution_date=False):
  if not test_cases:
    return None
  
  pythoncom.CoInitialize()
  hwp = Hwp(visible=False)

  print(f"\n한글 생성 중...")

  try:
    abs_temp_path = os.path.abspath(temp_path)
    abs_output_path = os.path.abspath(output_filename)

    hwp.open(abs_temp_path)
    time.sleep(0.5)

    def find_text(s):
      hwp.HAction.GetDefault("RepeatFind", hwp.HParameterSet.HFindReplace.HSet)
      hwp.HParameterSet.HFindReplace.FindString = s
      hwp.HParameterSet.HFindReplace.IgnoreMessage = 1
      return hwp.HAction.Execute("RepeatFind", hwp.HParameterSet.HFindReplace.HSet)
    
    def clear_and_write(text):
      hwp.Run("Cancel")
      hwp.Run("SelectAll")
      hwp.Run("Delete")
      if text:
        hwp.insert_text(str(text))
      hwp.Run("Cancel")

    def move_to_first_test_body_after_toc(max_pages=5):
      hwp.Run("MoveDocBegin")
      if not (find_text("목 차") or find_text("<목 차>")):
        return False

      for _ in range(max_pages):
        hwp.Run("MovePageDown")
        hwp.Run("MovePageBegin")
        if find_text("단위시험 ID"):
          hwp.Run("Cancel")
          return True
      hwp.Run("Cancel")
      return False

    def move_to_test_step_data_row():
      hwp.Run("MovePageBegin")
      for header in ("순서", "순번"):
        if not find_text(header):
          continue
        if hwp.TableLowerCell():
          hwp.Run("TableColBegin")
          return True
        hwp.Run("Cancel")
        hwp.Run("MovePageBegin")
      return False

    hwp.Run("MoveDocBegin")
    if find_text("목 차") or find_text("<목 차>"):
      hwp.Run("Cancel")
      if move_to_first_test_body_after_toc():
        if find_text("단위시험 ID"):
          hwp.Run("Cancel")
        hwp.Run("MoveUp"); hwp.Run("MoveUp"); hwp.Run("MoveUp")
        hwp.Run("MoveLineBegin")
        hwp.Run("Select")
        hwp.Run("MoveDocEnd")
        hwp.Run("Delete")
        time.sleep(0.2)

    try:
      # 클립보드 강제 초기화
      ctypes.windll.user32.OpenClipboard(None)
      ctypes.windll.user32.EmptyClipboard()
      ctypes.windll.user32.CloseClipboard()
    except: pass

    hwp.Run("MoveDocBegin")
    if find_text("단위시험 ID"):
      hwp.Run("Cancel")
      hwp.Run("MoveUp"); hwp.Run("MoveUp"); hwp.Run("MoveUp")
      hwp.Run("MoveLineBegin")
      hwp.Run("Select")
      hwp.Run("MoveDocEnd")
      hwp.Run("Copy")
      time.sleep(0.3)
      hwp.Run("Cancel")
    else:
      raise RuntimeError("HWPX 양식에서 '단위시험 ID' 영역을 찾을 수 없습니다.")

    # 데이터 그룹화
    grouped_tc = {}
    for step in test_cases:
      tc_id = step.get("단위시험_ID", "")
      if tc_id not in grouped_tc:
        grouped_tc[tc_id] = []
      grouped_tc[tc_id].append(step)

    num_cases = len(grouped_tc)

    if num_cases > 1:
      for _ in range(num_cases - 1):
        hwp.Run("MoveDocEnd")
        hwp.Run("BreakPage")
        hwp.Run("MovePageBegin")
        hwp.Run("Paste")
        time.sleep(0.1)

    hwp.Run("MoveDocBegin")
    for tc_idx, (tc_id, steps) in enumerate(grouped_tc.items(), start=1):
      first_step = steps[0]

      if find_text("단위시험 ID"):
        hwp.Run("Cancel")
        hwp.Run("MoveUp"); hwp.Run("MoveUp")
        hwp.Run("MoveLineBegin"); hwp.Run("Select"); hwp.Run("MoveLineEnd"); hwp.Run("Delete")
        hwp.insert_text(f"{tc_id} - {first_step.get('단위시험_명','')}")

        find_text("단위시험 ID"); hwp.Run("TableRightCell"); clear_and_write(tc_id)
        if find_text("수행자"):
          hwp.Run("TableRightCell"); clear_and_write(performer)
        find_text("단위시험 명"); hwp.Run("TableRightCell"); clear_and_write(first_step.get("단위시험_명",""))
        if clear_execution_date and find_text("수행 일자"):
          hwp.Run("TableRightCell"); clear_and_write("")
        find_text("사전조건"); hwp.Run("TableRightCell"); clear_and_write(first_step.get("사전조건",""))
        find_text("화면 ID"); hwp.Run("TableRightCell"); clear_and_write(first_step.get("화면_ID",""))

        if not move_to_test_step_data_row():
          raise RuntimeError(f"HWPX 양식에서 테스트 스텝 데이터 행으로 이동할 수 없습니다: {tc_id}")

        for col in range(5):
          clear_and_write("")
          if col < 4: hwp.Run("TableRightCell")
        
        hwp.Run("TableColBegin")
        rows_to_delete = 0
        while hwp.TableLowerCell():
          rows_to_delete += 1

        for _ in range(rows_to_delete):
          hwp.Run("TableDeleteRow")

        hwp.Run("TableColBegin")

        for i, step in enumerate(steps):
          if i > 0: hwp.Run("TableAppendRow")
          hwp.Run("TableColBegin")
          clear_and_write(str(step.get("순서", i+1)))
          hwp.Run("TableRightCell"); clear_and_write(step.get("테스트_케이스", ""))
          hwp.Run("TableRightCell"); clear_and_write("")
          hwp.Run("TableRightCell"); clear_and_write(step.get("예상_결과", ""))
          hwp.Run("TableRightCell"); clear_and_write(step.get("수행_결과", ""))

        hwp.Run("Cancel"); hwp.Run("MoveRight")
      else:
        raise RuntimeError(f"HWPX 양식에서 채울 단위시험 표를 찾을 수 없습니다: {tc_id}")

    hwp.save_as(abs_output_path)
  finally:
    try:
      if hwp:
        hwp.quit()
    except Exception as e:
      print(f"[경고] HWP 종료 중 오류: {e}")
    finally:
      pythoncom.CoUninitialize()

  move_blank_page_breaks_to_test_titles(abs_output_path)
  show_hwpx_headers_on_test_pages(abs_output_path)
  print(f"한글 저장 완료: {os.path.abspath(abs_output_path)}")
  return abs_output_path

def generate_test_cases(
    pdf_path: Path,
    model_name: str,
    ollama_url: str,
    output_dir: Path,
    template_path: Path,
    extracted_text: str | None = None,
    screen_blocks: list[dict] | None = None,
    cancel_check: Callable[[], None] | None = None,
    progress_callback: Callable[[str], None] | None = None,
    block_status_callback: Callable[[dict], None] | None = None,
) -> dict:
    pdf_path = Path(pdf_path)
    output_dir = Path(output_dir)
    template_path = Path(template_path)

    output_dir.mkdir(parents=True, exist_ok=True)

    _check_cancel(cancel_check)
    _report_progress(progress_callback, "단위시험케이스 템플릿 정보 확인 중")
    cover_author = extract_cover_author_from_document(template_path)
    _check_cancel(cancel_check)
    if extracted_text is None:
        _report_progress(progress_callback, f"설계서 텍스트 추출 중 | {pdf_path.name}")
    extract_text = extracted_text if extracted_text is not None else extract_text_from_pdf(pdf_path)
    _report_progress(progress_callback, f"설계서 텍스트 추출 완료 | {len(extract_text)}자")
    _check_cancel(cancel_check)
    try:
        tc_data = build_test_cases_from_text(
            extract_text,
            model_name,
            ollama_url,
            screen_blocks=screen_blocks,
            cancel_check=cancel_check,
            progress_callback=progress_callback,
            block_status_callback=block_status_callback,
        )
    except TestCaseGenerationError as exc:
        return {
            "ok": False,
            "error": str(exc),
            "files": [],
        }
    _check_cancel(cancel_check)

    if not tc_data:
        message = "단위시험케이스 생성 0건 | 설계서의 화면ID 또는 처리흐름을 찾지 못했거나 AI 응답에 유효한 test_cases가 없습니다."
        _report_progress(progress_callback, message)
        return {
            "ok": False,
            "error": message,
            "files": [],
        }

    _report_progress(progress_callback, f"단위시험케이스 XLSX 저장 중 | {len(tc_data)}건")
    excel_path = save_test_cases_to_excel(tc_data, output_dir, performer=cover_author)
    _check_cancel(cancel_check)

    files = []
    if excel_path:
        _report_progress(progress_callback, f"단위시험케이스 XLSX 저장 완료 | {excel_path.name}")
        files.append({
            "kind": "xlsx",
            "path": str(excel_path),
            "name": excel_path.name,
        })

    if HWP_AVAILABLE and template_path.exists():
        _check_cancel(cancel_check)
        hwpx_path = output_dir / f"generated_TC_{int(time.time())}.hwpx"
        _report_progress(progress_callback, f"단위시험케이스 HWPX 저장 중 | {template_path.name}")
        saved_hwpx = save_test_cases_to_hwpx(
            tc_data,
            temp_path=str(template_path),
            output_filename=str(hwpx_path),
            performer=cover_author,
        )
        if saved_hwpx:
            saved_hwpx = Path(saved_hwpx)
            _report_progress(progress_callback, f"단위시험케이스 HWPX 저장 완료 | {saved_hwpx.name}")
            files.append({
                "kind": "hwpx",
                "path": str(saved_hwpx),
                "name": saved_hwpx.name,
            })

    return {
        "ok": True,
        "count": len(tc_data),
        "files": files,
    }
