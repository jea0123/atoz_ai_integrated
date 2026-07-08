# check.html 결과 폴더를 기준으로 QA 산출물을 생성하고 기존 파일을 bak로 보관합니다.
from __future__ import annotations

from contextlib import redirect_stdout
from datetime import datetime
import filecmp
from pathlib import Path
import re
import shutil
import zipfile
from typing import Callable
from uuid import uuid4

from openpyxl import load_workbook

from app_runtime import RESULT_DIR, TEMP_DIR, log_event, log_message, remove_runtime_path
from cancellation import CancelledRequest
from document_update.hwpx_text import extract_document_text
from document_update.runtime_conversion import prepare_target_file
from qa_generation.generate_tc import (
    SIZE_LABELS,
    SIZE_ORDER,
    classify_document_size,
    extract_cover_author_from_document,
    extract_screen_blocks,
    extract_text_from_pdf,
    generate_test_cases,
)
from qa_generation.generate_ts import generate_integration_test_results, generate_test_scenarios
from web_uploads import safe_relative_upload_path


IGNORED_FOLDER_NAMES = {"bak", "backup", "백업"}
DESIGN_DOCUMENT_SUFFIXES = {".pdf", ".hwp", ".hwpx"}
TC_TEMPLATE_SUFFIXES = {".hwpx"}
TS_TEMPLATE_SUFFIXES = {".xlsx"}
QA_SOURCE_SUFFIXES = DESIGN_DOCUMENT_SUFFIXES | TC_TEMPLATE_SUFFIXES | TS_TEMPLATE_SUFFIXES
QA_UPLOAD_DUMP_DIRNAME = "qa-folder-dumps"
REQ_ID_PATTERN = re.compile(r"SFR-[A-Z0-9]+(?:-[A-Z0-9]+)*", re.IGNORECASE)
DESIGN_KEYWORDS = ("사용자인터페이스설계서", "사용자 인터페이스 설계서", "화면설계서", "화면정의서", "ui설계서")
TC_KEYWORDS = (
    "단위시험케이스",
    "단위시험 케이스",
    "단위테스트",
    "단위 테스트",
    "unittestcase",
    "unit test case",
)
TC_EXCLUDE_KEYWORDS = ("통합시험", "통합테스트", "시나리오", "결과서", "인수인계")
UNIT_RESULT_KEYWORDS = (
    "단위시험결과서",
    "단위시험 결과서",
    "단위테스트결과서",
    "단위 테스트 결과서",
    "unittestresult",
    "unit test result",
)
UNIT_RESULT_EXCLUDE_KEYWORDS = ("통합시험", "통합테스트", "시나리오", "케이스", "인수인계")
INTEGRATION_RESULT_KEYWORDS = (
    "통합시험결과서",
    "통합시험 결과서",
    "통합테스트결과서",
    "통합 테스트 결과서",
    "integrationtestresult",
    "integration test result",
)
INTEGRATION_RESULT_EXCLUDE_KEYWORDS = ("단위시험", "단위테스트", "케이스", "시나리오", "인수인계")


class QaFolderMatchingError(ValueError):
    def __init__(self, message: str, payload: dict[str, object]):
        super().__init__(message)
        self.payload = payload


def prepare_ui_design_for_processing(ui_design: Path, temp_dir: Path) -> tuple[Path, bool]:
    if ui_design.suffix.lower() == ".pdf":
        return ui_design, False
    return prepare_target_file(ui_design, temp_dir)


def qa_batch_log(request_id: str, message: str) -> None:
    log_message(f"QA 배치[{request_id}] {message}")


def qa_progress_log(request_id: str, requirement_id: str, message: str) -> None:
    if "단위시험케이스 AI 생성 시작" in message:
        match = re.search(r"화면 블록\s+(\d+)개", message)
        block_count = match.group(1) if match else "-"
        qa_batch_log(request_id, f"[{requirement_id}] 단위시험케이스 생성 시작 | 화면 {block_count}개")
        return

    if "단위시험케이스 AI 호출 중" in message:
        match = re.search(r"(\d+)/(\d+).*화면=([^|\s]+).*timeout=(\d+)s", message)
        if match:
            current, total, screen_id, timeout = match.groups()
            qa_batch_log(
                request_id,
                f"[{requirement_id}] 화면 처리 중 | {current}/{total} {screen_id} | timeout={timeout}초",
            )
        return

    if "단위시험케이스 AI 응답 완료" in message:
        match = re.search(r"(\d+)/(\d+).*화면=([^|\s]+).*생성\s+(\d+)건", message)
        if match:
            current, total, screen_id, count = match.groups()
            qa_batch_log(
                request_id,
                f"[{requirement_id}] 화면 처리 완료 | {current}/{total} {screen_id} | 생성 {count}건",
            )
        return

    if "단위시험케이스 AI 응답 보정 호출 중" in message:
        match = re.search(r"화면=([^|\s]+).*timeout=(\d+)s", message)
        if match:
            screen_id, timeout = match.groups()
            qa_batch_log(request_id, f"[{requirement_id}] 화면 응답 보정 중 | {screen_id} | timeout={timeout}초")
        return

    if "단위시험케이스 AI 응답 보정 완료" in message:
        match = re.search(r"화면=([^|\s]+).*생성\s+(\d+)건", message)
        if match:
            screen_id, count = match.groups()
            qa_batch_log(request_id, f"[{requirement_id}] 화면 응답 보정 완료 | {screen_id} | 생성 {count}건")
        return

    if "시간 초과" in message or "호출 실패" in message or "보정 실패" in message or "생성 0건" in message:
        qa_batch_log(request_id, f"[{requirement_id}] {message}")


def file_name(path: Path | str | None) -> str:
    return Path(str(path)).name if path else "-"


class DiscardedOutput:
    def write(self, value: str) -> int:
        return len(value)

    def flush(self) -> None:
        pass


def run_with_suppressed_output(callback):
    with redirect_stdout(DiscardedOutput()):
        return callback()


def create_uploaded_qa_source_dump(
        qa_source_items: list[tuple[str, bytes]] | None,
        dump_parent: Path | None = None,
) -> Path | None:
    # 브라우저 폴더 업로드는 원본 경로를 알 수 없으므로 서버 결과 폴더에 복사본을 만든다.
    entries: list[tuple[Path, bytes]] = []
    for index, (filename, payload) in enumerate(qa_source_items or [], start=1):
        if not payload:
            continue
        suffix = Path(filename).suffix.lower()
        if suffix not in QA_SOURCE_SUFFIXES:
            continue
        relative_path = safe_relative_upload_path(filename, f"qa-source-{index}{suffix}")
        entries.append((relative_path, payload))

    if not entries:
        return None

    top_parts = {
        relative_path.parts[0]
        for relative_path, _payload in entries
        if len(relative_path.parts) > 1
    }
    common_top = next(iter(top_parts)) if len(top_parts) == 1 else ""
    folder_name = common_top or "qa-upload"
    dump_parent = Path(dump_parent or (RESULT_DIR / QA_UPLOAD_DUMP_DIRNAME))
    dump_parent.mkdir(parents=True, exist_ok=True)
    dump_root = next_versioned_result_folder(dump_parent, folder_name)

    for relative_path, payload in entries:
        target_relative = Path(*relative_path.parts[1:]) if common_top and relative_path.parts[0] == common_top else relative_path
        if not target_relative.parts:
            target_relative = Path(relative_path.name)
        target_path = dump_root / target_relative
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(payload)

    return dump_root


def next_versioned_result_folder(parent: Path, folder_name: str) -> Path:
    safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", folder_name).strip(" ._") or "qa-upload"
    pattern = re.compile(rf"^{re.escape(safe_name)}_v0\.(\d+)$", re.IGNORECASE)
    highest_minor = 0
    for child in parent.iterdir():
        if not child.is_dir():
            continue
        match = pattern.fullmatch(child.name)
        if match:
            highest_minor = max(highest_minor, int(match.group(1)))

    index = highest_minor + 1
    candidate = parent / f"{safe_name}_v0.{index}"
    while candidate.exists():
        index += 1
        candidate = parent / f"{safe_name}_v0.{index}"
    return candidate


def requirement_processing_key(item: dict[str, object]) -> tuple[int, int]:
    return (
        SIZE_ORDER.get(str(item.get("_document_size") or "large"), SIZE_ORDER["large"]),
        int(item.get("_original_index") or 0),
    )


def original_requirement_key(item: dict[str, object]) -> int:
    return int(item.get("_original_index") or 0)


def public_requirement_item(item: dict[str, object]) -> dict[str, object]:
    return {
        key: value
        for key, value in item.items()
        if not str(key).startswith("_")
    }


def block_status_key(update: dict[str, object]) -> tuple[str, object]:
    screen_id = str(update.get("screen_id") or "")
    if screen_id:
        return ("screen_id", screen_id)
    unit_test_id = str(update.get("unit_test_id") or "")
    if unit_test_id:
        return ("unit_test_id", unit_test_id)
    return ("original_index", int(update.get("original_index") or 0))


def sorted_block_statuses(blocks: dict[tuple[str, object], dict[str, object]]) -> list[dict[str, object]]:
    return sorted(
        (dict(value) for value in blocks.values()),
        key=lambda value: int(value.get("original_index") or 0),
    )


def mark_queued_blocks_interrupted(
        blocks: dict[tuple[str, object], dict[str, object]],
        reason: str,
) -> bool:
    changed = False
    for block in blocks.values():
        if block.get("status") == "queued":
            block["status"] = "interrupted"
            block["error"] = reason
            changed = True
    return changed


def prepare_requirement_processing_items(
        requirement_items: list[dict[str, object]],
        request_id: str,
        temp_dir: Path,
) -> list[dict[str, object]]:
    prepared_items: list[dict[str, object]] = []
    for index, item in enumerate(requirement_items):
        prepared_item = {
            **item,
            "_original_index": index,
            "_document_size": "large",
            "_screen_block_count": 0,
            "_extracted_text": None,
            "_screen_blocks": None,
            "_ui_design_processing_path": "",
            "_ui_design_converted": False,
            "_preparse_error": "",
        }
        requirement_id = str(item.get("requirement_id") or "")
        ui_design = Path(str(item.get("ui_design_path") or ""))
        try:
            processing_design, converted = prepare_ui_design_for_processing(
                ui_design,
                temp_dir / "ui-design-processing" / safe_requirement_dirname(requirement_id or f"item-{index + 1}"),
            )
            extracted_text = extract_text_from_pdf(processing_design)
            screen_blocks = extract_screen_blocks(extracted_text)
            document_size = classify_document_size(len(screen_blocks))
            prepared_item.update({
                "_document_size": document_size,
                "_screen_block_count": len(screen_blocks),
                "_extracted_text": extracted_text,
                "_screen_blocks": screen_blocks,
                "_ui_design_processing_path": str(processing_design),
                "_ui_design_converted": converted,
            })
            qa_batch_log(
                request_id,
                f"[{requirement_id}] 문서 분류 | {SIZE_LABELS[document_size]}({document_size}) 화면 블록 {len(screen_blocks)}개",
            )
        except Exception as exc:
            prepared_item["_preparse_error"] = str(exc)
            qa_batch_log(request_id, f"[{requirement_id}] 문서 분류 실패 | {exc}")
        prepared_items.append(prepared_item)

    processing_items = sorted(prepared_items, key=requirement_processing_key)
    processing_order = " -> ".join(
        f"{item.get('requirement_id') or '-'}:{item.get('_document_size')}"
        for item in processing_items
    )
    qa_batch_log(request_id, f"처리 순서 최적화 | 소량 문서 우선 | {processing_order}")
    return processing_items


def preview_folder_qa_matching(
        dump_root: Path,
        *,
        ui_design_items: list[tuple[str, bytes]] | None = None,
        ui_design_root: Path | None = None,
        qa_source_items: list[tuple[str, bytes]] | None = None,
        qa_source_root: Path | None = None,
        tc_source_root: Path | None = None,
        unit_result_root: Path | None = None,
        ts_source_root: Path | None = None,
        integration_result_root: Path | None = None,
        request_id: str | None = None,
) -> dict[str, object]:
    request_id = request_id or uuid4().hex[:8]
    dump_root = Path(dump_root).expanduser().resolve()
    if not dump_root.exists() or not dump_root.is_dir():
        raise ValueError(f"결과 폴더를 찾지 못했습니다: {dump_root}")

    temp_dir = TEMP_DIR / f"qa-match-{request_id}"
    temp_dir.mkdir(parents=True, exist_ok=True)
    try:
        ui_design_paths = collect_design_documents(ui_design_root)
        ui_design_paths.extend(save_uploaded_design_documents(temp_dir, ui_design_items or []))
        qa_source_paths = collect_source_documents(qa_source_root)
        qa_source_is_override = bool(qa_source_root and str(qa_source_root).strip())
        uploaded_qa_source_paths = save_uploaded_source_documents(temp_dir, qa_source_items or [])
        ui_design_paths.extend(
            path for path in uploaded_qa_source_paths
            if path.suffix.lower() in DESIGN_DOCUMENT_SUFFIXES
        )
        tc_source_paths = collect_documents(tc_source_root, TC_TEMPLATE_SUFFIXES, "단위시험 폴더")
        unit_result_paths = collect_documents(unit_result_root, TC_TEMPLATE_SUFFIXES, "단위시험결과서 폴더")
        ts_source_paths = collect_documents(ts_source_root, TS_TEMPLATE_SUFFIXES, "통합시험 폴더")
        integration_result_paths = collect_documents(integration_result_root, TS_TEMPLATE_SUFFIXES, "통합시험결과서 폴더")
        selection = select_qa_source_files(
            dump_root,
            ui_design_paths,
            qa_source_paths=qa_source_paths,
            qa_source_is_override=qa_source_is_override,
            tc_source_paths=tc_source_paths,
            unit_result_paths=unit_result_paths,
            ts_source_paths=ts_source_paths,
            integration_result_paths=integration_result_paths,
        )
        requirement_items = build_requirement_work_items(selection)
        return {
            "ok": bool(requirement_items),
            "request_id": request_id,
            "dump_root": str(dump_root),
            "match_preview": True,
            "tc_count": 0,
            "ts_count": 0,
            "requirement_count": len(requirement_items),
            "processed_requirement_count": 0,
            "failed_requirement_count": 0,
            "role_counts": build_role_counts(selection),
            "source_files": serialize_selection(selection),
            "requirement_items": requirement_items,
            "missing_requirements": build_missing_requirement_report(selection),
            "placed_files": [],
            "files": [],
            "error": "" if requirement_items else "요구사항 ID 기준으로 함께 처리할 5종 세트를 찾지 못했습니다.",
        }
    finally:
        remove_runtime_path(temp_dir)


def run_folder_qa_pipeline(
        dump_root: Path,
        *,
        model_name: str,
        ollama_url: str,
        scenario_form_path: Path,
        result_form_path: Path,
        ui_design_items: list[tuple[str, bytes]] | None = None,
        ui_design_root: Path | None = None,
        qa_source_items: list[tuple[str, bytes]] | None = None,
        qa_source_root: Path | None = None,
        tc_source_root: Path | None = None,
        unit_result_root: Path | None = None,
        ts_source_root: Path | None = None,
        integration_result_root: Path | None = None,
        system_name: str = "",
        request_id: str | None = None,
        cancel_check: Callable[[], None] | None = None,
        status_callback: Callable[[dict[str, object]], None] | None = None,
        requirement_filter: str | None = None,
) -> dict[str, object]:
    # check.html이 만든 결과 폴더 안에서 QA 입력물을 찾아 생성 결과를 같은 위치에 배치한다.
    request_id = request_id or uuid4().hex[:8]
    dump_root = Path(dump_root).expanduser().resolve()
    if not dump_root.exists() or not dump_root.is_dir():
        raise ValueError(f"결과 폴더를 찾지 못했습니다: {dump_root}")

    temp_dir = TEMP_DIR / f"qa-folder-{request_id}"
    tc_output_dir = temp_dir / "tc-output"
    ts_output_dir = temp_dir / "ts-output"
    temp_dir.mkdir(parents=True, exist_ok=True)

    qa_batch_log(request_id, f"시작 | 대상 산출물 폴더: {dump_root}")

    try:
        if cancel_check:
            cancel_check()
        ui_design_paths = collect_design_documents(ui_design_root)
        ui_design_paths.extend(save_uploaded_design_documents(temp_dir, ui_design_items or []))
        qa_source_paths = collect_source_documents(qa_source_root)
        qa_source_is_override = bool(qa_source_root and str(qa_source_root).strip())
        uploaded_qa_source_paths = save_uploaded_source_documents(temp_dir, qa_source_items or [])
        ui_design_paths.extend(
            path for path in uploaded_qa_source_paths
            if path.suffix.lower() in DESIGN_DOCUMENT_SUFFIXES
        )
        tc_source_paths = collect_documents(tc_source_root, TC_TEMPLATE_SUFFIXES, "단위시험 폴더")
        unit_result_paths = collect_documents(unit_result_root, TC_TEMPLATE_SUFFIXES, "단위시험결과서 폴더")
        ts_source_paths = collect_documents(ts_source_root, TS_TEMPLATE_SUFFIXES, "통합시험 폴더")
        integration_result_paths = collect_documents(integration_result_root, TS_TEMPLATE_SUFFIXES, "통합시험결과서 폴더")
        selection = select_qa_source_files(
            dump_root,
            ui_design_paths,
            qa_source_paths=qa_source_paths,
            qa_source_is_override=qa_source_is_override,
            tc_source_paths=tc_source_paths,
            unit_result_paths=unit_result_paths,
            ts_source_paths=ts_source_paths,
            integration_result_paths=integration_result_paths,
        )
        all_requirement_items = build_requirement_work_items(selection)
        requirement_items = [
            item for item in all_requirement_items
            if not requirement_filter or str(item.get("requirement_id") or "") == requirement_filter
        ]
        role_counts = build_role_counts(selection)
        if status_callback:
            status_callback({
                "status": "running",
                "dump_root": str(dump_root),
                "requirement_count": len(requirement_items),
                "processed_requirement_count": 0,
                "failed_requirement_count": 0,
                "role_counts": role_counts,
                "source_files": serialize_selection(selection),
                "requirement_items": [
                    {**item, "status": "queued"}
                    for item in requirement_items
                ],
                "missing_requirements": build_missing_requirement_report(selection),
            })
        qa_batch_log(
            request_id,
            "입력 매칭 완료 | "
            f"요구사항 {len(requirement_items)}건 | "
            f"설계서 {role_counts.get('ui_design', 0)}개, "
            f"단위시험케이스 {role_counts.get('tc_template', 0)}개, "
            f"단위시험결과서 {role_counts.get('unit_result_template', 0)}개, "
            f"통합시험시나리오 {role_counts.get('ts_template', 0)}개, "
            f"통합시험결과서 {role_counts.get('integration_result_template', 0)}개",
        )
        if not requirement_items:
            missing_message = (
                f"재생성할 요구사항을 찾지 못했습니다: {requirement_filter}"
                if requirement_filter
                else "요구사항 ID 기준으로 함께 처리할 화면설계서, 단위시험케이스, 단위시험결과서, 통합시험시나리오, 통합시험결과서를 찾지 못했습니다."
            )
            payload = build_matching_failure_payload(
                request_id=request_id,
                dump_root=dump_root,
                selection=selection,
                message=missing_message,
            )
            qa_batch_log(request_id, "중단 | 처리 가능한 5종 세트를 찾지 못했습니다.")
            raise QaFolderMatchingError(str(payload["error"]), payload)

        placed_files: list[dict[str, object]] = []
        processed_items: list[dict[str, object]] = []
        block_status_by_requirement: dict[str, dict[tuple[str, object], dict[str, object]]] = {}
        active_requirement_id: str | None = None
        total_tc_count = 0
        total_ts_count = 0
        processing_items = prepare_requirement_processing_items(requirement_items, request_id, temp_dir)

        def get_requirement_blocks(requirement_id: str) -> list[dict[str, object]]:
            return sorted_block_statuses(block_status_by_requirement.get(requirement_id, {}))

        def update_requirement_block_status(requirement_id: str, update: dict[str, object]) -> None:
            blocks = block_status_by_requirement.setdefault(requirement_id, {})
            key = block_status_key(update)
            current = blocks.get(key, {})
            current.update(update)
            blocks[key] = current
            emit_progress()

        def interrupt_queued_requirement_blocks(requirement_id: str, reason: str) -> None:
            blocks = block_status_by_requirement.get(requirement_id, {})
            if mark_queued_blocks_interrupted(blocks, reason):
                emit_progress()

        def emit_progress(current_item: dict[str, object] | None = None) -> None:
            if not status_callback:
                return
            current_items = [
                {
                    **public_requirement_item(candidate),
                    "status": "running"
                    if candidate is current_item or str(candidate.get("requirement_id") or "") == active_requirement_id
                    else "queued",
                    "document_size": candidate.get("_document_size"),
                    "screen_block_count": candidate.get("_screen_block_count"),
                    "blocks": get_requirement_blocks(str(candidate.get("requirement_id") or "")),
                }
                for candidate in sorted(processing_items, key=original_requirement_key)
            ]
            for processed in processed_items:
                for current in current_items:
                    if current.get("requirement_id") == processed.get("requirement_id"):
                        current.update(public_requirement_item(processed))
                        current["status"] = processed.get("status")
                        current["blocks"] = get_requirement_blocks(str(current.get("requirement_id") or ""))
            status_callback({
                "status": "running",
                "processed_requirement_count": len([done for done in processed_items if done.get("status") != "error"]),
                "failed_requirement_count": len([done for done in processed_items if done.get("status") == "error"]),
                "tc_count": total_tc_count,
                "ts_count": total_ts_count,
                "placed_files": placed_files,
                "requirement_items": current_items,
            })

        if status_callback:
            emit_progress()

        for index, item in enumerate(processing_items, start=1):
            if cancel_check:
                cancel_check()
            requirement_id = str(item["requirement_id"])
            req_temp_dir = temp_dir / safe_requirement_dirname(requirement_id)
            item_tc_output_dir = req_temp_dir / "tc-output"
            item_unit_result_output_dir = req_temp_dir / "unit-result-output"
            item_ts_output_dir = req_temp_dir / "ts-output"
            item_integration_result_output_dir = req_temp_dir / "integration-result-output"
            try:
                active_requirement_id = requirement_id
                ui_design = Path(str(item.get("_ui_design_processing_path") or item["ui_design_path"]))
                tc_template = Path(str(item["tc_template_path"]))
                unit_result_template = Path(str(item["unit_result_template_path"]))
                ts_template = Path(str(item["ts_template_path"]))
                integration_result_template = Path(str(item["integration_result_template_path"]))
                tc_template_target = Path(str(item["tc_template_target_path"]))
                unit_result_template_target = Path(str(item["unit_result_template_target_path"]))
                ts_template_target = Path(str(item["ts_template_target_path"]))
                integration_result_template_target = Path(str(item["integration_result_template_target_path"]))
                preparse_error = str(item.get("_preparse_error") or "")
                if preparse_error:
                    raise RuntimeError(f"설계서 문서 분류/파싱 실패: {preparse_error}")
                emit_progress(item)
                qa_batch_log(
                    request_id,
                    f"[{requirement_id}] 처리 시작 | {index}/{len(processing_items)} | "
                    f"문서={SIZE_LABELS.get(str(item.get('_document_size')), str(item.get('_document_size')))} | "
                    f"화면 {int(item.get('_screen_block_count') or 0)}개",
                )

                tc_payload = run_with_suppressed_output(
                    lambda: generate_test_cases(
                        pdf_path=ui_design,
                        model_name=model_name,
                        ollama_url=ollama_url,
                        output_dir=item_tc_output_dir,
                        template_path=tc_template,
                        extracted_text=str(item.get("_extracted_text") or ""),
                        screen_blocks=item.get("_screen_blocks") if isinstance(item.get("_screen_blocks"), list) else None,
                        cancel_check=cancel_check,
                        progress_callback=lambda message: qa_progress_log(request_id, requirement_id, message),
                        block_status_callback=lambda update: update_requirement_block_status(requirement_id, update),
                    )
                )
                if cancel_check:
                    cancel_check()
                if not tc_payload.get("ok"):
                    raise RuntimeError(str(tc_payload.get("error") or "단위시험 케이스 생성에 실패했습니다."))

                tc_xlsx = find_generated_file(tc_payload, suffix=".xlsx", kind="xlsx")
                if tc_xlsx is None:
                    raise RuntimeError("통합시험 시나리오 생성에 사용할 단위시험 케이스 XLSX가 생성되지 않았습니다.")

                tc_hwpx = find_generated_file(tc_payload, suffix=".hwpx", kind="hwpx")
                if tc_hwpx is None:
                    raise RuntimeError("교체할 단위시험케이스 HWPX가 생성되지 않았습니다.")
                qa_batch_log(
                    request_id,
                    f"[{requirement_id}] 단위시험케이스 생성 완료 | {int(tc_payload.get('count') or 0)}행",
                )
                unit_result_hwpx = run_with_suppressed_output(
                    lambda: generate_unit_result_hwpx(
                        tc_hwpx,
                        unit_result_template,
                        item_unit_result_output_dir,
                    )
                )
                if cancel_check:
                    cancel_check()
                if unit_result_hwpx is None:
                    raise RuntimeError("교체할 단위시험결과서 HWPX가 생성되지 않았습니다.")
                qa_batch_log(
                    request_id,
                    f"[{requirement_id}] 단위시험결과서 생성 완료",
                )

                ts_payload = run_with_suppressed_output(
                    lambda: generate_test_scenarios(
                        template_xlsx_path=ts_template,
                        tc_xlsx_path=tc_xlsx,
                        ui_pdf_path=ui_design,
                        output_dir=item_ts_output_dir,
                        form_path=scenario_form_path,
                        system_name=system_name,
                        cancel_check=cancel_check,
                    )
                )
                if cancel_check:
                    cancel_check()
                if not ts_payload.get("ok"):
                    raise RuntimeError(str(ts_payload.get("error") or "통합시험 시나리오 생성에 실패했습니다."))

                ts_xlsx = find_generated_file(ts_payload, suffix=ts_template.suffix, kind="xlsx")
                if ts_xlsx is None:
                    raise RuntimeError("교체할 통합시험시나리오 XLSX가 생성되지 않았습니다.")
                qa_batch_log(
                    request_id,
                    f"[{requirement_id}] 통합시험시나리오 생성 완료 | {int(ts_payload.get('count') or 0)}행",
                )
                integration_result_payload = run_with_suppressed_output(
                    lambda: generate_integration_test_results(
                        template_xlsx_path=integration_result_template,
                        tc_xlsx_path=tc_xlsx,
                        ui_pdf_path=ui_design,
                        output_dir=item_integration_result_output_dir,
                        form_path=result_form_path,
                        system_name=system_name,
                        cancel_check=cancel_check,
                    )
                )
                if cancel_check:
                    cancel_check()
                if not integration_result_payload.get("ok"):
                    raise RuntimeError(str(integration_result_payload.get("error") or "통합시험 결과서 생성에 실패했습니다."))

                integration_result_xlsx = find_generated_file(
                    integration_result_payload,
                    suffix=integration_result_template.suffix,
                    kind="xlsx",
                )
                if integration_result_xlsx is None:
                    raise RuntimeError("교체할 통합시험결과서 XLSX가 생성되지 않았습니다.")
                qa_batch_log(
                    request_id,
                    f"[{requirement_id}] 통합시험결과서 생성 완료 | {int(integration_result_payload.get('count') or 0)}행",
                )

                if cancel_check:
                    cancel_check()
                placed_for_requirement = [
                    place_generated_file(tc_hwpx, tc_template_target, "tc_hwpx", "단위시험케이스", requirement_id),
                    place_generated_file(unit_result_hwpx, unit_result_template_target, "unit_result_hwpx", "단위시험결과서", requirement_id),
                    place_generated_file(ts_xlsx, ts_template_target, "ts_xlsx", "통합시험시나리오", requirement_id),
                    place_generated_file(integration_result_xlsx, integration_result_template_target, "integration_result_xlsx", "통합시험결과서", requirement_id),
                ]
                for placed in placed_for_requirement:
                    placed["__requirement_original_index"] = int(item.get("_original_index") or 0)
                qa_batch_log(
                    request_id,
                    f"[{requirement_id}] 산출물 배치 완료 | 4개 파일 교체",
                )
                placed_files.extend(placed_for_requirement)
                tc_count = int(tc_payload.get("count") or 0)
                ts_count = int(ts_payload.get("count") or 0)
                total_tc_count += tc_count
                total_ts_count += ts_count
                processed_items.append({
                    **public_requirement_item(item),
                    "status": "updated",
                    "tc_count": tc_count,
                    "ts_count": ts_count,
                    "placed_files": placed_for_requirement,
                    "blocks": get_requirement_blocks(requirement_id),
                    "_original_index": int(item.get("_original_index") or 0),
                })
                emit_progress()
            except CancelledRequest:
                raise
            except Exception as exc:
                qa_batch_log(request_id, f"[{requirement_id}] 실패 | {exc}")
                interrupt_queued_requirement_blocks(
                    requirement_id,
                    "앞 화면ID 실패로 실행하지 않았습니다.",
                )
                processed_items.append({
                    **public_requirement_item(item),
                    "status": "error",
                    "error": str(exc),
                    "blocks": get_requirement_blocks(requirement_id),
                    "_original_index": int(item.get("_original_index") or 0),
                })
                emit_progress()

        processed_items = sorted(processed_items, key=original_requirement_key)
        placed_files = sorted(
            placed_files,
            key=lambda placed: int(placed.get("__requirement_original_index") or 0),
        )
        for processed_item in processed_items:
            requirement_id = str(processed_item.get("requirement_id") or "")
            processed_item["blocks"] = get_requirement_blocks(requirement_id)
            processed_item.pop("_original_index", None)
            for placed in processed_item.get("placed_files") or []:
                if isinstance(placed, dict):
                    placed.pop("__requirement_original_index", None)
        for placed in placed_files:
            placed.pop("__requirement_original_index", None)

        failed_items = [item for item in processed_items if item.get("status") == "error"]
        payload = {
            "ok": bool(placed_files),
            "request_id": request_id,
            "dump_root": str(dump_root),
            "tc_count": total_tc_count,
            "ts_count": total_ts_count,
            "requirement_count": len(requirement_items),
            "processed_requirement_count": len(requirement_items) - len(failed_items),
            "failed_requirement_count": len(failed_items),
            "role_counts": role_counts,
            "source_files": serialize_selection(selection),
            "requirement_items": processed_items,
            "missing_requirements": build_missing_requirement_report(selection),
            "placed_files": placed_files,
            "error": "" if placed_files else "생성되어 배치된 QA 산출물이 없습니다.",
        }
        if status_callback:
            status_callback({
                **payload,
                "status": "done" if payload.get("ok") else "error",
            })
        qa_batch_log(
            request_id,
            "완료 | "
            f"성공 {payload['processed_requirement_count']}/{len(requirement_items)}건 · "
            f"실패 {len(failed_items)}건 · "
            f"배치 파일 {len(placed_files)}개",
        )
        return payload
    except QaFolderMatchingError:
        raise
    except Exception as exc:
        qa_batch_log(request_id, f"오류 종료 | {exc}")
        raise
    finally:
        remove_runtime_path(temp_dir)


def select_qa_source_files(
        dump_root: Path,
        ui_design_paths: list[Path],
        *,
        qa_source_paths: list[Path] | None = None,
        qa_source_is_override: bool = False,
        tc_source_paths: list[Path] | None = None,
        unit_result_paths: list[Path] | None = None,
        ts_source_paths: list[Path] | None = None,
        integration_result_paths: list[Path] | None = None,
) -> dict[str, dict[str, object]]:
    # QA 대상 문서를 요구사항 ID별로 찾는다. 설계서는 업로드 파일을 우선 사용하고, 없으면 결과 폴더에서 보조 탐색한다.
    files = list(iter_candidate_files(dump_root))
    qa_source_paths = list(qa_source_paths or [])
    tc_source_paths = list(tc_source_paths or [])
    unit_result_paths = list(unit_result_paths or [])
    ts_source_paths = list(ts_source_paths or [])
    integration_result_paths = list(integration_result_paths or [])
    qa_source_override_paths = qa_source_paths if qa_source_is_override else []
    base_artifact_paths = qa_source_override_paths if qa_source_is_override else files
    ui_design_candidates = unique_paths(ui_design_paths if ui_design_paths else qa_source_paths)
    tc_candidates = unique_paths(tc_source_paths)
    unit_result_candidates = unique_paths(unit_result_paths)
    ts_candidates = unique_paths(ts_source_paths)
    integration_result_candidates = unique_paths(integration_result_paths)

    fallback_ui_files = [
        path for path in files
        if path.suffix.lower() in DESIGN_DOCUMENT_SUFFIXES
        and requirement_ids_from_path(path)
        and score_keywords(
            searchable_text(dump_root, path),
            DESIGN_KEYWORDS,
        ) > 0
    ]
    return {
        "ui_design": {
            "label": "사용자인터페이스설계서",
            "by_requirement": index_design_files_by_requirement(unique_paths(
                ui_design_candidates if ui_design_candidates else fallback_ui_files
            )),
        },
        "tc_template": {
            "label": "단위시험케이스",
            "by_requirement": merge_artifact_inputs_with_targets(
                index_artifact_files_by_requirement(
                    dump_root,
                    tc_candidates,
                    suffixes=TC_TEMPLATE_SUFFIXES,
                    keywords=TC_KEYWORDS,
                    exclude_keywords=TC_EXCLUDE_KEYWORDS,
                ),
                index_artifact_files_by_requirement(
                    dump_root,
                    base_artifact_paths,
                    suffixes=TC_TEMPLATE_SUFFIXES,
                    keywords=TC_KEYWORDS,
                    exclude_keywords=TC_EXCLUDE_KEYWORDS,
                ),
            ),
        },
        "unit_result_template": {
            "label": "단위시험결과서",
            "by_requirement": merge_artifact_inputs_with_targets(
                index_artifact_files_by_requirement(
                    dump_root,
                    unit_result_candidates,
                    suffixes=TC_TEMPLATE_SUFFIXES,
                    keywords=UNIT_RESULT_KEYWORDS,
                    exclude_keywords=UNIT_RESULT_EXCLUDE_KEYWORDS,
                ),
                index_artifact_files_by_requirement(
                    dump_root,
                    base_artifact_paths,
                    suffixes=TC_TEMPLATE_SUFFIXES,
                    keywords=UNIT_RESULT_KEYWORDS,
                    exclude_keywords=UNIT_RESULT_EXCLUDE_KEYWORDS,
                ),
            ),
        },
        "ts_template": {
            "label": "통합시험시나리오",
            "by_requirement": merge_artifact_inputs_with_targets(
                index_artifact_files_by_requirement(
                    dump_root,
                    ts_candidates,
                    suffixes=TS_TEMPLATE_SUFFIXES,
                    keywords=(
                        "통합시험시나리오",
                        "통합시험 시나리오",
                        "통합테스트",
                        "통합 테스트",
                        "integrationtestscenario",
                        "integration test scenario",
                    ),
                    exclude_keywords=("단위시험", "단위테스트", "케이스", "결과서", "인수인계"),
                ),
                index_artifact_files_by_requirement(
                    dump_root,
                    base_artifact_paths,
                    suffixes=TS_TEMPLATE_SUFFIXES,
                    keywords=(
                        "통합시험시나리오",
                        "통합시험 시나리오",
                        "통합테스트",
                        "통합 테스트",
                        "integrationtestscenario",
                        "integration test scenario",
                    ),
                    exclude_keywords=("단위시험", "단위테스트", "케이스", "결과서", "인수인계"),
                ),
            ),
        },
        "integration_result_template": {
            "label": "통합시험결과서",
            "by_requirement": merge_artifact_inputs_with_targets(
                index_artifact_files_by_requirement(
                    dump_root,
                    integration_result_candidates,
                    suffixes=TS_TEMPLATE_SUFFIXES,
                    keywords=INTEGRATION_RESULT_KEYWORDS,
                    exclude_keywords=INTEGRATION_RESULT_EXCLUDE_KEYWORDS,
                ),
                index_artifact_files_by_requirement(
                    dump_root,
                    base_artifact_paths,
                    suffixes=TS_TEMPLATE_SUFFIXES,
                    keywords=INTEGRATION_RESULT_KEYWORDS,
                    exclude_keywords=INTEGRATION_RESULT_EXCLUDE_KEYWORDS,
                ),
            ),
        },
    }


def merge_artifact_inputs_with_targets(
        input_by_requirement: dict[str, dict[str, object]],
        target_by_requirement: dict[str, dict[str, object]],
) -> dict[str, dict[str, object]]:
    # 문서별 위치 지정은 입력 후보로만 쓰고, 교체 배치는 대상 폴더의 기존 파일을 기준으로 한다.
    result: dict[str, dict[str, object]] = {}
    for requirement_id, target in target_by_requirement.items():
        selected_input = input_by_requirement.get(requirement_id) or target
        if not selected_input:
            continue
        merged = dict(selected_input)
        merged["placement_path"] = target.get("path", "")
        merged["placement_score"] = target.get("score", 0)
        merged["placement_candidates"] = target.get("candidates", [])
        result[requirement_id] = merged
    return result


def iter_candidate_files(root: Path):
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if any(part.casefold() in IGNORED_FOLDER_NAMES for part in path.relative_to(root).parts[:-1]):
            continue
        yield path


def unique_paths(paths: list[Path]) -> list[Path]:
    result: list[Path] = []
    seen: set[str] = set()
    for path in paths:
        resolved = str(Path(path).resolve()).casefold()
        if resolved in seen:
            continue
        seen.add(resolved)
        result.append(Path(path))
    return result


def filter_role_files(
        root: Path,
        files: list[Path],
        *,
        suffixes: set[str],
        keywords: tuple[str, ...],
        exclude_keywords: tuple[str, ...] = (),
) -> list[Path]:
    result: list[Path] = []
    for path in files:
        if path.suffix.lower() not in suffixes:
            continue
        text = searchable_text(root, path)
        if score_keywords(text, keywords) - score_keywords(text, exclude_keywords) <= 0:
            continue
        result.append(path)
    return result


def collect_documents(root: Path | None, suffixes: set[str], label: str) -> list[Path]:
    # 사용자가 지정한 로컬 폴더에서 허용 확장자 파일을 모두 모은다.
    if root is None:
        return []

    root = Path(root).expanduser()
    if not str(root).strip():
        return []

    root = root.resolve()
    if not root.exists() or not root.is_dir():
        raise ValueError(f"{label}를 찾지 못했습니다: {root}")

    return [
        path
        for path in root.rglob("*")
        if path.is_file()
        and path.suffix.lower() in suffixes
        and not any(part.casefold() in IGNORED_FOLDER_NAMES for part in path.relative_to(root).parts[:-1])
    ]


def collect_design_documents(root: Path | None) -> list[Path]:
    # 사용자가 지정한 화면설계서 폴더에서 PDF/HWPX 설계서를 모은다.
    return collect_documents(root, DESIGN_DOCUMENT_SUFFIXES, "화면설계서 폴더")


def collect_source_documents(root: Path | None) -> list[Path]:
    # QA 원천 폴더에서 화면설계서/단위시험케이스/통합시험 파일 후보를 모두 모은다.
    return collect_documents(root, QA_SOURCE_SUFFIXES, "QA 원천 폴더")


def save_uploaded_design_documents(temp_dir: Path, items: list[tuple[str, bytes]]) -> list[Path]:
    # qa.html에서 직접 올린 화면/사용자인터페이스 설계서를 임시 폴더에 저장한다.
    design_dir = temp_dir / "ui-designs"
    saved: list[Path] = []
    for index, (filename, payload) in enumerate(items, start=1):
        if not payload:
            continue
        suffix = Path(filename).suffix.lower()
        if suffix not in DESIGN_DOCUMENT_SUFFIXES:
            continue

        relative_path = safe_relative_upload_path(filename, f"ui-design-{index}{suffix}")
        target_path = design_dir / relative_path
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(payload)
        saved.append(target_path)

    return saved


def save_uploaded_source_documents(temp_dir: Path, items: list[tuple[str, bytes]]) -> list[Path]:
    # qa.html에서 선택한 QA 원천 폴더를 임시 폴더에 상대 경로 그대로 저장한다.
    source_dir = temp_dir / "qa-sources"
    saved: list[Path] = []
    for index, (filename, payload) in enumerate(items, start=1):
        if not payload:
            continue
        suffix = Path(filename).suffix.lower()
        if suffix not in QA_SOURCE_SUFFIXES:
            continue

        relative_path = safe_relative_upload_path(filename, f"qa-source-{index}{suffix}")
        target_path = source_dir / relative_path
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(payload)
        saved.append(target_path)

    return saved


def index_design_files_by_requirement(files: list[Path]) -> dict[str, dict[str, object]]:
    # 업로드된 설계서 파일을 SFR 요구사항 ID별로 묶는다.
    result: dict[str, dict[str, object]] = {}
    for path in files:
        if path.suffix.lower() not in DESIGN_DOCUMENT_SUFFIXES:
            continue

        searchable = searchable_text(path.parent, path)
        design_score = score_keywords(searchable, DESIGN_KEYWORDS)
        if path.suffix.lower() in {".hwp", ".hwpx"} and design_score <= 0:
            non_design_score = score_keywords(
                searchable,
                TC_KEYWORDS + UNIT_RESULT_KEYWORDS + INTEGRATION_RESULT_KEYWORDS,
            )
            if non_design_score > 0:
                continue

        requirement_ids = requirement_ids_from_path(path)
        if not requirement_ids:
            requirement_ids = requirement_ids_from_document_text(path)

        for requirement_id in requirement_ids:
            score = 1000 + design_score
            upsert_requirement_file(result, requirement_id, path, score)

    return result


def index_artifact_files_by_requirement(
        root: Path,
        files: list[Path],
        *,
        suffixes: set[str],
        keywords: tuple[str, ...],
        exclude_keywords: tuple[str, ...] = (),
) -> dict[str, dict[str, object]]:
    # check 결과 폴더의 QA 산출물 양식을 SFR 요구사항 ID별로 묶는다.
    result: dict[str, dict[str, object]] = {}
    for path in files:
        if path.suffix.lower() not in suffixes:
            continue

        text = searchable_text(root, path)
        requirement_ids = requirement_ids_from_path(path)
        if not requirement_ids:
            requirement_ids = requirement_ids_from_artifact_content(path)
        if not requirement_ids:
            continue

        score = score_keywords(text, keywords) - score_keywords(text, exclude_keywords)
        if score <= 0:
            continue

        for requirement_id in requirement_ids:
            upsert_requirement_file(result, requirement_id, path, score)

    return result


def upsert_requirement_file(result: dict[str, dict[str, object]], requirement_id: str, path: Path, score: int) -> None:
    existing = result.get(requirement_id)
    candidate = {"path": str(path), "score": score}
    if existing is None:
        result[requirement_id] = {
            "path": str(path),
            "score": score,
            "candidates": [candidate],
        }
        return

    candidates = list(existing.get("candidates") or [])
    candidates.append(candidate)
    candidates = sorted(
        candidates,
        key=lambda item: (int(item.get("score") or 0), -len(str(item.get("path") or "")), str(item.get("path") or "").casefold()),
        reverse=True,
    )
    best = candidates[0]
    existing["path"] = best["path"]
    existing["score"] = best["score"]
    existing["candidates"] = candidates[:5]


def requirement_ids_from_path(path: Path) -> list[str]:
    return requirement_ids_from_text(path.as_posix())


def requirement_ids_from_text(value: str) -> list[str]:
    return sorted({
        match.group(0).upper()
        for match in REQ_ID_PATTERN.finditer(value)
    })


def requirement_ids_from_document_text(path: Path) -> list[str]:
    try:
        text = extract_document_text(path)
    except Exception as exc:
        log_event("qa.folder.design_text_error", path=str(path), error=str(exc))
        return []
    return requirement_ids_from_text(text)


def requirement_ids_from_artifact_content(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix in {".hwp", ".hwpx", ".pdf"}:
        return requirement_ids_from_document_text(path)
    if suffix == ".xlsx":
        return requirement_ids_from_spreadsheet(path)
    return []


def requirement_ids_from_spreadsheet(path: Path) -> list[str]:
    found: set[str] = set()
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        log_event("qa.folder.spreadsheet_text_error", path=str(path), error=str(exc))
        return []

    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    found.update(requirement_ids_from_text(str(cell.value)))
    except Exception as exc:
        log_event("qa.folder.spreadsheet_scan_error", path=str(path), error=str(exc))
    finally:
        workbook.close()

    return sorted(found)


def build_requirement_work_items(selection: dict[str, dict[str, object]]) -> list[dict[str, object]]:
    ui_by_req = by_requirement(selection, "ui_design")
    tc_by_req = by_requirement(selection, "tc_template")
    unit_result_by_req = by_requirement(selection, "unit_result_template")
    ts_by_req = by_requirement(selection, "ts_template")
    integration_result_by_req = by_requirement(selection, "integration_result_template")
    requirement_ids = sorted(
        set(ui_by_req)
        & set(tc_by_req)
        & set(unit_result_by_req)
        & set(ts_by_req)
        & set(integration_result_by_req)
    )
    return [
        {
            "requirement_id": requirement_id,
            "ui_design_path": ui_by_req[requirement_id]["path"],
            "tc_template_path": tc_by_req[requirement_id]["path"],
            "unit_result_template_path": unit_result_by_req[requirement_id]["path"],
            "ts_template_path": ts_by_req[requirement_id]["path"],
            "integration_result_template_path": integration_result_by_req[requirement_id]["path"],
            "tc_template_target_path": tc_by_req[requirement_id].get("placement_path") or tc_by_req[requirement_id]["path"],
            "unit_result_template_target_path": unit_result_by_req[requirement_id].get("placement_path") or unit_result_by_req[requirement_id]["path"],
            "ts_template_target_path": ts_by_req[requirement_id].get("placement_path") or ts_by_req[requirement_id]["path"],
            "integration_result_template_target_path": integration_result_by_req[requirement_id].get("placement_path") or integration_result_by_req[requirement_id]["path"],
        }
        for requirement_id in requirement_ids
    ]


def by_requirement(selection: dict[str, dict[str, object]], key: str) -> dict[str, dict[str, object]]:
    value = selection.get(key, {}).get("by_requirement", {})
    return value if isinstance(value, dict) else {}


def build_missing_requirement_report(selection: dict[str, dict[str, object]]) -> list[dict[str, object]]:
    role_labels = {
        "ui_design": "사용자인터페이스설계서",
        "tc_template": "단위시험케이스",
        "unit_result_template": "단위시험결과서",
        "ts_template": "통합시험시나리오",
        "integration_result_template": "통합시험결과서",
    }
    indexes = {
        key: by_requirement(selection, key)
        for key in role_labels
    }
    requirement_ids = sorted(set().union(*(set(index) for index in indexes.values())))
    return [
        {
            "requirement_id": requirement_id,
            "missing": [
                label
                for key, label in role_labels.items()
                if requirement_id not in indexes[key]
            ],
        }
        for requirement_id in requirement_ids
        if any(requirement_id not in index for index in indexes.values())
    ]


def build_role_counts(selection: dict[str, dict[str, object]]) -> dict[str, int]:
    return {
        key: len(by_requirement(selection, key))
        for key in ("ui_design", "tc_template", "unit_result_template", "ts_template", "integration_result_template")
    }


def build_matching_failure_payload(
        *,
        request_id: str,
        dump_root: Path,
        selection: dict[str, dict[str, object]],
        message: str,
) -> dict[str, object]:
    return {
        "ok": False,
        "request_id": request_id,
        "dump_root": str(dump_root),
        "error": message,
        "tc_count": 0,
        "ts_count": 0,
        "requirement_count": 0,
        "processed_requirement_count": 0,
        "failed_requirement_count": 0,
        "role_counts": build_role_counts(selection),
        "source_files": serialize_selection(selection),
        "requirement_items": [],
        "missing_requirements": build_missing_requirement_report(selection),
        "placed_files": [],
        "files": [],
    }


def searchable_text(root: Path, path: Path) -> str:
    try:
        relative = path.relative_to(root)
    except ValueError:
        relative = path
    return normalize_search_text(f"{relative.as_posix()} {path.stem}")


def normalize_search_text(value: str) -> str:
    return re.sub(r"[\s_\-()\[\]{}./\\]+", "", value).casefold()


def score_keywords(text: str, keywords: tuple[str, ...]) -> int:
    score = 0
    for keyword in keywords:
        normalized = normalize_search_text(keyword)
        if normalized and normalized in text:
            score += 100 + len(normalized)
    return score


def generate_unit_result_hwpx(tc_hwpx: Path, result_template: Path, output_dir: Path) -> Path | None:
    tc_hwpx = Path(tc_hwpx)
    result_template = Path(result_template)
    if not tc_hwpx.exists():
        return None

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = unique_file_path(output_dir / "generated_unit_result.hwpx")
    copy_hwpx_with_text_replacements(
        tc_hwpx,
        output_path,
        build_unit_result_replacements(tc_hwpx, result_template),
    )
    return output_path


def build_unit_result_replacements(tc_hwpx: Path, result_template: Path) -> dict[str, str]:
    replacements = {
        "단위시험케이스": "단위시험결과서",
    }

    source_author = extract_cover_author_from_document(tc_hwpx)
    result_author = extract_cover_author_from_document(result_template)
    if source_author and result_author and source_author != result_author:
        replacements[source_author] = result_author

    source_doc_no = extract_document_field(tc_hwpx, "문서번호")
    result_doc_no = extract_document_field(result_template, "문서번호")
    if source_doc_no and result_doc_no and source_doc_no != result_doc_no:
        replacements[source_doc_no] = result_doc_no

    source_version = extract_document_field(tc_hwpx, "문서버전")
    result_version = extract_document_field(result_template, "문서버전")
    if source_version and result_version and source_version != result_version:
        replacements[source_version] = result_version

    return replacements


def extract_document_field(document_path: Path, label: str) -> str:
    try:
        text = extract_document_text(Path(document_path))
    except Exception:
        return ""

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    normalized_label = re.sub(r"\s+", "", label)
    for index, line in enumerate(lines[:120]):
        if re.sub(r"\s+", "", line) != normalized_label:
            continue
        return lines[index + 1] if index + 1 < len(lines) else ""
    return ""


def copy_hwpx_with_text_replacements(source_path: Path, output_path: Path, replacements: dict[str, str]) -> None:
    with zipfile.ZipFile(source_path, "r") as source_zip, zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as output_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename.endswith(".xml") or item.filename.endswith(".txt"):
                text = data.decode("utf-8", errors="ignore")
                for old, new in replacements.items():
                    if old:
                        text = text.replace(old, new)
                data = text.encode("utf-8")
            output_zip.writestr(item, data)


def find_generated_file(payload: dict[str, object], *, suffix: str, kind: str) -> Path | None:
    files = payload.get("files")
    if not isinstance(files, list):
        return None

    normalized_suffix = suffix.lower()
    for item in files:
        if not isinstance(item, dict):
            continue
        path = Path(str(item.get("path") or ""))
        item_kind = str(item.get("kind") or "").lower()
        if path.exists() and path.suffix.lower() == normalized_suffix and item_kind == kind:
            return path

    for item in files:
        if not isinstance(item, dict):
            continue
        path = Path(str(item.get("path") or ""))
        if path.exists() and path.suffix.lower() == normalized_suffix:
            return path

    return None


def place_generated_file(
        source_path: Path,
        target_path: Path,
        kind: str,
        label: str,
        requirement_id: str = "",
) -> dict[str, object]:
    # 기존 산출물은 같은 폴더의 bak 하위로 이동하고 생성 파일을 기존 위치/이름에 맞춰 넣는다.
    source_path = Path(source_path)
    target_path = Path(target_path)
    if not source_path.exists() or not source_path.is_file():
        raise FileNotFoundError(f"생성 파일을 찾지 못했습니다: {source_path}")

    backup_path = ""
    if target_path.exists():
        backup_dir = target_path.parent / "bak"
        backup_dir.mkdir(exist_ok=True)
        backup_target = find_same_backup_file(backup_dir, target_path)
        if backup_target is None:
            backup_target = unique_file_path(backup_dir / target_path.name)
            shutil.move(str(target_path), str(backup_target))
        else:
            target_path.unlink()
        backup_path = str(backup_target)

    target_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(source_path), str(target_path))
    return {
        "kind": kind,
        "label": label,
        "requirement_id": requirement_id,
        "path": str(target_path),
        "backup_path": backup_path,
    }


def find_same_backup_file(backup_dir: Path, source_path: Path) -> Path | None:
    if not backup_dir.exists() or not source_path.exists():
        return None

    for candidate in backup_dir.glob(f"{source_path.stem}*{source_path.suffix}"):
        if not candidate.is_file():
            continue
        try:
            if filecmp.cmp(source_path, candidate, shallow=False):
                return candidate
        except OSError:
            continue
    return None


def unique_file_path(path: Path) -> Path:
    if not path.exists():
        return path

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = path.with_name(f"{path.stem}_{timestamp}{path.suffix}")
    index = 1
    while candidate.exists():
        candidate = path.with_name(f"{path.stem}_{timestamp}_{index}{path.suffix}")
        index += 1
    return candidate


def safe_requirement_dirname(requirement_id: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", requirement_id) or "requirement"


def serialize_selection(selection: dict[str, dict[str, object]]) -> list[dict[str, object]]:
    labels = {
        "ui_design": "사용자인터페이스설계서",
        "tc_template": "단위시험케이스",
        "unit_result_template": "단위시험결과서",
        "ts_template": "통합시험시나리오",
        "integration_result_template": "통합시험결과서",
    }
    items: list[dict[str, object]] = []
    for key, value in selection.items():
        by_req = value.get("by_requirement", {})
        if not isinstance(by_req, dict):
            continue
        for requirement_id, selected in sorted(by_req.items()):
            if not isinstance(selected, dict):
                continue
            items.append({
                "role": key,
                "label": labels.get(key, key),
                "requirement_id": requirement_id,
                "path": str(selected.get("path") or ""),
                "placement_path": str(selected.get("placement_path") or selected.get("path") or ""),
                "score": selected.get("score", 0),
                "candidates": selected.get("candidates", []),
            })
    return items
