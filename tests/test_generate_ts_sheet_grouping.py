import tempfile
import unittest
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Side

from qa_generation.generate_ts import (
    build_test_scenarios_from_unit_tests,
    save_integration_test_results_to_excel,
    save_test_scenarios_to_excel,
)


def create_base_workbook(path: Path):
    wb = openpyxl.Workbook()
    wb.active.title = "표지"
    wb.create_sheet("개정이력")
    wb.create_sheet("기존시트")
    wb.create_sheet("작성방법")
    wb.save(path)
    wb.close()


def create_form_workbook(path: Path, max_column: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "양식"
    ws["H1"] = "통합시험 시나리오"
    ws["B2"] = "시스템"
    ws["F2"] = "작성자"
    ws["B3"] = ""
    ws["B4"] = "시나리오ID"
    ws["F4"] = "시나리오명"
    ws["I4"] = "요구사항 ID"
    for col in range(1, max_column + 1):
        ws.cell(row=7, column=col).value = f"col{col}"
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)
    ws.cell(row=100, column=1).border = Border(bottom=Side(style="thin"))
    wb.save(path)
    wb.close()


def sample_scenarios():
    return [
        {
            "시스템": "시스템",
            "작성자": "작성자",
            "시나리오ID": "AT-IIL-002-01-01-01",
            "시나리오명": "첫 번째 화면",
            "요구사항_ID": "SFR-IIL-002",
            "케이스명": "케이스 A",
            "업무처리내용": "",
            "시험항목": "항목 A",
            "사전조건": "조건",
            "예상결과": "결과 A",
            "화면ID": "UI-IIL-002-01-01-01",
        },
        {
            "시스템": "시스템",
            "작성자": "작성자",
            "시나리오ID": "AT-IIL-002-01-01-01",
            "시나리오명": "첫 번째 화면",
            "요구사항_ID": "SFR-IIL-002",
            "케이스명": "케이스 A",
            "업무처리내용": "",
            "시험항목": "항목 A-2",
            "사전조건": "조건",
            "예상결과": "결과 A-2",
            "화면ID": "UI-IIL-002-01-01-01",
        },
        {
            "시스템": "시스템",
            "작성자": "작성자",
            "시나리오ID": "AT-IIL-002-01-01-02",
            "시나리오명": "두 번째 화면",
            "요구사항_ID": "SFR-IIL-002",
            "케이스명": "케이스 B",
            "업무처리내용": "",
            "시험항목": "항목 B",
            "사전조건": "조건",
            "예상결과": "결과 B",
            "화면ID": "UI-IIL-002-01-01-02",
        },
    ]


class GenerateTsSheetGroupingTest(unittest.TestCase):
    def test_build_scenarios_uses_system_name(self):
        scenarios = build_test_scenarios_from_unit_tests(
            [{"화면_ID": "UI-IIL-002-01-01-01", "테스트_케이스": "조회한다."}],
            author="작성자",
            system_name="수입식품통합시스템",
        )

        self.assertEqual("수입식품통합시스템", scenarios[0]["시스템"])

    def test_scenario_workbook_uses_requirement_sheet_and_repeated_blocks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            base = temp / "base.xlsx"
            form = temp / "form.xlsx"
            create_base_workbook(base)
            create_form_workbook(form, max_column=9)

            output = save_test_scenarios_to_excel(sample_scenarios(), base, temp, form)
            wb = openpyxl.load_workbook(output, data_only=True)
            try:
                self.assertIn("SFR-IIL-002", wb.sheetnames)
                self.assertNotIn("AT-IIL-002-01-01-01", wb.sheetnames)
                self.assertNotIn("AT-IIL-002-01-01-02", wb.sheetnames)

                ws = wb["SFR-IIL-002"]
                values = [cell.value for row in ws.iter_rows() for cell in row]
                self.assertIn("AT-IIL-002-01-01-01", values)
                self.assertIn("AT-IIL-002-01-01-02", values)
                self.assertIn("SFR-IIL-002", values)
                self.assertEqual(1, values.count("통합시험 시나리오"))
                self.assertEqual(1, values.count("시스템"))
                self.assertEqual("AT-IIL-002-01-01-01", ws["B4"].value)
                self.assertEqual("AT-IIL-002-01-01-02", ws["B10"].value)
                self.assertLess(ws.max_row, 100)
            finally:
                wb.close()

    def test_result_workbook_uses_requirement_sheet_and_repeated_blocks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            base = temp / "base.xlsx"
            form = temp / "result_form.xlsx"
            create_base_workbook(base)
            create_form_workbook(form, max_column=13)

            output = save_integration_test_results_to_excel(sample_scenarios(), base, temp, form)
            wb = openpyxl.load_workbook(output, data_only=True)
            try:
                self.assertIn("SFR-IIL-002", wb.sheetnames)
                self.assertNotIn("AT-IIL-002-01-01-01", wb.sheetnames)
                self.assertNotIn("AT-IIL-002-01-01-02", wb.sheetnames)

                ws = wb["SFR-IIL-002"]
                values = [cell.value for row in ws.iter_rows() for cell in row]
                self.assertIn("AT-IIL-002-01-01-01", values)
                self.assertIn("AT-IIL-002-01-01-02", values)
                self.assertIn("SFR-IIL-002", values)
                self.assertEqual(1, values.count("통합시험 시나리오"))
                self.assertEqual(1, values.count("시스템"))
                self.assertEqual("AT-IIL-002-01-01-01", ws["B4"].value)
                self.assertEqual("AT-IIL-002-01-01-02", ws["B10"].value)
                self.assertLess(ws.max_row, 100)
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
