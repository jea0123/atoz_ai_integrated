from contextlib import contextmanager
from pathlib import Path
import re
import shutil
import unittest
from uuid import uuid4
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from document_update.excel_ooxml import replace_cell_xml, replace_or_insert_cell_xml
from document_update.metadata_update import (
    WbsMetadata,
    extract_requirement_id_from_values,
    records_matching_path,
    update_revision_history_xml,
    update_unlabeled_metadata_row,
    write_updated_excel_metadata,
)


@contextmanager
def workspace_temp_dir():
    path = Path.cwd() / ".test-artifacts" / uuid4().hex
    path.mkdir(parents=True, exist_ok=True)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def rewrite_zip_text_member(path: Path, member_name: str, transform) -> None:
    temp_path = path.with_name(f"{path.stem}-zipedit{path.suffix}")
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == member_name:
                    data = transform(data.decode("utf-8")).encode("utf-8")
                zout.writestr(item, data)
    shutil.copyfile(temp_path, path)
    try:
        temp_path.unlink()
    except OSError:
        pass


def rewrite_zip_members(path: Path, transforms: dict[str, object], additions: dict[str, bytes] | None = None) -> None:
    temp_path = path.with_name(f"{path.stem}-zipedit{path.suffix}")
    additions = additions or {}
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            written = set()
            for item in zin.infolist():
                data = zin.read(item.filename)
                transform = transforms.get(item.filename)
                if transform is not None:
                    data = transform(data.decode("utf-8")).encode("utf-8")
                zout.writestr(item, data)
                written.add(item.filename)
            for name, data in additions.items():
                if name not in written:
                    zout.writestr(name, data)
    shutil.copyfile(temp_path, path)
    try:
        temp_path.unlink()
    except OSError:
        pass


class MetadataRequirementIdTest(unittest.TestCase):
    def test_replace_cell_xml_self_closing_cell_is_valid(self) -> None:
        updated = replace_cell_xml('<c r="A1" s="1"/>', "MFDS-PMC-08")

        ET.fromstring(updated)
        self.assertIn('r="A1"', updated)
        self.assertIn('s="1"', updated)
        self.assertIn('t="inlineStr"', updated)
        self.assertNotIn('/ t="inlineStr"', updated)

    def test_replace_cell_xml_does_not_overwrite_formula_cell(self) -> None:
        formula_cell = '<c r="A1" t="str"><f t="shared" ref="A1:A2" si="0">Sheet1!A1</f><v>old</v></c>'

        updated = replace_cell_xml(formula_cell, "new")

        self.assertEqual(formula_cell, updated)

    def test_replace_cell_xml_strips_cell_metadata_attrs(self) -> None:
        updated = replace_cell_xml('<c r="A1" s="1" t="s" cm="2" vm="3" ph="1"><v>0</v></c>', "text")

        ET.fromstring(updated)
        self.assertIn('s="1"', updated)
        self.assertNotIn(' cm=', updated)
        self.assertNotIn(' vm=', updated)
        self.assertNotIn(' ph=', updated)

    def test_insert_after_self_closing_cell_keeps_next_row_separate(self) -> None:
        sheet_xml = (
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData>'
            '<row r="5"><c r="H5" s="1"/><c r="J5" s="1"/></row>'
            '<row r="6"><c r="C6" s="2"><v>1</v></c></row>'
            '</sheetData>'
            '</worksheet>'
        )

        updated = replace_or_insert_cell_xml(sheet_xml, "H5", 5, 8, "2026-00-00")
        updated = replace_or_insert_cell_xml(updated, "I5", 5, 9, "author")
        root = ET.fromstring(updated)
        ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

        row5 = root.find(".//x:row[@r='5']", ns)
        row6 = root.find(".//x:row[@r='6']", ns)
        self.assertIsNotNone(row5)
        self.assertIsNotNone(row6)
        self.assertEqual(["H5", "I5", "J5"], [cell.attrib.get("r") for cell in row5.findall("x:c", ns)])
        self.assertEqual(["C6"], [cell.attrib.get("r") for cell in row6.findall("x:c", ns)])

    def test_extract_requirement_id_allows_variable_width_number(self) -> None:
        self.assertEqual(
            "SFR-OOO-13",
            extract_requirement_id_from_values("14-RequirementDefinition_SFR-OOO-13_v0.1.hwpx"),
        )

    def test_records_matching_path_filters_by_variable_width_requirement_id(self) -> None:
        records = [
            WbsMetadata(
                output_name="RequirementDefinition",
                author="A",
                revision_date="2026-01-01",
                wbs="",
                task="",
                row=1,
                requirement_id="SFR-OOO-13",
            ),
            WbsMetadata(
                output_name="RequirementDefinition",
                author="B",
                revision_date="2026-01-02",
                wbs="",
                task="",
                row=2,
                requirement_id="SFR-OOO-14",
            ),
        ]

        matches = records_matching_path(
            Path("14-RequirementDefinition_SFR-OOO-13_v0.1.hwpx"),
            records,
        )

        self.assertEqual(1, len(matches))
        self.assertEqual("SFR-OOO-13", matches[0].requirement_id)

    def test_unlabeled_header_short_document_code_updates_author_not_approval(self) -> None:
        row_xml = """
        <hp:tr>
          <hp:tc><hp:p><hp:run><hp:t>LS</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>v0.1</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>2025.03.23</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>임채현</hp:t></hp:run></hp:p></hp:tc>
        </hp:tr>
        """

        updated_row, count = update_unlabeled_metadata_row(
            row_xml,
            revision_date="2025-03-23",
            author="송아름",
        )

        self.assertEqual(2, count)
        self.assertIn("송아름", updated_row)
        self.assertNotIn("임채현", updated_row)
        self.assertIn("2025-03-23", updated_row)

    def test_unlabeled_header_space_date_updates_author(self) -> None:
        row_xml = """
        <hp:tr>
          <hp:tc><hp:p><hp:run><hp:t>LS</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>v0.1</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>2025 03 23</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>임채현</hp:t></hp:run></hp:p></hp:tc>
        </hp:tr>
        """

        updated_row, count = update_unlabeled_metadata_row(
            row_xml,
            revision_date="2025-03-23",
            author="송아름",
        )

        self.assertEqual(2, count)
        self.assertIn("송아름", updated_row)
        self.assertIn("2025-03-23", updated_row)

    def test_unlabeled_header_with_title_cell_updates_author_and_keeps_document_code(self) -> None:
        row_xml = """
        <hp:tr>
          <hp:tc><hp:p><hp:run><hp:t>테일러링결과서</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>LS</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>v0.1</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>2025.03.23</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>임채현</hp:t></hp:run></hp:p></hp:tc>
        </hp:tr>
        """

        updated_row, count = update_unlabeled_metadata_row(
            row_xml,
            revision_date="2025-03-23",
            author="송아름",
        )

        self.assertEqual(2, count)
        self.assertIn("테일러링결과서", updated_row)
        self.assertIn("LS", updated_row)
        self.assertIn("2025-03-23", updated_row)
        self.assertIn("송아름", updated_row)
        self.assertNotIn("임채현", updated_row)

    def test_unlabeled_header_numeric_document_code_updates_date_and_author(self) -> None:
        row_xml = """
        <hp:tr>
          <hp:tc><hp:p><hp:run><hp:t>10</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>v0.1</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>2025.03.28</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>임채현</hp:t></hp:run></hp:p></hp:tc>
        </hp:tr>
        """

        updated_row, count = update_unlabeled_metadata_row(
            row_xml,
            revision_date="2026-00-00",
            author="송아름",
        )

        self.assertEqual(2, count)
        self.assertIn("10", updated_row)
        self.assertIn("2026-00-00", updated_row)
        self.assertIn("송아름", updated_row)
        self.assertNotIn("임채현", updated_row)

    def test_unlabeled_header_version_like_document_code_is_kept_by_position(self) -> None:
        row_xml = """
        <hp:tr>
          <hp:tc><hp:p><hp:run><hp:t>1.0</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>v0.1</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>2025 03 28</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>임채현</hp:t></hp:run></hp:p></hp:tc>
        </hp:tr>
        """

        updated_row, count = update_unlabeled_metadata_row(
            row_xml,
            revision_date="2026-00-00",
            author="송아름",
        )

        self.assertEqual(2, count)
        self.assertIn("1.0", updated_row)
        self.assertIn("2026-00-00", updated_row)
        self.assertIn("송아름", updated_row)

    def test_unlabeled_header_blank_document_code_still_updates_metadata_slots(self) -> None:
        author = "\uc1a1\uc544\ub984"
        old_author = "\uae40\ub300\uba85"
        row_xml = f"""
        <hp:tr>
          <hp:tc><hp:p><hp:run><hp:t></hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>V0.0</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>2024.04.15</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>{old_author}</hp:t></hp:run></hp:p></hp:tc>
        </hp:tr>
        """

        updated_row, count = update_unlabeled_metadata_row(
            row_xml,
            revision_date="2026",
            author=author,
        )

        self.assertEqual(3, count)
        self.assertIn("v0.1", updated_row)
        self.assertIn("2026-00-00", updated_row)
        self.assertIn(author, updated_row)
        self.assertNotIn("V0.0", updated_row)
        self.assertNotIn("2024.04.15", updated_row)
        self.assertNotIn(old_author, updated_row)

    def test_unlabeled_header_blank_document_code_is_filled_when_document_number_is_known(self) -> None:
        author = "\uc1a1\uc544\ub984"
        row_xml = f"""
        <hp:tr>
          <hp:tc><hp:p><hp:run><hp:t></hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>V1.3</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>2025-10-17</hp:t></hp:run></hp:p></hp:tc>
          <hp:tc><hp:p><hp:run><hp:t>\uc784\ucc44\ud604</hp:t></hp:run></hp:p></hp:tc>
        </hp:tr>
        """

        updated_row, count = update_unlabeled_metadata_row(
            row_xml,
            revision_date="2026",
            author=author,
            document_number="MFDS-PMC-08",
        )

        self.assertEqual(4, count)
        self.assertIn("MFDS-PMC-08", updated_row)
        self.assertIn("v0.1", updated_row)
        self.assertIn("2026-00-00", updated_row)
        self.assertIn(author, updated_row)

    def test_blank_revision_history_first_row_is_filled(self) -> None:
        author = "\uc1a1\uc544\ub984"
        approval = "\uc784\ucc44\ud604"
        xml = """
        <root>
          <hp:tr>
            <hp:tc><hp:p><hp:run><hp:t>\ubc84\uc804</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uac1c\uc815\uc77c\uc790</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uac1c\uc815\uc0ac\uc720</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uac1c\uc815\ub0b4\uc5ed</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uc791\uc131\uc790</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uc2b9\uc778\uc790</hp:t></hp:run></hp:p></hp:tc>
          </hp:tr>
          <hp:tr>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
          </hp:tr>
        </root>
        """

        updated_xml, count = update_revision_history_xml(xml, "2026", author, approval)

        self.assertEqual(6, count)
        self.assertIn("0.1", updated_xml)
        self.assertIn("2026-00-00", updated_xml)
        self.assertIn("\uc81c\uc815", updated_xml)
        self.assertIn("\ucd5c\ucd08 \uc791\uc131", updated_xml)
        self.assertIn(author, updated_xml)
        self.assertIn(approval, updated_xml)

    def test_revision_history_reason_and_detail_headers_allow_footnote_suffixes(self) -> None:
        author = "\uc1a1\uc544\ub984"
        approval = "\uc784\ucc44\ud604"
        xml = """
        <root>
          <hp:tr>
            <hp:tc><hp:p><hp:run><hp:t>\ubc84\uc804</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uac1c\uc815\uc77c\uc790</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uac1c\uc815\uc0ac\uc7201)</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uac1c\uc815\ub0b4\uc5ed2)</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uc791\uc131\uc790</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uc2b9\uc778\uc790</hp:t></hp:run></hp:p></hp:tc>
          </hp:tr>
          <hp:tr>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
            <hp:tc><hp:p></hp:p></hp:tc>
          </hp:tr>
        </root>
        """

        updated_xml, count = update_revision_history_xml(xml, "2026", author, approval)

        self.assertEqual(6, count)
        self.assertIn("\uc81c\uc815", updated_xml)
        self.assertIn("\ucd5c\ucd08 \uc791\uc131", updated_xml)

    def test_excel_blank_header_id_and_blank_revision_history_are_filled(self) -> None:
        author = "\uc1a1\uc544\ub984"
        approval = "\uc784\ucc44\ud604"
        old_author = "\uae40\ub300\uba85"
        with workspace_temp_dir() as tmp:
            path = tmp / "metadata.xlsx"
            workbook = Workbook()
            cover = workbook.active
            cover.title = "\ud45c\uc9c0"
            cover["B1"] = "V0.0"
            cover["C1"] = "2024.04.15"
            cover["D1"] = old_author
            history = workbook.create_sheet("\uac1c\uc815\uc774\ub825")
            headers = [
                "\ubc84\uc804",
                "\uac1c\uc815\uc77c\uc790",
                "\uac1c\uc815\uc0ac\uc720",
                "\uac1c\uc815\ub0b4\uc5ed",
                "\uc791\uc131\uc790",
                "\uc2b9\uc778\uc790",
            ]
            for col, value in enumerate(headers, start=1):
                history.cell(row=1, column=col, value=value)
                history.cell(row=2, column=col, value="")
            workbook.save(path)
            workbook.close()

            cover_count, revision_count = write_updated_excel_metadata(path, author, "2026", approval, "MFDS-PMC-08")

            updated = load_workbook(path, data_only=False)
            try:
                cover = updated["\ud45c\uc9c0"]
                history = updated["\uac1c\uc815\uc774\ub825"]
                self.assertGreaterEqual(cover_count, 4)
                self.assertGreaterEqual(revision_count, 6)
                self.assertEqual("MFDS-PMC-08", cover["A1"].value)
                self.assertEqual("v0.1", cover["B1"].value)
                self.assertEqual("2026-00-00", cover["C1"].value)
                self.assertEqual(author, cover["D1"].value)
                self.assertEqual("0.1", history["A2"].value)
                self.assertEqual("2026-00-00", history["B2"].value)
                self.assertEqual("\uc81c\uc815", history["C2"].value)
                self.assertEqual("\ucd5c\ucd08 \uc791\uc131", history["D2"].value)
                self.assertEqual(author, history["E2"].value)
                self.assertEqual(approval, history["F2"].value)
            finally:
                updated.close()

    def test_excel_self_closing_header_id_cell_stays_openable_after_update(self) -> None:
        author = "\uc1a1\uc544\ub984"
        approval = "\uc784\ucc44\ud604"
        with workspace_temp_dir() as tmp:
            path = tmp / "self-closing.xlsx"
            workbook = Workbook()
            cover = workbook.active
            cover.title = "\ud45c\uc9c0"
            cover["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
            cover["B1"] = "V1.3"
            cover["C1"] = "2025-10-17"
            cover["D1"] = "\uc784\ucc44\ud604"
            workbook.save(path)
            workbook.close()

            def make_a1_self_closing(xml: str) -> str:
                updated_xml, count = re.subn(r'<c r="A1"[^>]*/?>.*?</c>', '<c r="A1" s="1"/>', xml, count=1)
                self.assertEqual(1, count)
                return updated_xml

            rewrite_zip_text_member(path, "xl/worksheets/sheet1.xml", make_a1_self_closing)

            cover_count, _revision_count = write_updated_excel_metadata(
                path,
                author,
                "2026",
                approval,
                "MFDS-PMC-08",
            )

            self.assertGreaterEqual(cover_count, 4)
            updated = load_workbook(path, data_only=False)
            try:
                cover = updated["\ud45c\uc9c0"]
                self.assertEqual("MFDS-PMC-08", cover["A1"].value)
                self.assertEqual("v0.1", cover["B1"].value)
                self.assertEqual("2026-00-00", cover["C1"].value)
                self.assertEqual(author, cover["D1"].value)
            finally:
                updated.close()

    def test_excel_metadata_update_removes_calc_chain_parts(self) -> None:
        author = "\uc1a1\uc544\ub984"
        approval = "\uc784\ucc44\ud604"
        with workspace_temp_dir() as tmp:
            path = tmp / "calc-chain.xlsx"
            workbook = Workbook()
            cover = workbook.active
            cover.title = "\ud45c\uc9c0"
            cover["B1"] = "V1.3"
            cover["C1"] = "2025-10-17"
            cover["D1"] = "\uc784\ucc44\ud604"
            workbook.save(path)
            workbook.close()

            rewrite_zip_members(
                path,
                {
                    "[Content_Types].xml": lambda xml: xml.replace(
                        "</Types>",
                        '<Override PartName="/xl/calcChain.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/></Types>',
                    ),
                    "xl/_rels/workbook.xml.rels": lambda xml: xml.replace(
                        "</Relationships>",
                        '<Relationship Id="rIdCalcChain" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain" Target="calcChain.xml"/></Relationships>',
                    ),
                },
                {
                    "xl/calcChain.xml": (
                        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        b'<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><c r="B1" i="1"/></calcChain>'
                    ),
                },
            )

            write_updated_excel_metadata(path, author, "2026", approval, "MFDS-PMC-08")

            with zipfile.ZipFile(path, "r") as zf:
                self.assertNotIn("xl/calcChain.xml", zf.namelist())
                self.assertNotIn("calcChain", zf.read("[Content_Types].xml").decode("utf-8"))
                self.assertNotIn("calcChain", zf.read("xl/_rels/workbook.xml.rels").decode("utf-8"))
            updated = load_workbook(path, data_only=False)
            try:
                self.assertEqual("MFDS-PMC-08", updated["\ud45c\uc9c0"]["A1"].value)
            finally:
                updated.close()

    def test_excel_labeled_cover_updates_document_number_and_metadata(self) -> None:
        author = "\uc1a1\uc544\ub984"
        approval = "\uc784\ucc44\ud604"
        old_author = "\uc784\ucc44\ud604"
        with workspace_temp_dir() as tmp:
            path = tmp / "cover.xlsx"
            workbook = Workbook()
            cover = workbook.active
            cover.title = "\ud45c\uc9c0"
            cover["I17"] = "\ubb38\uc11c\ubc88\ud638"
            cover["J17"] = "MFDS-PMC-08"
            cover["I18"] = "\ubb38\uc11c\ubc84\uc804"
            cover["J18"] = "V1.3"
            cover["I19"] = "\uac1c\uc815\uc77c\uc790"
            cover["J19"] = "2025-10-17"
            cover["I20"] = "\uc791\uc131\uc790"
            cover["J20"] = old_author
            workbook.save(path)
            workbook.close()

            cover_count, revision_count = write_updated_excel_metadata(
                path,
                author,
                "2026",
                approval,
                "20",
            )

            updated = load_workbook(path, data_only=False)
            try:
                cover = updated["\ud45c\uc9c0"]
                self.assertGreaterEqual(cover_count, 4)
                self.assertEqual(0, revision_count)
                self.assertEqual("20", cover["J17"].value)
                self.assertEqual("v0.1", cover["J18"].value)
                self.assertEqual("2026-00-00", cover["J19"].value)
                self.assertEqual(author, cover["J20"].value)
            finally:
                updated.close()


if __name__ == "__main__":
    unittest.main()
