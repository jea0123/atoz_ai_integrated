from pathlib import Path
from io import BytesIO
import shutil
import unittest
from unittest.mock import patch
from uuid import uuid4
import zipfile

from openpyxl import Workbook, load_workbook

from output_file_check.models import FileIdentity, StandardOutput
import output_file_check.template_workflow as workflow


def write_workbook(path: Path, cover_value: str, body_value: str = "") -> None:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "표지"
    cover["A1"] = cover_value
    body = workbook.create_sheet("본문")
    body["A1"] = body_value
    workbook.save(path)
    workbook.close()


def workbook_bytes(cover_value: str | None, body_value: str = "") -> bytes:
    workbook = Workbook()
    first = workbook.active
    if cover_value is None:
        first.title = "본문"
        first["A1"] = body_value
    else:
        first.title = "표지"
        first["A1"] = cover_value
        first["B1"] = "V0.0"
        first["C1"] = "2025.01.01"
        first["D1"] = "Old Author"
        history = workbook.create_sheet("개정이력")
        headers = ["버전", "개정일자", "개정사유", "개정내역", "작성자", "승인자"]
        for col, value in enumerate(headers, start=1):
            history.cell(row=1, column=col, value=value)
            history.cell(row=2, column=col, value="")
        body = workbook.create_sheet("본문")
        body["A1"] = body_value
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def write_hwpx(path: Path, section_xml: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("mimetype", "application/hwp+zip")
        zf.writestr("Contents/section0.xml", section_xml)
        zf.writestr("Contents/content.hpf", "<opf:package/>")
        zf.writestr("META-INF/container.xml", "<container/>")


class TemplateWorkflowModeTest(unittest.TestCase):
    def setUp(self) -> None:
        self.root = Path(".test-artifacts") / f"template-workflow-{uuid4().hex}"
        self.root.mkdir(parents=True, exist_ok=True)
        self.addCleanup(lambda: shutil.rmtree(self.root, ignore_errors=True))

        self.template_root = self.root / "templates"
        self.development_dir = self.template_root / "development"
        self.management_dir = self.template_root / "management"
        self.cover_dir = self.template_root / "cover"
        self.development_dir.mkdir(parents=True)
        self.management_dir.mkdir(parents=True)
        self.cover_dir.mkdir(parents=True)
        write_workbook(self.development_dir / "DOC-001-InterviewPlan.xlsx", "OLD COVER", "BODY")
        write_workbook(self.management_dir / "DOC-001-InterviewPlan.xlsx", "OLD COVER", "BODY")
        write_workbook(
            self.cover_dir / "{{DOCUMENT_ID}}-{{DOCUMENT_TITLE}}_{{VERSION}}_{{REQUIREMENT_IDS}}.xlsx",
            "{{PROJECT_NAME}} {{DOCUMENT_ID}} {{DOCUMENT_TITLE}} {{VERSION}} {{REQUIREMENT_IDS}} {{REVISION_DATE}} {{AUTHOR}} {{APPROVER}}",
        )

    def run_workflow(
        self,
        *,
        apply_mode: bool,
        include_requirement_files: bool = True,
        requirement_filenames: list[str] | None = None,
        artifact_filename: str = "DOC-001-InterviewPlan.xlsx",
        artifact_payload: bytes | None = None,
        identity: FileIdentity | None = None,
        category: str = "development",
    ) -> dict[str, object]:
        temp_dir = self.root / ("temp-apply" if apply_mode else "temp-match")
        temp_dir.mkdir()
        file_items = {
            "standard_file": [("standard.pdf", b"standard")],
            "artifact_files": [
                (
                    artifact_filename,
                    artifact_payload if artifact_payload is not None else workbook_bytes("Project", "INPUT BODY"),
                )
            ],
        }
        if include_requirement_files:
            names = requirement_filenames or ["REQ_SFR-002.txt"]
            file_items["requirement_files"] = [(name, b"req") for name in names]
        fields = {
            "artifact_category": category,
            "version": "v1.0",
            "revision_date": "2026",
            "author": "Alice",
            "approver": "Bob",
            "apply_mode": "true" if apply_mode else "false",
        }
        with (
            patch.object(workflow, "TEMPLATE_ROOT", self.template_root),
            patch.object(
                workflow,
                "TEMPLATE_CATEGORY_DIRS",
                {
                    "development": self.development_dir,
                    "management": self.management_dir,
                },
            ),
            patch.object(workflow, "COVER_TEMPLATE_DIR", self.cover_dir),
            patch.object(
                workflow,
                "extract_standard_text",
                return_value="2026 Project 문서관리표준",
            ),
            patch.object(
                workflow,
                "read_standard_project_title",
                return_value="Standard Project",
            ),
            patch.object(
                workflow,
                "read_standard_outputs",
                return_value=[StandardOutput("DOC-001", "InterviewPlan")],
            ),
            patch.object(
                workflow,
                "read_file_identity",
                return_value=identity or FileIdentity(
                    project_title="Project",
                    document_title="InterviewPlan",
                    document_number="DOC-001",
                ),
            ),
        ):
            return workflow.run_template_build(
                fields,
                file_items,
                temp_dir=temp_dir,
                result_dir=self.root / "results",
            )

    def test_match_mode_does_not_create_output_files(self) -> None:
        payload = self.run_workflow(apply_mode=False)

        self.assertFalse(payload["applied"])
        self.assertEqual(1, payload["matched_count"])
        self.assertEqual(0, payload["output_count"])
        self.assertEqual([], payload["files"])
        self.assertEqual("", payload["items"][0]["output_path"])

    def test_development_match_allows_missing_requirement_files(self) -> None:
        payload = self.run_workflow(apply_mode=False, include_requirement_files=False)

        self.assertFalse(payload["applied"])
        self.assertEqual(1, payload["matched_count"])
        self.assertEqual(0, payload["requirement_count"])
        self.assertEqual([], payload["requirement_ids"])

    def test_standard_output_matching_ignores_input_document_number(self) -> None:
        output, match_type = workflow.find_standard_output(
            Path("Unrelated.xlsx"),
            "Unrelated.xlsx",
            FileIdentity(document_number="DOC-001"),
            [StandardOutput("DOC-001", "InterviewPlan")],
        )

        self.assertIsNone(output)
        self.assertEqual("", match_type)

    def test_input_without_cover_identity_matches_by_filename(self) -> None:
        payload = self.run_workflow(
            apply_mode=False,
            include_requirement_files=False,
            artifact_filename="old-number-InterviewPlan.xlsx",
            identity=FileIdentity(project_title="", document_title="", document_number="OLD-999"),
        )

        self.assertEqual(1, payload["matched_count"])
        self.assertEqual("matched", payload["items"][0]["status"])
        self.assertEqual("filename_output_name", payload["items"][0]["standard_output"]["match_type"])

    def test_apply_mode_creates_output_files(self) -> None:
        payload = self.run_workflow(apply_mode=True)

        self.assertTrue(payload["applied"])
        self.assertEqual(1, payload["matched_count"])
        self.assertEqual(1, payload["output_count"])
        self.assertEqual(["SFR-002"], payload["requirement_ids"])
        self.assertTrue(Path(payload["items"][0]["output_path"]).exists())
        self.assertTrue(any(item["kind"] == "report" for item in payload["files"]))

        output_path = Path(payload["items"][0]["output_path"])
        self.assertEqual("DOC-001-InterviewPlan_v0.1_SFR-002.xlsx", output_path.name)
        workbook = load_workbook(output_path, data_only=True)
        try:
            self.assertEqual(
                "Standard Project DOC-001 InterviewPlan v0.1 SFR-002 2026-00-00 Alice Bob",
                workbook["표지"]["A1"].value,
            )
            self.assertEqual("BODY", workbook["본문"]["A1"].value)
        finally:
            workbook.close()

    def test_apply_mode_removes_template_work_folder(self) -> None:
        payload = self.run_workflow(apply_mode=True)

        self.assertFalse((Path(payload["dump_root"]) / "_template_work").exists())

    def test_apply_mode_creates_one_output_file_per_requirement_id(self) -> None:
        payload = self.run_workflow(
            apply_mode=True,
            requirement_filenames=["REQ_SFR-001.txt", "REQ_SFR-002.txt", "REQ_SFR-003.txt"],
        )

        self.assertEqual(["SFR-001", "SFR-002", "SFR-003"], payload["requirement_ids"])
        self.assertEqual(3, payload["output_count"])
        output_files = sorted(
            (item for item in payload["files"] if item["kind"] == "template_output"),
            key=lambda item: item["name"],
        )
        self.assertEqual(
            [
                "DOC-001-InterviewPlan_v0.1_SFR-001.xlsx",
                "DOC-001-InterviewPlan_v0.1_SFR-002.xlsx",
                "DOC-001-InterviewPlan_v0.1_SFR-003.xlsx",
            ],
            [item["name"] for item in output_files],
        )

        for item in output_files:
            output_path = Path(item["path"])
            self.assertTrue(output_path.exists())
            workbook = load_workbook(output_path, data_only=True)
            try:
                self.assertIn(item["requirement_id"], workbook.worksheets[0]["A1"].value)
            finally:
                workbook.close()

    def test_unmatched_input_with_cover_is_copied_and_existing_cover_is_updated(self) -> None:
        payload = self.run_workflow(
            apply_mode=True,
            artifact_filename="Unmatched.xlsx",
            artifact_payload=workbook_bytes("Old Project", "INPUT BODY"),
            identity=FileIdentity(
                project_title="Old Project",
                document_title="UnknownDoc",
                document_number="OLD-001",
                preview_text="Old Project\nUnknownDoc",
            ),
        )

        self.assertEqual(0, payload["matched_count"])
        self.assertEqual(1, payload["output_count"])
        self.assertEqual("fallback", payload["items"][0]["status"])
        self.assertTrue(payload["items"][0]["input_has_cover"])
        self.assertEqual("input_copy", payload["items"][0]["output_source"])
        output_path = Path(payload["items"][0]["output_path"])
        self.assertEqual("Unmatched_SFR-002.xlsx", output_path.name)

        workbook = load_workbook(output_path, data_only=False)
        try:
            cover = workbook["표지"]
            history = workbook["개정이력"]
            self.assertEqual("Standard Project", cover["A1"].value)
            self.assertEqual("v0.1", cover["B1"].value)
            self.assertEqual("2026-00-00", cover["C1"].value)
            self.assertEqual("Alice", cover["D1"].value)
            self.assertEqual("Alice", history["E2"].value)
            self.assertEqual("Bob", history["F2"].value)
            self.assertEqual("INPUT BODY", workbook["본문"]["A1"].value)
        finally:
            workbook.close()

    def test_input_without_cover_is_copied_without_creating_cover_even_when_matched(self) -> None:
        payload = self.run_workflow(
            apply_mode=True,
            artifact_payload=workbook_bytes(None, "INPUT BODY"),
            identity=FileIdentity(),
        )

        self.assertEqual(1, payload["matched_count"])
        self.assertEqual(1, payload["output_count"])
        self.assertFalse(payload["items"][0]["input_has_cover"])
        self.assertEqual("input_copy", payload["items"][0]["output_source"])
        output_path = Path(payload["items"][0]["output_path"])
        self.assertEqual("DOC-001-InterviewPlan_SFR-002.xlsx", output_path.name)

        workbook = load_workbook(output_path, data_only=False)
        try:
            self.assertEqual(["본문"], workbook.sheetnames)
            self.assertEqual("INPUT BODY", workbook["본문"]["A1"].value)
        finally:
            workbook.close()

    def test_excel_cover_is_inserted_when_body_has_no_cover_sheet(self) -> None:
        target = self.root / "body-only.xlsx"
        cover_template = self.root / "cover-template.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "본문"
        sheet["A1"] = "BODY"
        workbook.save(target)
        workbook.close()
        write_workbook(cover_template, "{{DOCUMENT_ID}} {{VERSION}}")

        count = workflow.apply_excel_cover_template(
            target,
            cover_template,
            {"DOCUMENT_ID": "DOC-001", "VERSION": "v0.1"},
        )

        self.assertGreaterEqual(count, 1)
        workbook = load_workbook(target, data_only=True)
        try:
            self.assertEqual("표지", workbook.sheetnames[0])
            self.assertIn("본문", workbook.sheetnames)
            self.assertEqual("DOC-001 v0.1", workbook["표지"]["A1"].value)
            self.assertEqual("BODY", workbook["본문"]["A1"].value)
        finally:
            workbook.close()

    def test_hwpx_cover_update_keeps_document_structure_and_replaces_tokens(self) -> None:
        target = self.root / "target.hwpx"
        cover_template = self.root / "cover.hwpx"
        write_hwpx(
            target,
            '<hp:sec xmlns:hp="http://www.hancom.co.kr/hwpml/2011/paragraph">'
            "<hp:p><hp:run><hp:t>{{PROJECT_NAME}}</hp:t></hp:run></hp:p>"
            "</hp:sec>",
        )
        write_hwpx(
            cover_template,
            '<hp:sec xmlns:hp="http://www.hancom.co.kr/hwpml/2011/paragraph">'
            "<hp:p><hp:run><hp:t>WRONG COVER</hp:t></hp:run></hp:p>"
            "</hp:sec>",
        )

        count = workflow.apply_hwpx_cover_template(
            target,
            cover_template,
            {
                "PROJECT_NAME": "Standard Project",
                "DOCUMENT_ID": "",
                "REVISION_DATE": "",
                "AUTHOR": "",
                "APPROVER": "",
            },
            self.root / "work",
        )

        self.assertGreaterEqual(count, 1)
        with zipfile.ZipFile(target) as zf:
            self.assertIsNone(zf.testzip())
            section_xml = zf.read("Contents/section0.xml").decode("utf-8")
        self.assertIn("Standard Project", section_xml)
        self.assertNotIn("WRONG COVER", section_xml)


if __name__ == "__main__":
    unittest.main()
