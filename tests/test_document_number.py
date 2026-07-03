from contextlib import contextmanager
from pathlib import Path
import shutil
import unittest
from uuid import uuid4
import zipfile

from openpyxl import Workbook, load_workbook

from document_update import ppt_ooxml
from document_update.document_number import (
    replace_unlabeled_header_version_block,
    write_updated_document,
    write_updated_project_title,
)
from document_update.excel_ooxml import find_excel_cover_identity


def write_hwpx(path: Path, section_xml: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("Contents/section0.xml", section_xml)


def write_xlsx_cover(path: Path, title: str, project_title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "표지"
    sheet["A1"] = title
    sheet["A2"] = project_title
    workbook.save(path)
    workbook.close()


def write_xlsx_standard_cover(path: Path, project_title: str, title: str, attachment_title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "표지"
    sheet["A1"] = project_title
    sheet["A2"] = title
    sheet["A3"] = attachment_title
    sheet["A5"] = "문서번호"
    sheet["B5"] = "MFDS-P9-LS"
    workbook.save(path)
    workbook.close()


def add_calc_chain_parts(path: Path) -> None:
    temp_path = path.with_name(f"{path.stem}-zipedit{path.suffix}")
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            written = set()
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "[Content_Types].xml":
                    data = data.decode("utf-8").replace(
                        "</Types>",
                        '<Override PartName="/xl/calcChain.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/></Types>',
                    ).encode("utf-8")
                elif item.filename == "xl/_rels/workbook.xml.rels":
                    data = data.decode("utf-8").replace(
                        "</Relationships>",
                        '<Relationship Id="rIdCalcChain" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain" Target="calcChain.xml"/></Relationships>',
                    ).encode("utf-8")
                zout.writestr(item, data)
                written.add(item.filename)
            if "xl/calcChain.xml" not in written:
                zout.writestr(
                    "xl/calcChain.xml",
                    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    b'<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><c r="A1" i="1"/></calcChain>',
                )
    shutil.copyfile(temp_path, path)
    try:
        temp_path.unlink()
    except OSError:
        pass


def write_pptx_cover(path: Path, title: str, project_title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    slide_xml = f"""
    <p:sld xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"
           xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
      <p:cSld>
        <p:spTree>
          <p:sp>
            <p:spPr><a:xfrm><a:off x="100000" y="100000"/></a:xfrm></p:spPr>
            <p:txBody><a:p><a:r><a:rPr sz="3200"/><a:t>{title}</a:t></a:r></a:p></p:txBody>
          </p:sp>
          <p:sp>
            <p:spPr><a:xfrm><a:off x="100000" y="300000"/></a:xfrm></p:spPr>
            <p:txBody><a:p><a:r><a:rPr sz="2800"/><a:t>{project_title}</a:t></a:r></a:p></p:txBody>
          </p:sp>
        </p:spTree>
      </p:cSld>
    </p:sld>
    """
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("ppt/slides/slide1.xml", slide_xml)


@contextmanager
def workspace_temp_dir():
    path = Path.cwd() / ".test-artifacts" / uuid4().hex
    path.mkdir(parents=True, exist_ok=True)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


class DocumentNumberUpdateTest(unittest.TestCase):
    def test_hwpx_nonstandard_cover_updates_labeled_project_title(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.hwpx"
            output = tmp / "output.hwpx"
            write_hwpx(
                source,
                """
                <root>
                  <hp:p><hp:run><hp:t>사업수행계획서</hp:t></hp:run></hp:p>
                  <hp:tr>
                    <hp:tc><hp:p><hp:run><hp:t>사업명</hp:t></hp:run></hp:p></hp:tc>
                    <hp:tc><hp:p><hp:run><hp:t>기존 프로젝트</hp:t></hp:run></hp:p></hp:tc>
                  </hp:tr>
                </root>
                """,
            )

            old_document_number, _backup, project_count, document_count, output_path = write_updated_document(
                source,
                new_document_number="MFDS-PP-01",
                new_project_title="신규 프로젝트",
                output_path=output,
                allow_missing_document_number=True,
            )

            with zipfile.ZipFile(output_path, "r") as zf:
                updated_xml = zf.read("Contents/section0.xml").decode("utf-8")

            self.assertEqual("", old_document_number)
            self.assertEqual(1, project_count)
            self.assertEqual(0, document_count)
            self.assertIn("신규 프로젝트", updated_xml)
            self.assertNotIn("기존 프로젝트", updated_xml)
            self.assertIn("사업수행계획서", updated_xml)

    def test_hwpx_nonstandard_cover_does_not_replace_document_title_without_project_label(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.hwpx"
            output = tmp / "output.hwpx"
            write_hwpx(
                source,
                """
                <root>
                  <hp:p><hp:run><hp:t>사업수행계획서</hp:t></hp:run></hp:p>
                  <hp:p><hp:run><hp:t>2025년 3월 14일</hp:t></hp:run></hp:p>
                </root>
                """,
            )

            old_document_number, _backup, project_count, document_count, output_path = write_updated_document(
                source,
                new_document_number="MFDS-PP-01",
                new_project_title="신규 프로젝트",
                output_path=output,
                allow_missing_document_number=True,
            )

            with zipfile.ZipFile(output_path, "r") as zf:
                updated_xml = zf.read("Contents/section0.xml").decode("utf-8")

            self.assertEqual("", old_document_number)
            self.assertEqual(0, project_count)
            self.assertEqual(0, document_count)
            self.assertIn("사업수행계획서", updated_xml)
            self.assertNotIn("신규 프로젝트", updated_xml)

    def test_project_title_only_update_replaces_visible_project_text(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.hwpx"
            output = tmp / "output.hwpx"
            write_hwpx(
                source,
                """
                <root>
                  <hp:p><hp:run><hp:t>2025년도 수입식품통합정보시스템 고도화</hp:t></hp:run></hp:p>
                  <hp:p><hp:run><hp:t>기술적용결과표</hp:t></hp:run></hp:p>
                </root>
                """,
            )

            replace_count, output_path = write_updated_project_title(
                source,
                "2025년도 수입식품통합정보시스템 고도화",
                "2026년도 수입식품통합정보시스템 고도화",
                output_path=output,
            )

            with zipfile.ZipFile(output_path, "r") as zf:
                updated_xml = zf.read("Contents/section0.xml").decode("utf-8")

            self.assertEqual(1, replace_count)
            self.assertIn("2026년도 수입식품통합정보시스템 고도화", updated_xml)
            self.assertIn("기술적용결과표", updated_xml)
            self.assertNotIn("2025년도 수입식품통합정보시스템 고도화", updated_xml)

    def test_hwpx_standard_cover_with_attachment_keeps_document_title(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.hwpx"
            output = tmp / "output.hwpx"
            write_hwpx(
                source,
                """
                <root>
                  <hp:p><hp:run><hp:t>2025년 수입식품통합정보시스템 고도화</hp:t></hp:run></hp:p>
                  <hp:p><hp:run><hp:t>품질관리계획서</hp:t></hp:run></hp:p>
                  <hp:p><hp:run><hp:t>[별첨1]품질목표정의서</hp:t></hp:run></hp:p>
                  <hp:tr>
                    <hp:tc><hp:p><hp:run><hp:t>문서번호</hp:t></hp:run></hp:p></hp:tc>
                    <hp:tc><hp:p><hp:run><hp:t>MFDS-P9-LS</hp:t></hp:run></hp:p></hp:tc>
                  </hp:tr>
                </root>
                """,
            )

            _old_document_number, _backup, project_count, document_count, output_path = write_updated_document(
                source,
                new_document_number="MFDS-QA-01",
                new_project_title="2026년도 수입식품통합정보시스템 고도화",
                output_path=output,
            )

            with zipfile.ZipFile(output_path, "r") as zf:
                updated_xml = zf.read("Contents/section0.xml").decode("utf-8")

            self.assertEqual(1, project_count)
            self.assertEqual(1, document_count)
            self.assertIn("2026년도 수입식품통합정보시스템 고도화", updated_xml)
            self.assertNotIn("2025년 수입식품통합정보시스템 고도화", updated_xml)
            self.assertIn("품질관리계획서", updated_xml)
            self.assertIn("[별첨1]품질목표정의서", updated_xml)

    def test_hwpx_update_preserves_body_tables_after_numbered_heading(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.hwpx"
            output = tmp / "output.hwpx"
            old_project = "2025년도 수입식품통합정보시스템 고도화"
            new_project = "2026년도 수입식품통합정보시스템 고도화"
            write_hwpx(
                source,
                f"""
                <root>
                  <hp:tr>
                    <hp:tc><hp:p><hp:run><hp:t>사업명</hp:t></hp:run></hp:p></hp:tc>
                    <hp:tc><hp:p><hp:run><hp:t>{old_project}</hp:t></hp:run></hp:p></hp:tc>
                  </hp:tr>
                  <hp:tr>
                    <hp:tc><hp:p><hp:run><hp:t>문서번호</hp:t></hp:run></hp:p></hp:tc>
                    <hp:tc><hp:p><hp:run><hp:t>MFDS-OLD-01</hp:t></hp:run></hp:p></hp:tc>
                  </hp:tr>
                  <hp:p><hp:run><hp:t>1. 부적합 현황</hp:t></hp:run></hp:p>
                  <hp:tr>
                    <hp:tc><hp:p><hp:run><hp:t>부적합관리번호</hp:t></hp:run></hp:p></hp:tc>
                    <hp:tc><hp:p><hp:run><hp:t>MFDS-OLD-01</hp:t></hp:run></hp:p></hp:tc>
                    <hp:tc><hp:p><hp:run><hp:t>{old_project}</hp:t></hp:run></hp:p></hp:tc>
                  </hp:tr>
                </root>
                """,
            )

            _old_document_number, _backup, project_count, document_count, output_path = write_updated_document(
                source,
                new_document_number="MFDS-QA-01",
                old_project_title=old_project,
                new_project_title=new_project,
                output_path=output,
            )

            with zipfile.ZipFile(output_path, "r") as zf:
                updated_xml = zf.read("Contents/section0.xml").decode("utf-8")

            self.assertEqual(1, project_count)
            self.assertEqual(1, document_count)
            self.assertIn(new_project, updated_xml)
            self.assertIn("MFDS-QA-01", updated_xml)
            self.assertIn("<hp:t>1. 부적합 현황</hp:t>", updated_xml)
            self.assertIn("<hp:t>부적합관리번호</hp:t>", updated_xml)
            self.assertIn(f"<hp:t>{old_project}</hp:t>", updated_xml)
            self.assertIn("<hp:t>MFDS-OLD-01</hp:t>", updated_xml)

    def test_xlsx_unlabeled_cover_updates_project_line_not_document_title(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.xlsx"
            output = tmp / "output.xlsx"
            title = "SW 규모(기능점수) 산정 양식"
            old_project = "2025년도 수입식품통합정보시스템 고도화"
            new_project = "2026년도 수입식품통합정보시스템 고도화"
            write_xlsx_cover(source, title, old_project)

            project_title, document_title = find_excel_cover_identity(source)
            self.assertEqual(old_project, project_title)
            self.assertEqual(title, document_title)

            _old_document_number, _backup, project_count, document_count, output_path = write_updated_document(
                source,
                new_document_number="MFDS-FP-01",
                new_project_title=new_project,
                output_path=output,
                allow_missing_document_number=True,
            )

            workbook = load_workbook(output_path)
            sheet = workbook["표지"]
            self.assertEqual(title, sheet["A1"].value)
            self.assertEqual(new_project, sheet["A2"].value)
            workbook.close()
            self.assertEqual(1, project_count)
            self.assertEqual(0, document_count)

    def test_xlsx_standard_cover_with_attachment_keeps_document_title(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.xlsx"
            output = tmp / "output.xlsx"
            old_project = "2025년 수입식품통합정보시스템 고도화"
            new_project = "2026년도 수입식품통합정보시스템 고도화"
            title = "품질관리계획서"
            attachment_title = "[별첨1]품질목표정의서"
            write_xlsx_standard_cover(source, old_project, title, attachment_title)

            project_title, document_title = find_excel_cover_identity(source)
            self.assertEqual(old_project, project_title)
            self.assertEqual(title, document_title)

            _old_document_number, _backup, project_count, document_count, output_path = write_updated_document(
                source,
                new_document_number="MFDS-QA-01",
                new_project_title=new_project,
                output_path=output,
            )

            workbook = load_workbook(output_path)
            sheet = workbook["표지"]
            self.assertEqual(new_project, sheet["A1"].value)
            self.assertEqual(title, sheet["A2"].value)
            self.assertEqual(attachment_title, sheet["A3"].value)
            self.assertEqual("MFDS-QA-01", sheet["B5"].value)
            workbook.close()
            self.assertEqual(1, project_count)
            self.assertEqual(1, document_count)

    def test_xlsx_document_update_removes_calc_chain_parts(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.xlsx"
            output = tmp / "output.xlsx"
            write_xlsx_standard_cover(source, "OLD PROJECT", "TITLE", "ATTACHMENT")
            add_calc_chain_parts(source)

            _old_document_number, _backup, project_count, document_count, output_path = write_updated_document(
                source,
                new_document_number="MFDS-QA-01",
                new_project_title="NEW PROJECT",
                output_path=output,
            )

            with zipfile.ZipFile(output_path, "r") as zf:
                self.assertNotIn("xl/calcChain.xml", zf.namelist())
                self.assertNotIn("calcChain", zf.read("[Content_Types].xml").decode("utf-8"))
                self.assertNotIn("calcChain", zf.read("xl/_rels/workbook.xml.rels").decode("utf-8"))
            workbook = load_workbook(output_path)
            try:
                sheet = workbook.active
                self.assertEqual("NEW PROJECT", sheet["A1"].value)
                self.assertEqual("MFDS-QA-01", sheet["B5"].value)
            finally:
                workbook.close()
            self.assertEqual(1, project_count)
            self.assertEqual(1, document_count)

    def test_pptx_unlabeled_cover_updates_matching_project_text_only(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.pptx"
            output = tmp / "output.pptx"
            title = "SW 규모(기능점수) 산정 양식"
            old_project = "2025년도 수입식품통합정보시스템 고도화"
            new_project = "2026년도 수입식품통합정보시스템 고도화"
            write_pptx_cover(source, title, old_project)

            _old_document_number, _backup, project_count, document_count, output_path = write_updated_document(
                source,
                new_document_number="MFDS-FP-01",
                new_project_title=new_project,
                output_path=output,
                allow_missing_document_number=True,
            )

            with zipfile.ZipFile(output_path, "r") as zf:
                updated_xml = zf.read("ppt/slides/slide1.xml").decode("utf-8")

            self.assertIn(title, updated_xml)
            self.assertIn(new_project, updated_xml)
            self.assertNotIn(old_project, updated_xml)
            self.assertEqual(1, project_count)
            self.assertEqual(0, document_count)

    def test_unlabeled_header_version_with_title_cell_keeps_document_code(self) -> None:
        header_xml = """
        <hp:header>
          <hp:tr>
            <hp:tc><hp:p><hp:run><hp:t>테일러링결과서</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>LS</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>v1.0</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>2025.03.23</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>임채현</hp:t></hp:run></hp:p></hp:tc>
          </hp:tr>
        </hp:header>
        """

        updated_header, count = replace_unlabeled_header_version_block(header_xml)

        self.assertEqual(1, count)
        self.assertIn("테일러링결과서", updated_header)
        self.assertIn("LS", updated_header)
        self.assertIn("v0.1", updated_header)
        self.assertNotIn("v1.0", updated_header)

    def test_unlabeled_header_version_like_code_is_kept_by_position(self) -> None:
        header_xml = """
        <hp:header>
          <hp:tr>
            <hp:tc><hp:p><hp:run><hp:t>1.0</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>v1.0</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>2025 03 28</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>임채현</hp:t></hp:run></hp:p></hp:tc>
          </hp:tr>
        </hp:header>
        """

        updated_header, count = replace_unlabeled_header_version_block(header_xml)

        self.assertEqual(1, count)
        self.assertIn("1.0", updated_header)
        self.assertIn("v0.1", updated_header)
        self.assertNotIn("v1.0", updated_header)

    def test_unlabeled_header_blank_document_code_updates_version_by_slot(self) -> None:
        header_xml = """
        <hp:header>
          <hp:tr>
            <hp:tc><hp:p><hp:run><hp:t></hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>V0.0</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>2024.04.15</hp:t></hp:run></hp:p></hp:tc>
            <hp:tc><hp:p><hp:run><hp:t>\uae40\ub300\uba85</hp:t></hp:run></hp:p></hp:tc>
          </hp:tr>
        </hp:header>
        """

        updated_header, count = replace_unlabeled_header_version_block(header_xml)

        self.assertEqual(1, count)
        self.assertIn("v0.1", updated_header)
        self.assertNotIn("V0.0", updated_header)

    def test_ppt_unlabeled_header_blank_document_code_finds_version_slot(self) -> None:
        row_xml = """
        <a:tr>
          <a:tc><a:txBody><a:p><a:r><a:t></a:t></a:r></a:p></a:txBody></a:tc>
          <a:tc><a:txBody><a:p><a:r><a:t>V0.0</a:t></a:r></a:p></a:txBody></a:tc>
          <a:tc><a:txBody><a:p><a:r><a:t>2024.04.15</a:t></a:r></a:p></a:txBody></a:tc>
          <a:tc><a:txBody><a:p><a:r><a:t>\uae40\ub300\uba85</a:t></a:r></a:p></a:txBody></a:tc>
        </a:tr>
        """
        cells = list(ppt_ooxml.TABLE_CELL_PATTERN.finditer(row_xml))

        self.assertEqual(1, ppt_ooxml.unlabeled_header_version_cell_index(cells))

    def test_hwpx_document_number_missing_still_raises_in_strict_mode(self) -> None:
        with workspace_temp_dir() as tmp:
            source = tmp / "source.hwpx"
            output = tmp / "output.hwpx"
            write_hwpx(
                source,
                """
                <root>
                  <hp:p><hp:run><hp:t>기존 프로젝트</hp:t></hp:run></hp:p>
                  <hp:p><hp:run><hp:t>사업수행계획서</hp:t></hp:run></hp:p>
                </root>
                """,
            )

            with self.assertRaises(RuntimeError):
                write_updated_document(
                    source,
                    new_document_number="MFDS-PP-01",
                    new_project_title="신규 프로젝트",
                    output_path=output,
                )


if __name__ == "__main__":
    unittest.main()
